"""
ZYN Capital — Excel Template Filler (Motor de Crédito v1)

Preenche a planilha-template ZYN_Motor_Credito_v1.xlsx com os dados
da análise de crédito (output de analyze_credit) e parâmetros da operação.

Fluxo:
  1. Copia o template .xlsx
  2. Seleciona a aba correspondente ao tipo_operacao
  3. Preenche as células de input (azul 0000FF, Montserrat 9)
  4. Insere fórmulas Excel para campos calculados
  5. Aplica formatação condicional (OK verde / NOK vermelho)
  6. Salva no output_path

Uso:
    from modules.excel_template_filler import generate_comite_excel
    path = generate_comite_excel(analise, parametros, "output/analise.xlsx")
"""

from __future__ import annotations

import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════════════════════
# Constants
# ═══════════════════════════════════════════════════════════════════════════════

TEMPLATE_PATH = Path(__file__).parent.parent / "templates" / "ZYN_Motor_Credito_v1.xlsx"

FONT_INPUT = Font(name="Montserrat", size=9, color="0000FF")
FONT_INPUT_BOLD = Font(name="Montserrat", size=9, color="0000FF", bold=True)
FONT_OK = Font(name="Montserrat", size=9, color="2E7D4F", bold=True)
FONT_NOK = Font(name="Montserrat", size=9, color="922B21", bold=True)

FILL_OK = PatternFill(start_color="EAF4EE", end_color="EAF4EE", fill_type="solid")
FILL_NOK = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")

SHEET_MAP = {
    "CRI": "CRI",
    "CRA": "CRA",
    "Debênture": "Debênture",
    "Debenture": "Debênture",
    "DEBENTURE": "Debênture",
    "Nota Comercial": "Nota Comercial",
    "NC": "Nota Comercial",
    "CPR-F": "CPR-F",
    "CPRF": "CPR-F",
    "CPR": "CPR-F",
}

# Score mapping for risk matrix: (Probabilidade, Impacto) -> Score numérico
RISK_SCORE_MAP = {
    ("Baixo", "Baixo"): 1,
    ("Baixo", "Médio"): 2,
    ("Baixo", "Alto"): 3,
    ("Médio", "Baixo"): 2,
    ("Médio", "Médio"): 4,
    ("Médio", "Alto"): 6,
    ("Alto", "Baixo"): 3,
    ("Alto", "Médio"): 6,
    ("Alto", "Alto"): 9,
}


# ═══════════════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════════════


def _safe_get(d: dict | None, *keys: str, default: Any = None) -> Any:
    """Navega nested dict com segurança."""
    val = d
    for k in keys:
        if not isinstance(val, dict):
            return default
        val = val.get(k, default)
    return val if val is not None else default


def _parse_numeric(val: Any) -> float | int | None:
    """Converte string/percentual para número. Ex: '30%' -> 0.30, '1.5x' -> 1.5."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if not s:
        return None
    # Remove separador de milhar e normaliza decimal
    s = s.replace("R$", "").replace("mil", "").strip()
    s = s.replace(" ", "")
    # Percentual
    if s.endswith("%"):
        try:
            return float(s[:-1].replace(",", ".")) / 100.0
        except ValueError:
            return None
    # Multiplicador
    if s.lower().endswith("x"):
        try:
            return float(s[:-1].replace(",", "."))
        except ValueError:
            return None
    # Número direto
    try:
        # Padrão BR: 1.234,56
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        return float(s)
    except ValueError:
        return None


def _parse_bps(val: Any) -> float | None:
    """Extrai spread em bps de strings como 'CDI + 3.5%' ou '350'."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    # Tenta extrair número após '+'
    if "+" in s:
        parts = s.split("+")
        right = parts[-1].strip().replace("%", "").replace(",", ".")
        try:
            pct = float(right)
            # Se < 20, provavelmente é percentual -> converter para bps
            if pct < 20:
                return pct * 100
            return pct
        except ValueError:
            pass
    try:
        return float(s.replace(",", "."))
    except ValueError:
        return None


def _parse_indexador(taxa_str: Any) -> str | None:
    """Extrai indexador de string de taxa como 'CDI + 3.5%'."""
    if not taxa_str:
        return None
    s = str(taxa_str).upper()
    for idx in ["CDI", "IPCA", "IGPM", "IGP-M", "PRÉ", "PRE", "SELIC"]:
        if idx in s:
            return idx.replace("PRE", "Pré").replace("PRÉ", "Pré")
    return None


def _set_cell(ws: Worksheet, row: int, col: int, value: Any,
              font: Font = None, fill: PatternFill = None) -> None:
    """Define valor e estilo de uma célula, tratando merged cells."""
    if value is None:
        return
    # Safety: convert non-scalar types that openpyxl cannot write to a cell
    if isinstance(value, list):
        value = "; ".join(str(v) for v in value)
    elif isinstance(value, dict):
        value = str(value)
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill


def _set_input(ws: Worksheet, row: int, col: int, value: Any) -> None:
    """Define valor com formatação padrão de input (azul, Montserrat 9)."""
    _set_cell(ws, row, col, value, font=FONT_INPUT)


def _set_status(ws: Worksheet, row: int, col: int, is_ok: bool) -> None:
    """Define célula de status OK/NOK com formatação condicional."""
    if is_ok:
        _set_cell(ws, row, col, "OK", font=FONT_OK, fill=FILL_OK)
    else:
        _set_cell(ws, row, col, "NOK", font=FONT_NOK, fill=FILL_NOK)


def _set_formula(ws: Worksheet, row: int, col: int, formula: str) -> None:
    """Define fórmula Excel na célula."""
    cell = ws.cell(row=row, column=col)
    cell.value = formula
    cell.font = FONT_INPUT


def _format_volume(vol: Any) -> float | None:
    """Converte volume para R$ mil."""
    v = _parse_numeric(vol)
    if v is None:
        return None
    # Se > 100.000, provavelmente está em R$ (não R$ mil)
    if v > 100_000:
        return round(v / 1000, 1)
    return v


def _get_risk_score(prob: str, impact: str) -> int | None:
    """Calcula score do risco a partir de probabilidade e impacto."""
    return RISK_SCORE_MAP.get((prob, impact))


def _extract_covenant_status(analise: dict, covenant_key: str,
                             valor_atual: Any = None,
                             limite: Any = None,
                             is_max: bool = True) -> bool:
    """Determina se covenant está OK baseado em valor vs limite."""
    if valor_atual is None or limite is None:
        return True  # Default OK se sem dados
    va = _parse_numeric(valor_atual)
    lm = _parse_numeric(limite)
    if va is None or lm is None:
        return True
    if is_max:
        return va <= lm
    return va >= lm


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet Fillers
# ═══════════════════════════════════════════════════════════════════════════════


def _fill_cri(ws: Worksheet, analise: dict, parametros: dict) -> None:
    """Preenche aba CRI (rows 7-78)."""
    kpis = analise.get("kpis", {})
    rating = _safe_get(analise, "rating_final", "nota")
    parecer = _safe_get(analise, "rating_final", "parecer")
    justificativa = _safe_get(analise, "rating_final", "justificativa")
    recomendacoes = _safe_get(analise, "rating_final", "recomendacoes", default=[])
    operacao = analise.get("operacao", {})
    riscos = analise.get("riscos", {})
    covenants = analise.get("covenants", {})

    # ── Section 1: Ficha (rows 7-15) ──
    _set_input(ws, 7, 2, parametros.get("tomador"))
    _set_input(ws, 8, 2, parametros.get("cnpj"))
    _set_input(ws, 9, 2, parametros.get("sede"))
    _set_input(ws, 10, 2, parametros.get("setor"))
    _set_input(ws, 11, 2, rating)
    _set_input(ws, 12, 2, "ZYN Capital")
    _set_input(ws, 13, 2, datetime.now().strftime("%d/%m/%Y"))
    _set_input(ws, 14, 2, _safe_get(operacao, "securitizadora") or parametros.get("securitizadora"))
    _set_input(ws, 15, 2, _safe_get(operacao, "agente_fiduciario") or parametros.get("agente_fiduciario"))

    # ── Section 2: Estrutura (rows 19-28) — Série 1 = col B ──
    vol_mil = _format_volume(parametros.get("volume"))
    _set_input(ws, 19, 2, vol_mil)
    _set_input(ws, 20, 2, parametros.get("taxa"))
    _set_input(ws, 21, 2, _parse_indexador(parametros.get("taxa")))
    spread = _parse_bps(parametros.get("taxa"))
    _set_input(ws, 22, 2, spread)
    _set_input(ws, 23, 2, parametros.get("prazo_meses"))
    _set_input(ws, 24, 2, _safe_get(operacao, "amortizacao") or parametros.get("amortizacao"))
    _set_input(ws, 25, 2, _safe_get(operacao, "carencia") or parametros.get("carencia"))
    _set_input(ws, 26, 2, _safe_get(operacao, "subordinacao") or parametros.get("subordinacao"))
    _set_input(ws, 27, 2, rating)

    # ── Section 3: Lastro e Garantias (rows 31-38) ──
    _set_input(ws, 31, 2, _safe_get(operacao, "tipo_lastro") or parametros.get("tipo_lastro"))
    _set_input(ws, 32, 2, _safe_get(operacao, "descricao_lastro"))
    _set_input(ws, 33, 2, _format_volume(_safe_get(operacao, "valor_lastro")))
    ltv = _parse_numeric(kpis.get("ltv"))
    if ltv is not None:
        _set_input(ws, 34, 2, f"{ltv * 100:.1f}%" if ltv < 1 else f"{ltv:.1f}%")
    _set_input(ws, 35, 2, parametros.get("garantias"))
    _set_input(ws, 36, 2, _safe_get(operacao, "alienacao_fiduciaria"))
    _set_input(ws, 37, 2, _safe_get(operacao, "cessao_recebiveis"))
    _set_input(ws, 38, 2, _format_volume(_safe_get(operacao, "fundo_liquidez")))

    # ── Section 4: Financeira (rows 42-50) — col D = Ano 0 (base) ──
    _set_input(ws, 42, 4, _parse_numeric(kpis.get("receita_liquida")))
    _set_input(ws, 43, 4, _parse_numeric(kpis.get("ebitda")))
    # Margem EBITDA = formula
    _set_formula(ws, 44, 4, "=IF(D42=0,\"\",D43/D42)")
    _set_input(ws, 45, 4, _parse_numeric(kpis.get("divida_liquida")))
    # Dív.Líq/EBITDA = formula
    _set_formula(ws, 46, 4, "=IF(D43=0,\"\",D45/D43)")
    ebit_desp = _parse_numeric(kpis.get("ebit_desp_fin"))
    _set_input(ws, 47, 4, ebit_desp)
    _set_input(ws, 48, 4, _parse_numeric(kpis.get("patrimonio_liquido")))
    _set_input(ws, 49, 4, _parse_numeric(kpis.get("fco")))
    # FCO / Serviço Dívida — usar DSCR se disponível
    dscr = _parse_numeric(kpis.get("dscr"))
    if dscr is not None:
        _set_input(ws, 50, 4, dscr)

    # Preencher dados históricos se disponíveis (col B = Ano-2, C = Ano-1)
    historico = analise.get("historico_financeiro", {})
    for year_offset, col in [(-2, 2), (-1, 3)]:
        year_key = f"ano_{year_offset}" if year_offset < 0 else f"ano_mais_{year_offset}"
        year_data = historico.get(year_key, {})
        if year_data:
            _set_input(ws, 42, col, _parse_numeric(year_data.get("receita_liquida")))
            _set_input(ws, 43, col, _parse_numeric(year_data.get("ebitda")))
            col_letter = "B" if col == 2 else "C"
            _set_formula(ws, 44, col, f"=IF({col_letter}42=0,\"\",{col_letter}43/{col_letter}42)")
            _set_input(ws, 45, col, _parse_numeric(year_data.get("divida_liquida")))
            _set_formula(ws, 46, col, f"=IF({col_letter}43=0,\"\",{col_letter}45/{col_letter}43)")
            _set_input(ws, 47, col, _parse_numeric(year_data.get("ebit_desp_fin")))
            _set_input(ws, 48, col, _parse_numeric(year_data.get("patrimonio_liquido")))
            _set_input(ws, 49, col, _parse_numeric(year_data.get("fco")))

    # ── Section 5: Covenants (rows 54-60) ──
    tabela_cov = covenants.get("tabela_covenants", [])
    cov_rows = [54, 55, 56, 57, 58, 59, 60]
    # Map known covenants to KPIs for status evaluation
    cov_kpi_defaults = {
        0: ("divida_liquida_ebitda", True),   # Dív Líq / EBITDA máx
        1: ("dscr", False),                   # DSCR mínimo
        2: ("ltv", True),                     # LTV máx
        3: ("liquidez_corrente", False),      # Liquidez Corrente mín
        4: ("patrimonio_liquido", False),      # PL mín
    }

    for i, row in enumerate(cov_rows):
        if i < len(tabela_cov):
            cov = tabela_cov[i]
            limite = cov.get("limite") or cov.get("valor_limite")
            atual = cov.get("atual") or cov.get("valor_atual")
            status = cov.get("status")
            _set_input(ws, row, 2, limite)
            _set_input(ws, row, 3, atual)
            if status is not None:
                is_ok = str(status).upper() in ("OK", "CONFORME", "APROVADO", "TRUE", "SIM")
                _set_status(ws, row, 4, is_ok)
            _set_input(ws, row, 5, cov.get("observacao") or cov.get("obs"))
        else:
            # Fill from KPIs if available
            if i in cov_kpi_defaults:
                kpi_key, is_max = cov_kpi_defaults[i]
                kpi_val = _parse_numeric(kpis.get(kpi_key))
                if kpi_val is not None:
                    _set_input(ws, row, 3, kpi_val)

    # ── Section 6: Riscos (rows 64-69) ──
    matriz = riscos.get("matriz_riscos", [])
    risk_rows = [64, 65, 66, 67, 68, 69]
    for i, row in enumerate(risk_rows):
        if i < len(matriz):
            r = matriz[i]
            prob = r.get("probabilidade") or r.get("prob")
            imp = r.get("impacto")
            mit = r.get("mitigante") or r.get("mitigacao")
            _set_input(ws, row, 2, prob)
            _set_input(ws, row, 3, imp)
            score = _get_risk_score(str(prob), str(imp)) if prob and imp else None
            _set_input(ws, row, 4, score)
            _set_input(ws, row, 5, mit)

    # ── Section 7: Parecer (rows 72-78) ──
    score = _safe_get(analise, "rating_final", "score")
    _set_input(ws, 72, 2, score)

    if parecer:
        parecer_upper = str(parecer).upper()
        is_approved = parecer_upper in ("APROVADO", "APPROVED")
        font = FONT_OK if is_approved else FONT_NOK
        fill = FILL_OK if is_approved else FILL_NOK
        _set_cell(ws, 73, 2, parecer_upper, font=font, fill=fill)

    # Condições/Ressalvas
    cond_text = ""
    if justificativa:
        cond_text = justificativa
    if recomendacoes:
        if cond_text:
            cond_text += "\n"
        cond_text += "; ".join(str(r) for r in recomendacoes)
    _set_input(ws, 74, 2, cond_text or None)

    _set_input(ws, 77, 2, parametros.get("analista", "Equipe ZYN Capital"))
    _set_input(ws, 78, 2, parametros.get("socio", "Danilo Salasar"))


def _fill_cra(ws: Worksheet, analise: dict, parametros: dict) -> None:
    """Preenche aba CRA (rows 7-143)."""
    kpis = analise.get("kpis", {})
    rating = _safe_get(analise, "rating_final", "nota")
    parecer = _safe_get(analise, "rating_final", "parecer")
    operacao = analise.get("operacao", {})
    producao = analise.get("producao", {})
    riscos = analise.get("riscos", {})
    covenants = analise.get("covenants", {})
    recomendacoes = _safe_get(analise, "rating_final", "recomendacoes", default=[])

    # ── Section 1: Ficha (rows 7-16, merged B:E) ──
    _set_input(ws, 7, 2, parametros.get("tomador"))
    _set_input(ws, 8, 2, parametros.get("cnpj"))
    _set_input(ws, 9, 2, parametros.get("sede"))
    _set_input(ws, 10, 2, parametros.get("setor") or _safe_get(producao, "atividade"))
    _set_input(ws, 11, 2, _safe_get(producao, "hectares") or parametros.get("hectares"))
    _set_input(ws, 12, 2, _safe_get(producao, "certificacao") or parametros.get("certificacao"))
    _set_input(ws, 13, 2, _safe_get(operacao, "securitizadora") or parametros.get("securitizadora"))
    _set_input(ws, 14, 2, _safe_get(operacao, "agente_fiduciario") or parametros.get("agente_fiduciario"))
    _set_input(ws, 15, 2, rating)
    _set_input(ws, 16, 2, "ZYN Capital")

    # ── Section 2: Estrutura (rows 20-52, every 4 rows, cols B=Sênior) ──
    vol_mil = _format_volume(parametros.get("volume"))
    struct_rows = {
        20: vol_mil,
        24: parametros.get("taxa"),
        28: _parse_indexador(parametros.get("taxa")),
        32: _parse_bps(parametros.get("taxa")),
        36: parametros.get("prazo_meses"),
        40: _safe_get(operacao, "amortizacao") or parametros.get("amortizacao"),
        44: _safe_get(operacao, "carencia") or parametros.get("carencia"),
        48: _safe_get(operacao, "subordinacao") or parametros.get("subordinacao"),
        52: rating,
    }
    for row, val in struct_rows.items():
        _set_input(ws, row, 2, val)

    # ── Section 3: Lastro Agro (rows 58-67, merged B:E) ──
    _set_input(ws, 58, 2, _safe_get(operacao, "tipo_lastro") or parametros.get("tipo_lastro"))
    _set_input(ws, 59, 2, _safe_get(producao, "cultura") or parametros.get("cultura"))
    _set_input(ws, 60, 2, _safe_get(producao, "safra"))
    _set_input(ws, 61, 2, _safe_get(producao, "volume_sacas"))
    _set_input(ws, 62, 2, _safe_get(producao, "preco_referencia"))
    _set_input(ws, 63, 2, _format_volume(_safe_get(operacao, "valor_lastro")))
    ltv = _parse_numeric(kpis.get("ltv"))
    if ltv is not None:
        _set_input(ws, 64, 2, f"{ltv * 100:.1f}%" if ltv < 1 else f"{ltv:.1f}%")
    _set_input(ws, 65, 2, _safe_get(operacao, "armazem"))
    _set_input(ws, 66, 2, _safe_get(operacao, "registro_cetip"))
    _set_input(ws, 67, 2, _safe_get(operacao, "seguro_rural"))

    # ── Section 4: Financeira (rows 71-107, every 4 rows, cols B-E) ──
    # Col D = Safra Base
    fin_rows = {
        71: kpis.get("receita_liquida") or kpis.get("receita_bruta"),
        75: kpis.get("cpv") or kpis.get("custo_safra"),
        83: kpis.get("ebitda"),
        91: kpis.get("divida_liquida") or kpis.get("divida_total"),
        99: kpis.get("fco"),
        103: kpis.get("dscr"),
        107: _safe_get(producao, "terras_proprias"),
    }
    for row, val in fin_rows.items():
        _set_input(ws, row, 4, _parse_numeric(val))

    # Formulas for calculated fields
    _set_formula(ws, 79, 4, "=IF(D71=0,\"\",(D71-D75)/D71)")    # Margem Bruta
    _set_formula(ws, 87, 4, "=IF(D71=0,\"\",D83/D71)")           # Margem EBITDA
    _set_formula(ws, 95, 4, "=IF(D83=0,\"\",D91/D83)")           # Dív Líq / EBITDA

    # Histórico
    historico = analise.get("historico_financeiro", {})
    for year_offset, col in [(-2, 2), (-1, 3)]:
        year_key = f"ano_{year_offset}" if year_offset < 0 else f"safra_{year_offset}"
        alt_key = f"safra_{year_offset}"
        year_data = historico.get(year_key, historico.get(alt_key, {}))
        if year_data:
            col_letter = "B" if col == 2 else "C"
            _set_input(ws, 71, col, _parse_numeric(year_data.get("receita_liquida") or year_data.get("receita_bruta")))
            _set_input(ws, 75, col, _parse_numeric(year_data.get("cpv") or year_data.get("custo_safra")))
            _set_formula(ws, 79, col, f"=IF({col_letter}71=0,\"\",({col_letter}71-{col_letter}75)/{col_letter}71)")
            _set_input(ws, 83, col, _parse_numeric(year_data.get("ebitda")))
            _set_formula(ws, 87, col, f"=IF({col_letter}71=0,\"\",{col_letter}83/{col_letter}71)")
            _set_input(ws, 91, col, _parse_numeric(year_data.get("divida_liquida") or year_data.get("divida_total")))
            _set_formula(ws, 95, col, f"=IF({col_letter}83=0,\"\",{col_letter}91/{col_letter}83)")
            _set_input(ws, 99, col, _parse_numeric(year_data.get("fco")))
            _set_input(ws, 103, col, _parse_numeric(year_data.get("dscr")))

    # ── Section 5: ESG (rows 114-119) — col B = Status, col C = Observação ──
    esg = analise.get("esg", {})
    esg_items = esg.get("itens", [])
    esg_rows = [114, 115, 116, 117, 118, 119]
    for i, row in enumerate(esg_rows):
        if i < len(esg_items):
            item = esg_items[i]
            status = item.get("status", "Conforme")
            _set_input(ws, row, 2, status)
            _set_input(ws, row, 3, item.get("observacao") or item.get("obs"))

    # ── Section 6: Covenants (rows 123-128) ──
    tabela_cov = covenants.get("tabela_covenants", [])
    cov_rows = [123, 124, 125, 126, 127, 128]
    for i, row in enumerate(cov_rows):
        if i < len(tabela_cov):
            cov = tabela_cov[i]
            _set_input(ws, row, 2, cov.get("limite") or cov.get("valor_limite"))
            _set_input(ws, row, 3, cov.get("atual") or cov.get("valor_atual"))
            status = cov.get("status")
            if status is not None:
                is_ok = str(status).upper() in ("OK", "CONFORME", "APROVADO", "TRUE", "SIM")
                _set_status(ws, row, 4, is_ok)

    # ── Section 7: Riscos (rows 132-137) ──
    matriz = riscos.get("matriz_riscos", [])
    risk_rows = [132, 133, 134, 135, 136, 137]
    for i, row in enumerate(risk_rows):
        if i < len(matriz):
            r = matriz[i]
            _set_input(ws, row, 2, r.get("probabilidade") or r.get("prob"))
            _set_input(ws, row, 3, r.get("impacto"))
            _set_input(ws, row, 4, r.get("mitigante") or r.get("mitigacao"))

    # ── Section 8: Parecer (rows 140-143) ──
    _set_input(ws, 140, 2, _safe_get(analise, "rating_final", "score"))
    if parecer:
        parecer_upper = str(parecer).upper()
        is_approved = parecer_upper in ("APROVADO", "APPROVED")
        _set_cell(ws, 141, 2, parecer_upper,
                  font=FONT_OK if is_approved else FONT_NOK,
                  fill=FILL_OK if is_approved else FILL_NOK)
    _set_input(ws, 142, 2, parametros.get("analista", "Equipe ZYN Capital"))
    _set_input(ws, 143, 2, parametros.get("socio", "Danilo Salasar"))


def _fill_debenture(ws: Worksheet, analise: dict, parametros: dict) -> None:
    """Preenche aba Debênture (rows 7-182)."""
    kpis = analise.get("kpis", {})
    rating = _safe_get(analise, "rating_final", "nota")
    parecer = _safe_get(analise, "rating_final", "parecer")
    operacao = analise.get("operacao", {})
    capital = analise.get("capital", {})
    riscos = analise.get("riscos", {})
    covenants = analise.get("covenants", {})

    # ── Section 1: Ficha Emissor (rows 7-16, merged B:E) ──
    _set_input(ws, 7, 2, parametros.get("tomador"))
    _set_input(ws, 8, 2, parametros.get("cnpj"))
    _set_input(ws, 9, 2, parametros.get("sede"))
    _set_input(ws, 10, 2, parametros.get("setor"))
    _set_input(ws, 11, 2, _parse_numeric(kpis.get("receita_liquida")))
    _set_input(ws, 12, 2, _safe_get(operacao, "acoes_listadas") or parametros.get("acoes_listadas"))
    _set_input(ws, 13, 2, _safe_get(operacao, "codigo_b3") or parametros.get("codigo_b3"))
    _set_input(ws, 14, 2, _safe_get(operacao, "rating_externo") or parametros.get("rating_externo"))
    _set_input(ws, 15, 2, rating)
    _set_input(ws, 16, 2, "ZYN Capital")

    # ── Section 2: Emissão (rows 20-76, every 4, col B = 1ª Série) ──
    _set_input(ws, 20, 2, _safe_get(operacao, "numero_emissao") or parametros.get("numero_emissao"))
    _set_input(ws, 24, 2, _format_volume(parametros.get("volume")))
    _set_input(ws, 28, 2, _safe_get(operacao, "tipo_debenture") or "Simples")
    _set_input(ws, 32, 2, _safe_get(operacao, "especie") or parametros.get("especie"))
    _set_input(ws, 36, 2, parametros.get("taxa"))
    _set_input(ws, 40, 2, _parse_indexador(parametros.get("taxa")))
    _set_input(ws, 44, 2, _parse_bps(parametros.get("taxa")))
    _set_input(ws, 48, 2, parametros.get("prazo_meses"))
    _set_input(ws, 52, 2, _safe_get(operacao, "data_emissao") or parametros.get("data_emissao"))
    _set_input(ws, 56, 2, _safe_get(operacao, "data_vencimento") or parametros.get("data_vencimento"))
    _set_input(ws, 60, 2, _safe_get(operacao, "amortizacao") or parametros.get("amortizacao"))
    _set_input(ws, 64, 2, _safe_get(operacao, "remuneracao") or parametros.get("remuneracao"))
    _set_input(ws, 68, 2, _safe_get(operacao, "isin"))
    _set_input(ws, 72, 2, _safe_get(operacao, "escriturador") or parametros.get("escriturador"))
    _set_input(ws, 76, 2, _safe_get(operacao, "agente_fiduciario") or parametros.get("agente_fiduciario"))

    # ── Section 3: Destinação (rows 82-87, merged B:E) ──
    dest = operacao.get("destinacao", {})
    _set_input(ws, 82, 2, dest.get("finalidade") or _safe_get(operacao, "finalidade"))
    _set_input(ws, 83, 2, _format_volume(dest.get("capex")))
    _set_input(ws, 84, 2, _format_volume(dest.get("capital_giro")))
    _set_input(ws, 85, 2, _format_volume(dest.get("refinanciamento")))
    _set_input(ws, 86, 2, _format_volume(dest.get("outros")))
    # Total = formula
    _set_formula(ws, 87, 2, "=SUM(B83:B86)")

    # ── Section 4: Financeira (rows 91-139, every 4, cols B-E) ──
    # Col D = Ano 0
    fin_map = {
        91: "receita_liquida",
        95: "ebitda",
        103: "ebit",
        107: "lucro_liquido",
        111: "divida_bruta",
        115: "divida_liquida",
        127: "patrimonio_liquido",
        131: "fco",
        135: "dscr",
        139: "capex",
    }
    for row, key in fin_map.items():
        val = _parse_numeric(kpis.get(key))
        if val is not None:
            _set_input(ws, row, 4, val)

    # Formulas
    _set_formula(ws, 99, 4, "=IF(D91=0,\"\",D95/D91)")     # Margem EBITDA
    _set_formula(ws, 119, 4, "=IF(D95=0,\"\",D115/D95)")   # Dív Líq / EBITDA
    _set_formula(ws, 123, 4, "=IF(D95=0,\"\",D103/D95)")   # EBIT / Desp Fin (proxy)

    # Histórico
    historico = analise.get("historico_financeiro", {})
    for year_offset, col in [(-2, 2), (-1, 3)]:
        year_key = f"ano_{year_offset}"
        year_data = historico.get(year_key, {})
        if year_data:
            cl = "B" if col == 2 else "C"
            for row, key in fin_map.items():
                val = _parse_numeric(year_data.get(key))
                if val is not None:
                    _set_input(ws, row, col, val)
            _set_formula(ws, 99, col, f"=IF({cl}91=0,\"\",{cl}95/{cl}91)")
            _set_formula(ws, 119, col, f"=IF({cl}95=0,\"\",{cl}115/{cl}95)")
            _set_formula(ws, 123, col, f"=IF({cl}95=0,\"\",{cl}103/{cl}95)")

    # ── Section 5: Garantias/Covenants (rows 146-153) — col B = Descrição, col C = Status ──
    tabela_cov = covenants.get("tabela_covenants", [])
    # Rows 146-148 = guarantees, 149-153 = covenants
    garantias_list = operacao.get("garantias_detalhadas", [])
    gar_rows = [146, 147, 148]
    for i, row in enumerate(gar_rows):
        if i < len(garantias_list):
            g = garantias_list[i]
            _set_input(ws, row, 2, g.get("descricao") or str(g))
            status = g.get("status")
            if status:
                is_ok = str(status).upper() in ("OK", "CONFORME", "TRUE", "SIM")
                _set_status(ws, row, 3, is_ok)
        elif i == 0 and parametros.get("garantias"):
            _set_input(ws, row, 2, parametros.get("garantias"))
            _set_status(ws, row, 3, True)

    cov_rows = [149, 150, 151, 152, 153]
    for i, row in enumerate(cov_rows):
        if i < len(tabela_cov):
            cov = tabela_cov[i]
            desc = cov.get("limite") or cov.get("valor_limite") or cov.get("descricao")
            _set_input(ws, row, 2, desc)
            status = cov.get("status")
            if status:
                is_ok = str(status).upper() in ("OK", "CONFORME", "APROVADO", "TRUE", "SIM")
                _set_status(ws, row, 3, is_ok)

    # ── Section 6: Peers (rows 157-177) — fill from analise if available ──
    peers = analise.get("peers", [])
    for i, peer in enumerate(peers[:21]):  # max 21 rows
        row = 157 + i
        _set_input(ws, row, 1, peer.get("empresa"))
        _set_input(ws, row, 2, peer.get("setor"))
        _set_input(ws, row, 3, peer.get("rating"))
        _set_input(ws, row, 4, peer.get("spread"))
        _set_input(ws, row, 5, peer.get("observacao") or peer.get("obs"))

    # ── Section 7: Parecer (rows 179-182) ──
    _set_input(ws, 179, 2, _safe_get(analise, "rating_final", "score"))
    if parecer:
        parecer_upper = str(parecer).upper()
        is_approved = parecer_upper in ("APROVADO", "APPROVED")
        _set_cell(ws, 180, 2, parecer_upper,
                  font=FONT_OK if is_approved else FONT_NOK,
                  fill=FILL_OK if is_approved else FILL_NOK)
    _set_input(ws, 181, 2, parametros.get("analista", "Equipe ZYN Capital"))
    _set_input(ws, 182, 2, parametros.get("socio", "Danilo Salasar"))


def _fill_nota_comercial(ws: Worksheet, analise: dict, parametros: dict) -> None:
    """Preenche aba Nota Comercial (rows 7-120)."""
    kpis = analise.get("kpis", {})
    rating = _safe_get(analise, "rating_final", "nota")
    parecer = _safe_get(analise, "rating_final", "parecer")
    operacao = analise.get("operacao", {})
    covenants = analise.get("covenants", {})

    # ── Section 1: Ficha (rows 7-14, merged B:E) ──
    _set_input(ws, 7, 2, parametros.get("tomador"))
    _set_input(ws, 8, 2, parametros.get("cnpj"))
    _set_input(ws, 9, 2, parametros.get("sede"))
    _set_input(ws, 10, 2, parametros.get("setor"))
    _set_input(ws, 11, 2, _safe_get(operacao, "cagr_receita") or parametros.get("cagr_receita"))
    _set_input(ws, 12, 2, _parse_numeric(kpis.get("receita_liquida")) or _format_volume(parametros.get("faturamento")))
    _set_input(ws, 13, 2, rating)
    _set_input(ws, 14, 2, "ZYN Capital")

    # ── Section 2: Características (rows 18-54, every 4, col B = Série 1) ──
    vol_mil = _format_volume(parametros.get("volume"))
    char_rows = {
        18: vol_mil,
        22: parametros.get("taxa"),
        26: _parse_indexador(parametros.get("taxa")),
        30: _parse_bps(parametros.get("taxa")),
        34: parametros.get("prazo_meses"),
        38: _safe_get(operacao, "vencimento") or parametros.get("data_vencimento"),
        42: _safe_get(operacao, "amortizacao") or parametros.get("amortizacao"),
        46: _safe_get(operacao, "forma_pagamento_juros") or parametros.get("forma_juros"),
        50: _safe_get(operacao, "isin"),
        54: _safe_get(operacao, "escriturador") or parametros.get("escriturador"),
    }
    for row, val in char_rows.items():
        _set_input(ws, row, 2, val)

    # ── Section 3: Crédito Short-term (rows 61-69) — col B=Atual, C=Limite, D=Status ──
    credit_map = {
        61: ("receita_liquida", None),
        62: ("ebitda", None),
        63: ("margem_ebitda", None),
        64: ("divida_liquida", None),
        65: ("divida_liquida_ebitda", None),
        66: ("liquidez_corrente", None),
        67: ("liquidez_imediata", None),
        68: ("prazo_medio_recebimento", None),
        69: ("capital_giro", None),
    }
    tabela_cov = covenants.get("tabela_covenants", [])
    for row, (key, _) in credit_map.items():
        val = _parse_numeric(kpis.get(key))
        if val is not None:
            _set_input(ws, row, 2, val)

    # Fill covenant limits if available
    cov_names_nc = [
        "receita", "ebitda", "margem", "divida", "div_ebitda",
        "liquidez_corrente", "liquidez_imediata", "pmr", "capital_giro",
    ]
    for i, cov in enumerate(tabela_cov):
        if i >= 9:
            break
        row = 61 + i
        limite = cov.get("limite") or cov.get("valor_limite")
        _set_input(ws, row, 3, limite)
        status = cov.get("status")
        if status:
            is_ok = str(status).upper() in ("OK", "CONFORME", "APROVADO", "TRUE", "SIM")
            _set_status(ws, row, 4, is_ok)

    # ── Section 4: Garantias (rows 72-75, merged B:E) ──
    _set_input(ws, 72, 2, parametros.get("garantias") or _safe_get(operacao, "tipo_garantia"))
    _set_input(ws, 73, 2, _format_volume(_safe_get(operacao, "valor_garantia")))
    ltv = _parse_numeric(kpis.get("ltv"))
    if ltv is not None:
        _set_input(ws, 74, 2, f"{ltv * 100:.1f}%" if ltv < 1 else f"{ltv:.1f}%")
    _set_input(ws, 75, 2, _safe_get(operacao, "documentacao_garantia"))

    # ── Section 5: Fluxo de Caixa (rows 79-112, every 3 rows) ──
    fluxo = operacao.get("fluxo_caixa", [])
    fluxo_rows = [79, 82, 85, 88, 91, 94, 97, 100, 103, 106, 109, 112]
    for i, row in enumerate(fluxo_rows):
        if i < len(fluxo):
            f = fluxo[i]
            _set_input(ws, row, 2, f.get("entrada"))
            _set_input(ws, row, 3, f.get("saida"))
            # Saldo = formula
            if i == 0:
                _set_formula(ws, row, 4, f"=B{row}-C{row}")
            else:
                prev_row = fluxo_rows[i - 1]
                _set_formula(ws, row, 4, f"=D{prev_row}+B{row}-C{row}")

    # ── Section 6: Parecer (rows 117-120) ──
    _set_input(ws, 117, 2, _safe_get(analise, "rating_final", "score"))
    if parecer:
        parecer_upper = str(parecer).upper()
        is_approved = parecer_upper in ("APROVADO", "APPROVED")
        _set_cell(ws, 118, 2, parecer_upper,
                  font=FONT_OK if is_approved else FONT_NOK,
                  fill=FILL_OK if is_approved else FILL_NOK)
    _set_input(ws, 119, 2, parametros.get("analista", "Equipe ZYN Capital"))
    _set_input(ws, 120, 2, parametros.get("socio", "Danilo Salasar"))


def _fill_cprf(ws: Worksheet, analise: dict, parametros: dict) -> None:
    """Preenche aba CPR-F (rows 7-124)."""
    kpis = analise.get("kpis", {})
    rating = _safe_get(analise, "rating_final", "nota")
    parecer = _safe_get(analise, "rating_final", "parecer")
    operacao = analise.get("operacao", {})
    producao = analise.get("producao", {})
    riscos = analise.get("riscos", {})
    dados_fazenda = analise.get("dados_fazenda", {})
    onus = analise.get("onus", {})

    # ── Section 1: Ficha Emitente (rows 7-16, merged B:E) ──
    _set_input(ws, 7, 2, parametros.get("tomador"))
    _set_input(ws, 8, 2, parametros.get("cnpj") or parametros.get("cpf"))
    _set_input(ws, 9, 2, parametros.get("sede"))
    _set_input(ws, 10, 2, parametros.get("tipo_pessoa") or _safe_get(operacao, "tipo_pessoa"))
    _set_input(ws, 11, 2, _parse_numeric(dados_fazenda.get("area_total") or _safe_get(producao, "area_total")))
    _set_input(ws, 12, 2, _parse_numeric(dados_fazenda.get("area_propria") or _safe_get(producao, "area_propria")))
    _set_input(ws, 13, 2, _parse_numeric(dados_fazenda.get("area_arrendada") or _safe_get(producao, "area_arrendada")))
    _set_input(ws, 14, 2, dados_fazenda.get("culturas") or _safe_get(producao, "cultura") or parametros.get("cultura"))
    _set_input(ws, 15, 2, dados_fazenda.get("experiencia_anos") or _safe_get(producao, "experiencia"))
    _set_input(ws, 16, 2, rating)

    # ── Section 2: Características CPR-F (rows 19-28, merged B:E) ──
    _set_input(ws, 19, 2, _format_volume(parametros.get("volume")))
    _set_input(ws, 20, 2, parametros.get("taxa"))
    _set_input(ws, 21, 2, _parse_indexador(parametros.get("taxa")))
    _set_input(ws, 22, 2, _parse_bps(parametros.get("taxa")))
    _set_input(ws, 23, 2, parametros.get("prazo_meses"))
    _set_input(ws, 24, 2, _safe_get(operacao, "data_emissao") or parametros.get("data_emissao"))
    _set_input(ws, 25, 2, _safe_get(operacao, "data_vencimento") or parametros.get("data_vencimento"))
    _set_input(ws, 26, 2, _safe_get(operacao, "amortizacao") or parametros.get("amortizacao", "Bullet"))
    _set_input(ws, 27, 2, _safe_get(operacao, "credor") or parametros.get("credor"))
    _set_input(ws, 28, 2, _safe_get(operacao, "cartorio") or parametros.get("cartorio"))

    # ── Section 3: Agronômica (rows 32-64, every 4, cols B-E) ──
    # Col D = Safra Base
    agro_map = {
        32: dados_fazenda.get("area_plantada") or _safe_get(producao, "area_plantada"),
        36: dados_fazenda.get("produtividade") or _safe_get(producao, "produtividade"),
        40: dados_fazenda.get("producao_total") or _safe_get(producao, "producao_total"),
        44: dados_fazenda.get("preco_medio") or _safe_get(producao, "preco_medio"),
        48: kpis.get("receita_liquida") or kpis.get("receita_bruta") or dados_fazenda.get("receita_bruta"),
        52: dados_fazenda.get("custo_producao_ha") or _safe_get(producao, "custo_ha"),
        56: dados_fazenda.get("custo_total") or _safe_get(producao, "custo_total"),
        64: kpis.get("divida_liquida") or dados_fazenda.get("divida_agro"),
    }
    for row, val in agro_map.items():
        _set_input(ws, row, 4, _parse_numeric(val))

    # Formulas
    # Produção Total = Área × Produtividade
    _set_formula(ws, 40, 4, "=IF(D32=0,\"\",D32*D36)")
    # Receita Bruta = Produção × Preço
    _set_formula(ws, 48, 4, "=IF(D40=0,\"\",D40*D44/1000)")
    # Custo Total = Custo/ha × Área
    _set_formula(ws, 56, 4, "=IF(D32=0,\"\",D52*D32/1000)")
    # Margem Operacional
    _set_formula(ws, 60, 4, "=IF(D48=0,\"\",(D48-D56)/D48)")

    # Histórico
    historico = analise.get("historico_financeiro", {})
    for year_offset, col in [(-2, 2), (-1, 3)]:
        year_key = f"safra_{year_offset}" if year_offset < 0 else f"ano_{year_offset}"
        alt_key = f"ano_{year_offset}"
        year_data = historico.get(year_key, historico.get(alt_key, {}))
        if year_data:
            cl = "B" if col == 2 else "C"
            _set_input(ws, 32, col, _parse_numeric(year_data.get("area_plantada")))
            _set_input(ws, 36, col, _parse_numeric(year_data.get("produtividade")))
            _set_formula(ws, 40, col, f"=IF({cl}32=0,\"\",{cl}32*{cl}36)")
            _set_input(ws, 44, col, _parse_numeric(year_data.get("preco_medio")))
            _set_formula(ws, 48, col, f"=IF({cl}40=0,\"\",{cl}40*{cl}44/1000)")
            _set_input(ws, 52, col, _parse_numeric(year_data.get("custo_producao_ha") or year_data.get("custo_ha")))
            _set_formula(ws, 56, col, f"=IF({cl}32=0,\"\",{cl}52*{cl}32/1000)")
            _set_formula(ws, 60, col, f"=IF({cl}48=0,\"\",({cl}48-{cl}56)/{cl}48)")
            _set_input(ws, 64, col, _parse_numeric(year_data.get("divida_agro") or year_data.get("divida_liquida")))

    # ── Section 4: Garantias (rows 71-86, every 3, cols B-D) ──
    garantias = operacao.get("garantias_detalhadas", [])
    gar_rows = [71, 74, 77, 80, 83, 86]
    for i, row in enumerate(gar_rows):
        if i < len(garantias):
            g = garantias[i]
            _set_input(ws, row, 2, g.get("descricao") or g.get("matricula") or str(g))
            _set_input(ws, row, 3, _format_volume(g.get("valor")))
            ltv_g = _parse_numeric(g.get("ltv"))
            if ltv_g is not None:
                _set_input(ws, row, 4, f"{ltv_g * 100:.1f}%" if ltv_g < 1 else f"{ltv_g:.1f}%")
        elif i == 0 and parametros.get("garantias"):
            _set_input(ws, row, 2, parametros.get("garantias"))

    # ── Section 5: Compliance (rows 92-99) — col B=Status, C=Observação ──
    compliance = onus.get("itens", []) or analise.get("compliance", {}).get("itens", [])
    comp_rows = [92, 93, 94, 95, 96, 97, 98, 99]
    for i, row in enumerate(comp_rows):
        if i < len(compliance):
            c = compliance[i]
            _set_input(ws, row, 2, c.get("status", "Regular"))
            _set_input(ws, row, 3, c.get("observacao") or c.get("obs"))

    # ── Section 6: Sensibilidade (rows 103-115, every 4, cols B-E) ──
    sensibilidade = analise.get("sensibilidade", {})
    cenarios = sensibilidade.get("cenarios", [])
    sens_rows = [103, 107, 111, 115]
    sens_labels = ["base", "bull", "bear", "estresse"]
    for i, row in enumerate(sens_rows):
        if i < len(cenarios):
            c = cenarios[i]
            _set_input(ws, row, 2, _parse_numeric(c.get("produtividade")))
            _set_input(ws, row, 3, _parse_numeric(c.get("preco")))
            _set_input(ws, row, 4, _parse_numeric(c.get("receita")) or _format_volume(c.get("receita")))
            cobre = c.get("cobre_servico")
            if cobre is not None:
                _set_input(ws, row, 5, "S" if cobre else "N")
        elif i == 0:
            # Base scenario from agro data (D columns)
            _set_formula(ws, row, 2, "=D36")
            _set_formula(ws, row, 3, "=D44")
            _set_formula(ws, row, 4, "=D48")

    # Bull = +15%
    if len(cenarios) < 2:
        _set_formula(ws, 107, 2, "=B103*1.15")
        _set_formula(ws, 107, 3, "=C103*1.15")
        _set_formula(ws, 107, 4, "=B107*C107*D32/1000")
    # Bear = -15%
    if len(cenarios) < 3:
        _set_formula(ws, 111, 2, "=B103*0.85")
        _set_formula(ws, 111, 3, "=C103*0.85")
        _set_formula(ws, 111, 4, "=B111*C111*D32/1000")
    # Estresse = -30%
    if len(cenarios) < 4:
        _set_formula(ws, 115, 2, "=B103*0.70")
        _set_formula(ws, 115, 3, "=C103*0.70")
        _set_formula(ws, 115, 4, "=B115*C115*D32/1000")

    # ── Section 7: Parecer (rows 121-124) ──
    _set_input(ws, 121, 2, _safe_get(analise, "rating_final", "score"))
    if parecer:
        parecer_upper = str(parecer).upper()
        is_approved = parecer_upper in ("APROVADO", "APPROVED")
        _set_cell(ws, 122, 2, parecer_upper,
                  font=FONT_OK if is_approved else FONT_NOK,
                  fill=FILL_OK if is_approved else FILL_NOK)
    _set_input(ws, 123, 2, parametros.get("analista", "Equipe ZYN Capital"))
    _set_input(ws, 124, 2, parametros.get("socio", "Danilo Salasar"))


# ═══════════════════════════════════════════════════════════════════════════════
# Main Entry Point
# ═══════════════════════════════════════════════════════════════════════════════

FILLER_MAP = {
    "CRI": _fill_cri,
    "CRA": _fill_cra,
    "Debênture": _fill_debenture,
    "Nota Comercial": _fill_nota_comercial,
    "CPR-F": _fill_cprf,
}


def generate_comite_excel(
    analise: dict,
    parametros: dict,
    output_path: str,
) -> str:
    """
    Gera planilha de comitê preenchida a partir do template ZYN Motor de Crédito.

    Args:
        analise: Dict com resultado de analyze_credit() — rating, kpis, seções, riscos, etc.
        parametros: Dict com parâmetros da operação — tomador, cnpj, tipo_operacao, volume, etc.
        output_path: Caminho do arquivo .xlsx de saída.

    Returns:
        Caminho absoluto do arquivo gerado.

    Raises:
        FileNotFoundError: Se o template não for encontrado.
        ValueError: Se tipo_operacao não for reconhecido.
    """
    # Validar template
    template = Path(TEMPLATE_PATH)
    if not template.exists():
        raise FileNotFoundError(
            f"Template não encontrado: {template}. "
            "Verifique se ZYN_Motor_Credito_v1.xlsx está em templates/."
        )

    # Resolver tipo de operação
    tipo_raw = parametros.get("tipo_operacao", "")
    sheet_name = SHEET_MAP.get(tipo_raw) or SHEET_MAP.get(tipo_raw.upper())
    if not sheet_name:
        # Fuzzy match
        tipo_upper = tipo_raw.upper().replace("-", "").replace(" ", "")
        for key, val in SHEET_MAP.items():
            if key.upper().replace("-", "").replace(" ", "") == tipo_upper:
                sheet_name = val
                break
    if not sheet_name:
        valid = sorted(set(SHEET_MAP.values()))
        raise ValueError(
            f"tipo_operacao '{tipo_raw}' não reconhecido. "
            f"Valores válidos: {', '.join(valid)}"
        )

    # Copiar template
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, out)

    # Abrir e preencher
    wb = load_workbook(out)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba '{sheet_name}' não encontrada no template.")

    ws = wb[sheet_name]

    filler = FILLER_MAP.get(sheet_name)
    if filler is None:
        raise ValueError(f"Filler não implementado para aba '{sheet_name}'.")

    try:
        filler(ws, analise or {}, parametros or {})
    except Exception as e:
        logger.error("Erro ao preencher aba %s: %s", sheet_name, e, exc_info=True)
        raise RuntimeError(f"Erro ao preencher aba {sheet_name}: {e}") from e

    # Salvar
    wb.save(out)
    wb.close()

    abs_path = str(out.resolve())
    logger.info("Planilha de comitê gerada: %s (aba: %s)", abs_path, sheet_name)
    return abs_path
