"""
ZYN Capital — Gerador de Excel para Consulta Agro
Gera relatorio completo em .xlsx a partir de dados de consulta ambiental.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# ZYN Visual Constants
# ---------------------------------------------------------------------------
NAVY = "223040"
GRAY = "8B9197"
GREEN = "2E7D4F"
YELLOW = "EAB308"
RED = "DC2626"
WHITE = "FFFFFF"
LIGHT_GREEN = "E8F5E9"
LIGHT_YELLOW = "FFF9C4"
LIGHT_RED = "FFEBEE"
LIGHT_GRAY = "F5F5F5"

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, color=NAVY, size=14)
SUBTITLE_FONT = Font(name="Calibri", bold=True, color=NAVY, size=11)
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color=GRAY),
    right=Side(style="thin", color=GRAY),
    top=Side(style="thin", color=GRAY),
    bottom=Side(style="thin", color=GRAY),
)

SCORE_FILLS = {
    "Verde": PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid"),
    "Amarelo": PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid"),
    "Vermelho": PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid"),
}
SCORE_FONTS = {
    "Verde": Font(name="Calibri", bold=True, color=GREEN, size=12),
    "Amarelo": Font(name="Calibri", bold=True, color=YELLOW, size=12),
    "Vermelho": Font(name="Calibri", bold=True, color=RED, size=12),
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _auto_width(ws, min_width=10, max_width=50):
    """Ajusta largura das colunas automaticamente."""
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = length
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _write_header_row(ws, row, headers):
    """Escreve linha de cabecalho com estilo ZYN."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _write_row(ws, row, values, font=None, fill=None):
    """Escreve linha de dados."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = font or NORMAL_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def _bool_to_text(val):
    return "Sim" if val else "Nao"


def _safe_get(d, *keys, default="N/D"):
    """Navega dict aninhado com seguranca."""
    current = d
    for k in keys:
        if isinstance(current, dict):
            current = current.get(k, default)
        else:
            return default
    return current if current is not None else default


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------
def _build_resumo(wb, consulta: dict):
    """Sheet 1: Resumo geral da consulta."""
    ws = wb.active
    ws.title = "Resumo"

    # Title row
    ws.merge_cells("A1:F1")
    cell = ws.cell(row=1, column=1, value="ZYN Capital -- Consulta Agro")
    cell.font = Font(name="Calibri", bold=True, color=WHITE, size=16)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Subtitle
    ws.merge_cells("A2:F2")
    cell = ws.cell(row=2, column=1, value="Relatorio Ambiental Automatizado")
    cell.font = Font(name="Calibri", italic=True, color=GRAY, size=10)
    cell.alignment = Alignment(horizontal="center")

    # Metadata
    row = 4
    meta = [
        ("Data da Consulta", consulta.get("timestamp", "N/D")),
        ("Busca", consulta.get("busca", "N/D")),
        ("Tipo", consulta.get("tipo", "N/D")),
    ]
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2, value=value).font = NORMAL_FONT
        row += 1

    # Score ambiental
    row += 1
    resultado = consulta.get("resultado", {})
    score = resultado.get("score_ambiental_grupo", "N/D")

    ws.cell(row=row, column=1, value="Score Ambiental do Grupo").font = SUBTITLE_FONT
    score_cell = ws.cell(row=row, column=2, value=score)
    score_cell.font = SCORE_FONTS.get(score, BOLD_FONT)
    if score in SCORE_FILLS:
        score_cell.fill = SCORE_FILLS[score]
    row += 1

    # Summary metrics
    row += 1
    ws.cell(row=row, column=1, value="Total de Propriedades").font = BOLD_FONT
    ws.cell(row=row, column=2, value=resultado.get("total_propriedades", 0)).font = NORMAL_FONT
    row += 1

    ws.cell(row=row, column=1, value="Area Total (ha)").font = BOLD_FONT
    area = resultado.get("area_total_ha", 0)
    ws.cell(row=row, column=2, value=f"{area:,.1f}".replace(",", ".")).font = NORMAL_FONT
    row += 1

    # Area breakdown (sum across all properties)
    propriedades = resultado.get("propriedades", [])
    area_consolidada = sum(p.get("areas", {}).get("area_consolidada_ha", 0) or 0 for p in propriedades)
    area_agricultavel = sum(p.get("areas", {}).get("area_agricultavel_ha", 0) or 0 for p in propriedades)
    area_pastagem = sum(p.get("areas", {}).get("area_pastagem_ha", 0) or 0 for p in propriedades)
    area_solo = sum(p.get("areas", {}).get("area_solo_exposto_ha", 0) or 0 for p in propriedades)

    if area_consolidada > 0 or area_agricultavel > 0:
        ws.cell(row=row, column=1, value="Area Consolidada (ha)").font = BOLD_FONT
        ws.cell(row=row, column=2, value=f"{area_consolidada:,.1f}".replace(",", ".")).font = NORMAL_FONT
        row += 1

        ws.cell(row=row, column=1, value="Area Agricultavel (ha)").font = BOLD_FONT
        ws.cell(row=row, column=2, value=f"{area_agricultavel:,.1f}".replace(",", ".")).font = NORMAL_FONT
        row += 1

        ws.cell(row=row, column=1, value="Area Pastagem (ha)").font = BOLD_FONT
        ws.cell(row=row, column=2, value=f"{area_pastagem:,.1f}".replace(",", ".")).font = NORMAL_FONT
        row += 1

        ws.cell(row=row, column=1, value="Solo Exposto (ha)").font = BOLD_FONT
        ws.cell(row=row, column=2, value=f"{area_solo:,.1f}".replace(",", ".")).font = NORMAL_FONT
        row += 1

    # Consolidated alerts
    alertas = resultado.get("alertas_consolidados", [])
    if alertas:
        row += 1
        ws.cell(row=row, column=1, value="Alertas Consolidados").font = SUBTITLE_FONT
        row += 1
        for alerta in alertas:
            ws.cell(row=row, column=1, value=alerta).font = NORMAL_FONT
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            row += 1

    _auto_width(ws)
    ws.sheet_properties.tabColor = NAVY


def _build_propriedades(wb, consulta: dict):
    """Sheet 2: Tabela de propriedades."""
    ws = wb.create_sheet("Propriedades")

    headers = [
        "CAR Code", "Score",
        "Area Total (ha)", "Area Consolidada (ha)", "Area Agricultavel (ha)",
        "Area Pastagem (ha)", "Solo Exposto (ha)",
        "NDVI Medio", "NDVI Tendencia",
        "Embargo", "Quilombola", "Terra Indigena", "UC",
        "Assentamento", "Alertas",
    ]
    _write_header_row(ws, 1, headers)
    ws.freeze_panes = "A2"

    resultado = consulta.get("resultado", {})
    propriedades = resultado.get("propriedades", [])

    for i, prop in enumerate(propriedades):
        row = i + 2
        score = prop.get("score_ambiental", "N/D")
        ndvi = prop.get("ndvi", {})
        embargos = prop.get("embargos", {})
        sobrep = prop.get("sobreposicoes", {})
        alertas_prop = prop.get("alertas", [])
        areas = prop.get("areas", {})

        values = [
            prop.get("car_code", "N/D"),
            score,
            areas.get("area_total_ha", 0) or 0,
            areas.get("area_consolidada_ha", 0) or 0,
            areas.get("area_agricultavel_ha", 0) or 0,
            areas.get("area_pastagem_ha", 0) or 0,
            areas.get("area_solo_exposto_ha", 0) or 0,
            f"{ndvi.get('ndvi_mean', 0):.3f}" if ndvi.get("ndvi_mean") else "N/D",
            ndvi.get("tendencia", "N/D"),
            _bool_to_text(embargos.get("tem_embargo", False)),
            _bool_to_text(_safe_get(sobrep, "quilombolas", "tem_sobreposicao", default=False)),
            _bool_to_text(_safe_get(sobrep, "terras_indigenas", "tem_sobreposicao", default=False)),
            _bool_to_text(_safe_get(sobrep, "unidades_conservacao", "tem_sobreposicao", default=False)),
            _bool_to_text(_safe_get(sobrep, "assentamentos", "tem_sobreposicao", default=False)),
            "; ".join(alertas_prop) if alertas_prop else "Nenhum",
        ]

        fill = SCORE_FILLS.get(score)
        _write_row(ws, row, values, fill=fill)

    _auto_width(ws)
    ws.sheet_properties.tabColor = GREEN


def _build_ndvi(wb, consulta: dict):
    """Sheet 3: NDVI detalhado por propriedade."""
    ws = wb.create_sheet("NDVI Detalhado")

    headers = [
        "CAR Code", "NDVI Medio", "NDVI Mediana", "NDVI Min",
        "NDVI Max", "Tendencia", "Cobertura Vegetal (%)", "Data Cena",
    ]
    _write_header_row(ws, 1, headers)
    ws.freeze_panes = "A2"

    resultado = consulta.get("resultado", {})
    propriedades = resultado.get("propriedades", [])

    for i, prop in enumerate(propriedades):
        row = i + 2
        ndvi = prop.get("ndvi", {})
        values = [
            prop.get("car_code", "N/D"),
            ndvi.get("ndvi_mean", "N/D"),
            ndvi.get("ndvi_median", "N/D"),
            ndvi.get("ndvi_min", "N/D"),
            ndvi.get("ndvi_max", "N/D"),
            ndvi.get("tendencia", "N/D"),
            ndvi.get("cobertura_vegetal_pct", "N/D"),
            ndvi.get("data_cena", "N/D"),
        ]
        _write_row(ws, row, values)

    _auto_width(ws)
    ws.sheet_properties.tabColor = GREEN


def _build_embargos_sobreposicoes(wb, consulta: dict):
    """Sheet 4: Embargos e sobreposicoes detalhados."""
    ws = wb.create_sheet("Embargos & Sobreposicoes")

    headers = [
        "CAR Code", "Embargo Ativo", "Detalhes Embargo",
        "Terra Indigena", "Quilombola", "Unidade Conservacao", "Assentamento",
        "Detalhes Sobreposicao",
    ]
    _write_header_row(ws, 1, headers)
    ws.freeze_panes = "A2"

    resultado = consulta.get("resultado", {})
    propriedades = resultado.get("propriedades", [])

    for i, prop in enumerate(propriedades):
        row = i + 2
        embargos = prop.get("embargos", {})
        sobrep = prop.get("sobreposicoes", {})

        # Collect overlap details
        detalhes_parts = []
        for key, label in [
            ("terras_indigenas", "TI"),
            ("quilombolas", "Quilombola"),
            ("unidades_conservacao", "UC"),
            ("assentamentos", "Assentamento"),
        ]:
            sub = sobrep.get(key, {})
            if sub.get("tem_sobreposicao"):
                det = sub.get("detalhes", sub.get("nome", ""))
                if det:
                    detalhes_parts.append(f"{label}: {det}")

        values = [
            prop.get("car_code", "N/D"),
            _bool_to_text(embargos.get("tem_embargo", False)),
            embargos.get("detalhes", "Nenhum") if embargos.get("tem_embargo") else "Nenhum",
            _bool_to_text(_safe_get(sobrep, "terras_indigenas", "tem_sobreposicao", default=False)),
            _bool_to_text(_safe_get(sobrep, "quilombolas", "tem_sobreposicao", default=False)),
            _bool_to_text(_safe_get(sobrep, "unidades_conservacao", "tem_sobreposicao", default=False)),
            _bool_to_text(_safe_get(sobrep, "assentamentos", "tem_sobreposicao", default=False)),
            "; ".join(detalhes_parts) if detalhes_parts else "Nenhuma sobreposicao",
        ]

        # Highlight rows with issues
        has_issue = embargos.get("tem_embargo", False) or any(
            sobrep.get(k, {}).get("tem_sobreposicao", False)
            for k in ("terras_indigenas", "quilombolas", "unidades_conservacao", "assentamentos")
        )
        fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid") if has_issue else None
        _write_row(ws, row, values, fill=fill)

    _auto_width(ws)
    ws.sheet_properties.tabColor = RED


def _build_cruzamento_sigef(wb, consulta: dict):
    """Sheet 5: Cruzamento SIGEF (se disponivel)."""
    cruzamento = consulta.get("cruzamento", {})
    if not cruzamento:
        return

    ws = wb.create_sheet("Cruzamento SIGEF")

    # Summary section
    ws.merge_cells("A1:D1")
    cell = ws.cell(row=1, column=1, value="Cruzamento SIGEF")
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center")

    row = 3
    summary = [
        ("Cobertura (%)", f"{cruzamento.get('cobertura_pct', 0):.1f}%"),
        ("Propriedades Cadastradas", cruzamento.get("propriedades_cadastradas", 0)),
        ("Propriedades nos Documentos", cruzamento.get("propriedades_documentos", 0)),
    ]
    for label, value in summary:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2, value=value).font = NORMAL_FONT
        row += 1

    # Matches
    matches = cruzamento.get("matches", [])
    if matches:
        row += 1
        ws.cell(row=row, column=1, value="Propriedades Combinadas").font = SUBTITLE_FONT
        row += 1
        _write_header_row(ws, row, ["CAR Code", "Status", "Detalhes"])
        row += 1
        for m in matches:
            car = m.get("car_code", m) if isinstance(m, dict) else str(m)
            status = m.get("status", "Match") if isinstance(m, dict) else "Match"
            det = m.get("detalhes", "") if isinstance(m, dict) else ""
            _write_row(ws, row, [car, status, det])
            row += 1

    # Unregistered CARs
    nao_cad = cruzamento.get("nao_cadastradas", [])
    if nao_cad:
        row += 1
        ws.cell(row=row, column=1, value="CARs Nao Monitorados").font = SUBTITLE_FONT
        row += 1
        _write_header_row(ws, row, ["CAR Code", "Observacao"])
        row += 1
        for nc in nao_cad:
            car = nc.get("car_code", nc) if isinstance(nc, dict) else str(nc)
            obs = nc.get("observacao", "Nao monitorado") if isinstance(nc, dict) else "Nao monitorado"
            _write_row(ws, row, [car, obs],
                       fill=PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid"))
            row += 1

    # Alerts
    alertas = cruzamento.get("alertas", [])
    if alertas:
        row += 1
        ws.cell(row=row, column=1, value="Alertas SIGEF").font = SUBTITLE_FONT
        row += 1
        for a in alertas:
            ws.cell(row=row, column=1, value=a).font = NORMAL_FONT
            row += 1

    _auto_width(ws)
    ws.sheet_properties.tabColor = NAVY


def _build_alertas(wb, consulta: dict):
    """Sheet 6: Todos os alertas consolidados."""
    ws = wb.create_sheet("Alertas")

    headers = ["Severidade", "Propriedade", "Descricao"]
    _write_header_row(ws, 1, headers)
    ws.freeze_panes = "A2"

    resultado = consulta.get("resultado", {})
    propriedades = resultado.get("propriedades", [])

    row = 2

    # Per-property alerts
    for prop in propriedades:
        car = prop.get("car_code", "N/D")
        alertas_prop = prop.get("alertas", [])
        for alerta in alertas_prop:
            # Determine severity
            if "Vermelho" in alerta or "\U0001f534" in alerta:
                severity = "\U0001f534 Critico"
                fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
            elif "\u26a0" in alerta or "Amarelo" in alerta:
                severity = "\u26a0\ufe0f Atencao"
                fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
            else:
                severity = "\u2705 Info"
                fill = None

            _write_row(ws, row, [severity, car, alerta], fill=fill)
            row += 1

    # Consolidated alerts (not tied to a single property)
    alertas_consolidados = resultado.get("alertas_consolidados", [])
    seen = set()
    for prop in propriedades:
        seen.update(prop.get("alertas", []))

    for alerta in alertas_consolidados:
        if alerta in seen:
            continue
        if "Vermelho" in alerta or "\U0001f534" in alerta:
            severity = "\U0001f534 Critico"
            fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
        elif "\u26a0" in alerta or "Amarelo" in alerta:
            severity = "\u26a0\ufe0f Atencao"
            fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
        else:
            severity = "\u2705 Info"
            fill = None

        _write_row(ws, row, [severity, "Grupo", alerta], fill=fill)
        row += 1

    # SIGEF alerts
    cruzamento = consulta.get("cruzamento", {})
    for alerta in cruzamento.get("alertas", []):
        if "Vermelho" in alerta or "\U0001f534" in alerta:
            severity = "\U0001f534 Critico"
            fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
        elif "\u26a0" in alerta:
            severity = "\u26a0\ufe0f Atencao"
            fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
        else:
            severity = "\u2705 Info"
            fill = None

        _write_row(ws, row, [severity, "SIGEF", alerta], fill=fill)
        row += 1

    if row == 2:
        _write_row(ws, row, ["\u2705 Info", "Grupo", "Nenhum alerta identificado"])

    _auto_width(ws)
    ws.sheet_properties.tabColor = YELLOW


# ---------------------------------------------------------------------------
# Main public function
# ---------------------------------------------------------------------------
def generate_agro_excel(consulta: dict) -> bytes:
    """Generate Excel report from Dados Fazenda consultation.

    Args:
        consulta: dict with keys:
            - busca: str (CPF/CNPJ/CAR input)
            - tipo: str (CPF/CNPJ/CAR)
            - timestamp: str
            - resultado: dict with:
                - total_propriedades: int
                - area_total_ha: float
                - propriedades: list of dicts, each with:
                    - car_code: str
                    - score_ambiental: str (Verde/Amarelo/Vermelho)
                    - ndvi: dict (ndvi_mean, ndvi_median, tendencia, cobertura_vegetal_pct, data_cena)
                    - embargos: dict (tem_embargo, detalhes)
                    - sobreposicoes: dict with quilombolas, terras_indigenas, assentamentos, unidades_conservacao
                        each having tem_sobreposicao: bool
                    - incra: dict
                    - alertas: list of str
                - alertas_consolidados: list of str
                - score_ambiental_grupo: str
            - cruzamento: dict with:
                - cobertura_pct: float
                - propriedades_cadastradas: int
                - propriedades_documentos: int
                - matches: list
                - nao_cadastradas: list
                - alertas: list

    Returns:
        bytes of the xlsx file
    """
    wb = Workbook()

    _build_resumo(wb, consulta)
    _build_propriedades(wb, consulta)
    _build_ndvi(wb, consulta)
    _build_embargos_sobreposicoes(wb, consulta)
    _build_cruzamento_sigef(wb, consulta)
    _build_alertas(wb, consulta)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
