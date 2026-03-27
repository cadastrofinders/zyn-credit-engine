"""
ZYN Capital — Gerador de Planilha de Análise Técnica de Crédito
Gera modelo financeiro .xlsx boutique-quality a partir da análise de crédito MAC ZYN v3.

Suporta:
- Análise por empresa individual
- Consolidado de grupo econômico
- Indicadores financeiros e contábeis
- Simulador de viabilidade / stress
- Resumo executivo com parecer
- Cruzamento histórico
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ═══════════════════════════════════════════════════════════════════════════════
# ZYN Color Palette
# ═══════════════════════════════════════════════════════════════════════════════
NAVY = "223040"
DARK_SLATE = "3A4F63"
GREEN = "2E7D4F"
GREEN_LIGHT = "D5F5E3"
GRAY = "8B9197"
LIGHT_GRAY = "F2F4F6"
WHITE = "FFFFFF"
RED = "922B21"
RED_LIGHT = "F5B7B1"
YELLOW = "7D6608"
YELLOW_LIGHT = "FCF3CF"
BLUE = "2471A3"
ORANGE = "CA6F1E"
BORDER_COLOR = "D0D5DA"

# ── Fills ──────────────────────────────────────────────────────────────────────
FILL_NAVY = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
FILL_DARK_SLATE = PatternFill(start_color=DARK_SLATE, end_color=DARK_SLATE, fill_type="solid")
FILL_GREEN = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
FILL_GREEN_LIGHT = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")
FILL_RED_LIGHT = PatternFill(start_color=RED_LIGHT, end_color=RED_LIGHT, fill_type="solid")
FILL_YELLOW_LIGHT = PatternFill(start_color=YELLOW_LIGHT, end_color=YELLOW_LIGHT, fill_type="solid")
FILL_LIGHT = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

# ── Fonts ──────────────────────────────────────────────────────────────────────
FONT_TITLE = Font(name="Calibri", size=14, bold=True, color=WHITE)
FONT_SUBTITLE = Font(name="Calibri", size=11, bold=False, color=WHITE)
FONT_SECTION = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color=WHITE)
FONT_LABEL = Font(name="Calibri", size=10, color=DARK_SLATE)
FONT_LABEL_BOLD = Font(name="Calibri", size=10, bold=True, color=DARK_SLATE)
FONT_VALUE = Font(name="Calibri", size=10, color=NAVY)
FONT_VALUE_BOLD = Font(name="Calibri", size=10, bold=True, color=NAVY)
FONT_SMALL = Font(name="Calibri", size=9, color=GRAY, italic=True)
FONT_ALERT = Font(name="Calibri", size=10, bold=True, color=RED)
FONT_OK = Font(name="Calibri", size=10, bold=True, color=GREEN)
FONT_WARN = Font(name="Calibri", size=10, bold=True, color=YELLOW)
FONT_RATING_GREEN = Font(name="Calibri", size=14, bold=True, color=GREEN)
FONT_RATING_BLUE = Font(name="Calibri", size=14, bold=True, color=BLUE)
FONT_RATING_ORANGE = Font(name="Calibri", size=14, bold=True, color=ORANGE)
FONT_RATING_RED = Font(name="Calibri", size=14, bold=True, color=RED)
FONT_FOOTER = Font(name="Calibri", size=8, color=GRAY, italic=True)

# ── Borders ────────────────────────────────────────────────────────────────────
THIN_SIDE = Side(style="thin", color=BORDER_COLOR)
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color=NAVY))
NO_BORDER = Border()

# ── Alignments ─────────────────────────────────────────────────────────────────
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ── Number Formats (Brazilian) ─────────────────────────────────────────────────
# openpyxl uses Excel format codes; the locale rendering depends on Excel,
# but we use codes that produce correct output in pt-BR Excel.
BRL_FORMAT = '#,##0'
BRL_DECIMAL_FORMAT = '#,##0.00'
BRL_FULL_FORMAT = 'R$ #,##0'
BRL_FULL_DECIMAL = 'R$ #,##0.00'
PCT_FORMAT = '0.0%'
MULT_FORMAT = '0.00"x"'
DATE_FORMAT = 'DD/MM/YYYY'
INT_FORMAT = '#,##0'


# ═══════════════════════════════════════════════════════════════════════════════
# Utilities
# ═══════════════════════════════════════════════════════════════════════════════

def _safe(val, default=0) -> float:
    """Safely convert any value to a float number."""
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        cleaned = val.strip()
        # Remove currency symbols, "R$", "x", "%", dots-as-thousands, swap comma→dot
        cleaned = re.sub(r'[Rr]\$\s*', '', cleaned)
        cleaned = cleaned.replace('x', '').replace('X', '')
        cleaned = cleaned.replace('%', '')
        # Brazilian number: 1.234.567,89 → 1234567.89
        if ',' in cleaned and '.' in cleaned:
            cleaned = cleaned.replace('.', '').replace(',', '.')
        elif ',' in cleaned:
            cleaned = cleaned.replace(',', '.')
        try:
            return float(cleaned)
        except (ValueError, AttributeError):
            pass
    return float(default)


def _safe_str(val, default="—") -> str:
    """Safely convert to display string."""
    if val is None or val == "":
        return default
    return str(val)


def _safe_pct(val) -> float:
    """Convert a percentage value that may be 0-100 or 0-1 to 0-1 range."""
    v = _safe(val)
    if v > 1:
        return v / 100
    return v


def _rating_font(nota: str) -> Font:
    """Return appropriately colored font for a rating grade."""
    n = str(nota).upper().strip()
    if n in ("AAA", "AA", "A", "A+", "A-", "AA+", "AA-"):
        return FONT_RATING_GREEN
    if n in ("BBB", "BBB+", "BBB-", "BB+"):
        return FONT_RATING_BLUE
    if n in ("BB", "BB-", "B", "B+", "B-"):
        return FONT_RATING_ORANGE
    if n in ("CCC", "CC", "C", "D", "E"):
        return FONT_RATING_RED
    # Fallback: single-letter legacy
    if n.startswith("A"):
        return FONT_RATING_GREEN
    if n.startswith("B"):
        return FONT_RATING_BLUE
    if n.startswith("C"):
        return FONT_RATING_ORANGE
    if n in ("D", "E"):
        return FONT_RATING_RED
    return Font(name="Calibri", size=14, bold=True, color=NAVY)


def _status_font(status: str) -> Font:
    """Return font for a status string."""
    s = status.lower().strip()
    if s in ("ok", "viável", "forte", "adequado", "saudável", "cumprido"):
        return FONT_OK
    if s in ("atenção", "atencao", "alerta"):
        return FONT_WARN
    if s in ("crítico", "critico", "inviável", "fraco"):
        return FONT_ALERT
    return FONT_VALUE


def _status_fill(status: str) -> PatternFill | None:
    """Return conditional fill for status cells."""
    s = status.lower().strip()
    if s in ("ok", "viável", "forte", "adequado", "saudável", "cumprido"):
        return FILL_GREEN_LIGHT
    if s in ("atenção", "atencao", "alerta"):
        return FILL_YELLOW_LIGHT
    if s in ("crítico", "critico", "inviável", "fraco"):
        return FILL_RED_LIGHT
    return None


def _criticidade_fill(crit: str) -> PatternFill:
    """Return fill for criticality level."""
    c = crit.lower().strip()
    if c in ("alta", "crítica", "critica"):
        return FILL_RED_LIGHT
    if c in ("média", "media"):
        return FILL_YELLOW_LIGHT
    if c in ("baixa",):
        return FILL_GREEN_LIGHT
    return FILL_WHITE


def _criticidade_font(crit: str) -> Font:
    """Return font for criticality level."""
    c = crit.lower().strip()
    if c in ("alta", "crítica", "critica"):
        return FONT_ALERT
    if c in ("média", "media"):
        return FONT_WARN
    if c in ("baixa",):
        return FONT_OK
    return FONT_LABEL


# ── Layout Helpers ─────────────────────────────────────────────────────────────

def _set_col_widths(ws: Worksheet, widths: list[float]):
    """Set column widths with a minimum of 8."""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(w, 8)


def _write_title_bar(ws: Worksheet, row: int, text: str, cols: int = 7):
    """Write the branded title bar (navy background, white text, merged)."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_TITLE
    cell.fill = FILL_NAVY
    cell.alignment = ALIGN_CENTER
    ws.row_dimensions[row].height = 30
    # Paint all merged cells
    for c in range(2, cols + 1):
        ws.cell(row=row, column=c).fill = FILL_NAVY


def _write_subtitle_bar(ws: Worksheet, row: int, text: str, cols: int = 7):
    """Write a subtitle bar under the title."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_SUBTITLE
    cell.fill = FILL_NAVY
    cell.alignment = ALIGN_CENTER
    ws.row_dimensions[row].height = 20
    for c in range(2, cols + 1):
        ws.cell(row=row, column=c).fill = FILL_NAVY


def _section_header(ws: Worksheet, row: int, title: str, cols: int = 7) -> int:
    """Write a dark-slate section header spanning full width. Returns next row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = FONT_SECTION
    cell.fill = FILL_DARK_SLATE
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24
    for c in range(2, cols + 1):
        ws.cell(row=row, column=c).fill = FILL_DARK_SLATE
    return row + 1


def _write_header_row(ws: Worksheet, row: int, values: list[str], cols: int = 7):
    """Write a table header row (navy bg, white bold text, bottom border)."""
    for col_idx in range(1, cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = values[col_idx - 1] if col_idx - 1 < len(values) else ""
        cell.font = FONT_HEADER
        cell.fill = FILL_NAVY
        cell.alignment = ALIGN_CENTER
        cell.border = Border(
            left=THIN_SIDE, right=THIN_SIDE,
            top=THIN_SIDE,
            bottom=Side(style="medium", color=NAVY),
        )
    ws.row_dimensions[row].height = 22


def _write_data_row(
    ws: Worksheet,
    row: int,
    values: list,
    fonts: list[Font | None] | None = None,
    formats: list[str | None] | None = None,
    fills: list[PatternFill | None] | None = None,
    cols: int = 7,
    row_alt: bool | None = None,
):
    """Write a data row with proper formatting and alternating fill."""
    is_alt = row_alt if row_alt is not None else (row % 2 == 0)
    default_fill = FILL_LIGHT if is_alt else FILL_WHITE

    for col_idx in range(1, cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        val = values[col_idx - 1] if col_idx - 1 < len(values) else ""
        cell.value = val
        cell.font = (fonts[col_idx - 1] if fonts and col_idx - 1 < len(fonts) and fonts[col_idx - 1] else FONT_VALUE)
        cell.alignment = ALIGN_LEFT if col_idx == 1 else ALIGN_RIGHT
        cell.border = THIN_BORDER

        # Custom fill takes priority, then alternating
        custom_fill = fills[col_idx - 1] if fills and col_idx - 1 < len(fills) else None
        cell.fill = custom_fill if custom_fill else default_fill

        # Number format
        if formats and col_idx - 1 < len(formats) and formats[col_idx - 1]:
            cell.number_format = formats[col_idx - 1]


def _write_kv_row(ws: Worksheet, row: int, label: str, value, cols: int = 7,
                  value_font: Font = None, number_format: str = None,
                  merge_value: bool = True):
    """Write a key-value pair row: label in col A, value in col B (optionally merged)."""
    is_alt = row % 2 == 0
    fill = FILL_LIGHT if is_alt else FILL_WHITE

    lbl_cell = ws.cell(row=row, column=1, value=label)
    lbl_cell.font = FONT_LABEL_BOLD
    lbl_cell.fill = fill
    lbl_cell.border = THIN_BORDER
    lbl_cell.alignment = ALIGN_LEFT

    if merge_value and cols > 2:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=cols)

    val_cell = ws.cell(row=row, column=2, value=value)
    val_cell.font = value_font or FONT_VALUE_BOLD
    val_cell.fill = fill
    val_cell.border = THIN_BORDER
    val_cell.alignment = ALIGN_LEFT
    if number_format:
        val_cell.number_format = number_format

    # Fill remaining merged cells
    for c in range(3, cols + 1):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).border = THIN_BORDER


def _write_footer(ws: Worksheet, row: int, cols: int = 7):
    """Write the ZYN branded footer."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1,
                   value="ZYN Capital — Crédito Estruturado & M&A | São Paulo | Confidencial")
    cell.font = FONT_FOOTER
    cell.alignment = ALIGN_CENTER
    for c in range(2, cols + 1):
        ws.cell(row=row, column=c).border = NO_BORDER


def _setup_print(ws: Worksheet):
    """Configure print settings: landscape, fit to width, repeat headers."""
    ws.sheet_properties.pageSetUpPr = None  # reset
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = '1:2'


def _freeze_pane(ws: Worksheet, cell_ref: str = "A3"):
    """Freeze panes at the given cell reference."""
    ws.freeze_panes = cell_ref


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Resumo Executivo
# ═══════════════════════════════════════════════════════════════════════════════

def _build_resumo(wb: Workbook, analise: dict, op: dict):
    ws = wb.active
    ws.title = "Resumo Executivo"
    _set_col_widths(ws, [32, 22, 18, 18, 18, 18, 18])
    _setup_print(ws)

    rating = analise.get("rating_final", {})
    kpis = analise.get("kpis", {})
    tomador_info = analise.get("tomador", {})

    # ── Title Bar ──
    _write_title_bar(ws, 1, "ZYN Capital — Análise Técnica de Crédito")
    _write_subtitle_bar(ws, 2, f"Data: {datetime.now().strftime('%d/%m/%Y')}  |  CONFIDENCIAL")
    _freeze_pane(ws, "A4")

    # ── Identificação do Tomador ──
    r = 4
    r = _section_header(ws, r, "IDENTIFICAÇÃO DO TOMADOR")

    info_pairs = [
        ("Razão Social", _safe_str(tomador_info.get("razao_social", op.get("tomador")))),
        ("CNPJ", _safe_str(tomador_info.get("cnpj", op.get("cnpj")))),
        ("Grupo Econômico", _safe_str(tomador_info.get("grupo_economico"))),
        ("Tipo de Operação", _safe_str(op.get("tipo_operacao"))),
        ("Volume Solicitado", _safe(op.get("volume", 0)), BRL_FULL_FORMAT),
        ("Prazo", f"{op.get('prazo_meses', '—')} meses"),
        ("Taxa Indicativa", _safe_str(op.get("taxa"))),
        ("Garantias", ", ".join(op.get("garantias", [])) or "—"),
    ]
    for item in info_pairs:
        label = item[0]
        val = item[1]
        fmt = item[2] if len(item) > 2 else None
        _write_kv_row(ws, r, label, val, number_format=fmt)
        r += 1

    # ── Rating & Parecer ──
    r += 1
    r = _section_header(ws, r, "RATING & PARECER")

    nota = rating.get("nota", "—")
    # Rating badge
    lbl = ws.cell(row=r, column=1, value="Rating Final")
    lbl.font = FONT_LABEL_BOLD
    lbl.border = THIN_BORDER
    lbl.fill = FILL_WHITE
    nota_cell = ws.cell(row=r, column=2, value=nota)
    nota_cell.font = _rating_font(nota)
    nota_cell.border = THIN_BORDER
    nota_cell.alignment = ALIGN_CENTER
    # Rating badge background
    sf = _status_fill("ok" if nota.upper().startswith("A") else ("atenção" if nota.upper().startswith("B") else "crítico"))
    if sf:
        nota_cell.fill = sf
    for c in range(3, 8):
        ws.cell(row=r, column=c).border = THIN_BORDER
        ws.cell(row=r, column=c).fill = FILL_WHITE
    r += 1

    _write_kv_row(ws, r, "Parecer", _safe_str(rating.get("parecer")))
    r += 1
    _write_kv_row(ws, r, "Justificativa", _safe_str(rating.get("justificativa")),
                  value_font=FONT_VALUE)
    ws.cell(row=r, column=2).alignment = ALIGN_WRAP
    ws.row_dimensions[r].height = 40
    r += 1

    # ── KPIs ──
    r += 1
    r = _section_header(ws, r, "INDICADORES-CHAVE (KPIs)")
    _write_header_row(ws, r, ["Indicador", "Valor", "Benchmark", "Status", "", "", ""])
    r += 1

    dscr = _safe(kpis.get("dscr"))
    ltv = _safe_pct(kpis.get("ltv"))
    div_ebitda = _safe(kpis.get("divida_liquida_ebitda"))
    margem = _safe_pct(kpis.get("margem_ebitda"))

    kpi_rows = [
        ("Receita Líquida", _safe(kpis.get("receita_liquida")), "—", "", BRL_FULL_FORMAT),
        ("EBITDA", _safe(kpis.get("ebitda")), "—", "", BRL_FULL_FORMAT),
        ("Margem EBITDA", margem, "> 15%",
         "OK" if margem > 0.15 else "Atenção", PCT_FORMAT),
        ("Dív.Líq. / EBITDA", div_ebitda, "< 3,5x",
         "OK" if div_ebitda < 3.5 else ("Atenção" if div_ebitda < 5 else "Crítico"), MULT_FORMAT),
        ("LTV", ltv, "< 80%",
         "OK" if ltv < 0.80 else ("Atenção" if ltv < 1.0 else "Crítico"), PCT_FORMAT),
        ("DSCR", dscr, "> 1,2x",
         "OK" if dscr > 1.2 else ("Atenção" if dscr > 1.0 else "Crítico"), MULT_FORMAT),
    ]

    for label, val, bench, status, fmt in kpi_rows:
        sf = _status_font(status) if status else FONT_VALUE
        sfill = _status_fill(status) if status else None
        _write_data_row(
            ws, r,
            [label, val, bench, status, "", "", ""],
            fonts=[FONT_LABEL_BOLD, FONT_VALUE_BOLD, FONT_SMALL, sf, None, None, None],
            formats=[None, fmt, None, None, None, None, None],
            fills=[None, None, None, sfill, None, None, None],
        )
        r += 1

    # ── Ratings por Seção ──
    r += 1
    r = _section_header(ws, r, "RATINGS POR SEÇÃO")
    _write_header_row(ws, r, ["Seção", "Rating", "Análise Resumida", "", "", "", ""])
    r += 1

    secoes = ["tomador", "patrimonio", "producao", "capital", "operacao",
              "pagamento", "onus", "riscos", "covenants", "cronograma"]
    for s in secoes:
        sec = analise.get(s, {})
        rating_sec = _safe_str(sec.get("rating_secao"))
        analise_text = _safe_str(sec.get("analise"))

        is_alt = r % 2 == 0
        fill = FILL_LIGHT if is_alt else FILL_WHITE

        ws.cell(row=r, column=1, value=s.replace("_", " ").title()).font = FONT_LABEL_BOLD
        ws.cell(row=r, column=1).fill = fill
        ws.cell(row=r, column=1).border = THIN_BORDER

        rating_cell = ws.cell(row=r, column=2, value=rating_sec)
        rating_cell.font = _rating_font(rating_sec) if rating_sec != "—" else FONT_VALUE_BOLD
        rating_cell.font = Font(name="Calibri", size=11, bold=True,
                                color=_rating_font(rating_sec).color if rating_sec != "—" else NAVY)
        rating_cell.alignment = ALIGN_CENTER
        rating_cell.fill = fill
        rating_cell.border = THIN_BORDER

        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        analise_cell = ws.cell(row=r, column=3, value=analise_text)
        analise_cell.font = FONT_SMALL
        analise_cell.alignment = ALIGN_WRAP
        analise_cell.fill = fill
        analise_cell.border = THIN_BORDER
        for c in range(4, 8):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    # ── Recomendações ──
    recs = rating.get("recomendacoes", [])
    if recs:
        r += 1
        r = _section_header(ws, r, "RECOMENDAÇÕES")
        for i, rec in enumerate(recs, 1):
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            cell = ws.cell(row=r, column=1, value=f"  {i}.  {rec}")
            cell.font = FONT_VALUE
            cell.alignment = ALIGN_WRAP
            cell.border = THIN_BORDER
            is_alt = r % 2 == 0
            cell.fill = FILL_LIGHT if is_alt else FILL_WHITE
            row_fill = FILL_LIGHT if is_alt else FILL_WHITE
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = row_fill
                ws.cell(row=r, column=c).border = THIN_BORDER
            r += 1

    # ── Footer ──
    r += 2
    _write_footer(ws, r)

    # Auto-filter on KPIs section would be on specific rows only;
    # we skip it here since this sheet is a summary layout.


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Indicadores Financeiros
# ═══════════════════════════════════════════════════════════════════════════════

def _build_indicadores(wb: Workbook, analise: dict, op: dict):
    ws = wb.create_sheet("Indicadores Financeiros")
    _set_col_widths(ws, [36, 22, 18, 18, 18, 16, 16])
    _setup_print(ws)

    kpis = analise.get("kpis", {})
    capital = analise.get("capital", {})
    indicadores = capital.get("indicadores", {})

    _write_title_bar(ws, 1, "Indicadores Financeiros & Contábeis")
    _write_subtitle_bar(ws, 2, f"{_safe_str(op.get('tomador', ''))}  |  {datetime.now().strftime('%d/%m/%Y')}")
    _freeze_pane(ws, "A4")

    # ── Indicadores de Resultado ──
    r = 4
    r = _section_header(ws, r, "INDICADORES DE RESULTADO")
    _write_header_row(ws, r, ["Indicador", "Valor", "Referência", "Status", "", "", ""])
    r += 1

    receita = _safe(kpis.get("receita_liquida"))
    ebitda = _safe(kpis.get("ebitda"))
    margem = _safe_pct(kpis.get("margem_ebitda"))

    result_rows = [
        ("Receita Líquida", receita, "—", "", BRL_FULL_FORMAT),
        ("EBITDA", ebitda, "—", "", BRL_FULL_FORMAT),
        ("Margem EBITDA", margem, "> 15%",
         "Forte" if margem > 0.20 else ("Adequado" if margem > 0.15 else "Atenção"), PCT_FORMAT),
    ]
    for label, val, ref, status, fmt in result_rows:
        sf = _status_font(status) if status else FONT_VALUE
        sfill = _status_fill(status) if status else None
        _write_data_row(
            ws, r,
            [label, val, ref, status, "", "", ""],
            fonts=[FONT_LABEL_BOLD, FONT_VALUE_BOLD, FONT_SMALL, sf, None, None, None],
            formats=[None, fmt, None, None, None, None, None],
            fills=[None, None, None, sfill, None, None, None],
        )
        r += 1

    # ── Indicadores de Endividamento ──
    r += 1
    r = _section_header(ws, r, "INDICADORES DE ENDIVIDAMENTO")
    _write_header_row(ws, r, ["Indicador", "Valor", "Lim. Atenção", "Lim. Crítico", "Status", "", ""])
    r += 1

    div_ebitda = _safe(indicadores.get("divida_liquida_ebitda", kpis.get("divida_liquida_ebitda")))
    div_pl = _safe(indicadores.get("divida_pl"))
    liq_corrente = _safe(indicadores.get("liquidez_corrente"))
    roe = _safe(indicadores.get("roe"))
    dscr = _safe(kpis.get("dscr"))
    ltv = _safe_pct(kpis.get("ltv"))

    debt_rows = [
        ("Dív.Líq. / EBITDA", div_ebitda, 3.5, 5.0, MULT_FORMAT, False),
        ("Dívida / PL", div_pl, 1.5, 3.0, MULT_FORMAT, False),
        ("Liquidez Corrente", liq_corrente, 1.0, 0.8, MULT_FORMAT, True),
        ("DSCR", dscr, 1.2, 1.0, MULT_FORMAT, True),
        ("LTV", ltv, 0.80, 1.0, PCT_FORMAT, False),
        ("ROE", _safe_pct(roe), 0.10, 0.05, PCT_FORMAT, True),
    ]
    for label, val, atencao, critico, fmt, higher_better in debt_rows:
        if higher_better:
            status = "Forte" if val > atencao else ("Adequado" if val > critico else "Fraco")
        else:
            status = "Forte" if val < atencao else ("Atenção" if val < critico else "Crítico")
        sf = _status_font(status)
        sfill = _status_fill(status)
        _write_data_row(
            ws, r,
            [label, val, atencao, critico, status, "", ""],
            fonts=[FONT_LABEL_BOLD, FONT_VALUE_BOLD, FONT_SMALL, FONT_SMALL, sf, None, None],
            formats=[None, fmt, fmt, fmt, None, None, None],
            fills=[None, None, None, None, sfill, None, None],
        )
        r += 1

    # Auto-filter on the debt table
    ws.auto_filter.ref = f"A{r - len(debt_rows) - 1}:G{r - 1}"

    # ── Estrutura de Capital ──
    r += 1
    r = _section_header(ws, r, "ESTRUTURA DE CAPITAL")

    cap_text = _safe_str(capital.get("estrutura_capital"))
    endiv_text = _safe_str(capital.get("endividamento"))
    analise_text = _safe_str(capital.get("analise"))

    for label, val in [("Estrutura", cap_text), ("Endividamento", endiv_text), ("Análise", analise_text)]:
        _write_kv_row(ws, r, label, val, value_font=FONT_VALUE)
        ws.cell(row=r, column=2).alignment = ALIGN_WRAP
        ws.row_dimensions[r].height = max(20, min(60, len(str(val)) // 3))
        r += 1

    # ── Flags de Atenção ──
    r += 1
    r = _section_header(ws, r, "FLAGS DE ATENÇÃO")

    todas_flags = []
    for s in ["tomador", "patrimonio", "producao", "capital", "operacao",
              "pagamento", "onus", "riscos", "covenants", "cronograma"]:
        for flag in analise.get(s, {}).get("flags", []):
            todas_flags.append((s.replace("_", " ").title(), flag))

    if todas_flags:
        _write_header_row(ws, r, ["Seção", "Flag de Atenção", "", "", "", "", ""])
        r += 1
        for secao, flag in todas_flags:
            ws.cell(row=r, column=1, value=secao).font = FONT_LABEL_BOLD
            ws.cell(row=r, column=1).border = THIN_BORDER
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            flag_cell = ws.cell(row=r, column=2, value=f"⚠  {flag}")
            flag_cell.font = FONT_ALERT
            flag_cell.alignment = ALIGN_WRAP
            flag_cell.border = THIN_BORDER
            flag_cell.fill = FILL_RED_LIGHT
            is_alt = r % 2 == 0
            ws.cell(row=r, column=1).fill = FILL_LIGHT if is_alt else FILL_WHITE
            for c in range(3, 8):
                ws.cell(row=r, column=c).border = THIN_BORDER
                ws.cell(row=r, column=c).fill = FILL_RED_LIGHT
            r += 1
    else:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        cell = ws.cell(row=r, column=1, value="Nenhuma flag de atenção identificada.")
        cell.font = FONT_OK
        cell.fill = FILL_GREEN_LIGHT
        cell.alignment = ALIGN_CENTER
        for c in range(2, 8):
            ws.cell(row=r, column=c).fill = FILL_GREEN_LIGHT
        r += 1

    # Footer
    r += 2
    _write_footer(ws, r)


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 3 — Análise de Crédito Detalhada
# ═══════════════════════════════════════════════════════════════════════════════

def _build_credito(wb: Workbook, analise: dict, op: dict):
    ws = wb.create_sheet("Análise de Crédito")
    _set_col_widths(ws, [28, 42, 18, 18, 18, 16, 16])
    _setup_print(ws)

    _write_title_bar(ws, 1, "Análise de Crédito Detalhada — MAC ZYN v3")
    _write_subtitle_bar(ws, 2, f"{_safe_str(op.get('tomador', ''))}  |  {datetime.now().strftime('%d/%m/%Y')}")
    _freeze_pane(ws, "A4")

    r = 4
    secoes = [
        ("tomador", "Tomador"),
        ("patrimonio", "Patrimônio & Garantias"),
        ("producao", "Produção"),
        ("capital", "Estrutura de Capital"),
        ("operacao", "Operação"),
        ("pagamento", "Capacidade de Pagamento"),
        ("onus", "Ônus & Gravames"),
        ("riscos", "Riscos"),
        ("covenants", "Covenants"),
        ("cronograma", "Cronograma"),
    ]

    for key, titulo in secoes:
        sec = analise.get(key, {})
        r = _section_header(ws, r, titulo.upper())

        # Main fields
        for field, val in sec.items():
            if field in ("flags", "rating_secao", "analise"):
                continue

            field_label = field.replace("_", " ").title()

            if isinstance(val, list):
                if val and isinstance(val[0], dict):
                    # Table for structured lists
                    headers = list(val[0].keys())
                    display_headers = [h.replace("_", " ").title() for h in headers]
                    padded = display_headers + [""] * (7 - len(display_headers))
                    _write_header_row(ws, r, padded[:7])
                    r += 1
                    for item in val:
                        row_vals = [_safe_str(item.get(h)) for h in headers]
                        row_vals += [""] * (7 - len(row_vals))
                        _write_data_row(ws, r, row_vals[:7],
                                        fonts=[FONT_VALUE] * 7)
                        r += 1
                elif val:
                    _write_kv_row(ws, r, field_label, ", ".join(str(v) for v in val),
                                  value_font=FONT_VALUE)
                    r += 1
            elif isinstance(val, dict):
                # Sub-dictionary: each key-value as a row
                for sub_key, sub_val in val.items():
                    sub_label = sub_key.replace("_", " ").title()
                    _write_kv_row(ws, r, f"  {sub_label}", _safe_str(sub_val),
                                  value_font=FONT_VALUE, merge_value=False)
                    r += 1
            else:
                _write_kv_row(ws, r, field_label, _safe_str(val), value_font=FONT_VALUE)
                # Enable word wrap for long text
                text_val = str(val) if val else ""
                if len(text_val) > 60:
                    ws.cell(row=r, column=2).alignment = ALIGN_WRAP
                    ws.row_dimensions[r].height = max(20, min(80, len(text_val) // 3))
                r += 1

        # Rating + Análise da seção
        rating_sec = _safe_str(sec.get("rating_secao"))
        lbl = ws.cell(row=r, column=1, value="Rating da Seção")
        lbl.font = FONT_LABEL_BOLD
        lbl.border = THIN_BORDER
        lbl.fill = FILL_LIGHT
        rc = ws.cell(row=r, column=2, value=rating_sec)
        rc.font = Font(name="Calibri", size=11, bold=True,
                       color=_rating_font(rating_sec).color if rating_sec != "—" else NAVY)
        rc.border = THIN_BORDER
        rc.alignment = ALIGN_CENTER
        rc.fill = FILL_LIGHT
        for c in range(3, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).fill = FILL_LIGHT
        r += 1

        # Section analysis text
        analise_text = _safe_str(sec.get("analise"))
        ws.cell(row=r, column=1, value="Análise").font = FONT_LABEL_BOLD
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        atxt = ws.cell(row=r, column=2, value=analise_text)
        atxt.font = FONT_VALUE
        atxt.alignment = ALIGN_WRAP
        atxt.border = THIN_BORDER
        ws.row_dimensions[r].height = max(25, min(80, len(analise_text) // 3))
        for c in range(3, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 2

    # Footer
    r += 1
    _write_footer(ws, r)


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 4 — Riscos & Viabilidade
# ═══════════════════════════════════════════════════════════════════════════════

def _build_riscos(wb: Workbook, analise: dict, op: dict):
    ws = wb.create_sheet("Riscos & Viabilidade")
    _set_col_widths(ws, [28, 20, 16, 16, 36, 16, 16])
    _setup_print(ws)

    _write_title_bar(ws, 1, "Matriz de Riscos & Análise de Viabilidade")
    _write_subtitle_bar(ws, 2, f"{_safe_str(op.get('tomador', ''))}  |  {datetime.now().strftime('%d/%m/%Y')}")
    _freeze_pane(ws, "A4")

    riscos = analise.get("riscos", {})
    r = 4

    # ── Categorias de Risco ──
    r = _section_header(ws, r, "CATEGORIAS DE RISCO")
    for cat in ["mercado", "credito", "operacional", "legal"]:
        val = _safe_str(riscos.get(cat))
        _write_kv_row(ws, r, f"Risco {cat.title()}", val, value_font=FONT_VALUE)
        ws.cell(row=r, column=2).alignment = ALIGN_WRAP
        if len(val) > 60:
            ws.row_dimensions[r].height = max(20, min(60, len(val) // 3))
        r += 1

    # ── Matriz de Riscos ──
    r += 1
    matriz = riscos.get("matriz_riscos", [])
    if matriz:
        r = _section_header(ws, r, "MATRIZ DE RISCOS")
        _write_header_row(ws, r, ["Risco", "Probabilidade", "Impacto", "Severidade", "Mitigante", "", ""])
        r += 1

        for item in matriz:
            prob = _safe_str(item.get("probabilidade"))
            impacto = _safe_str(item.get("impacto"))

            # Determine severity color
            severity_map = {"alto": "Crítico", "alta": "Crítico", "médio": "Atenção",
                            "média": "Atenção", "baixo": "OK", "baixa": "OK"}
            sev_prob = severity_map.get(prob.lower(), "")
            sev_imp = severity_map.get(impacto.lower(), "")
            # Combined: worst of the two
            severity_order = {"Crítico": 3, "Atenção": 2, "OK": 1, "": 0}
            combined = sev_prob if severity_order.get(sev_prob, 0) >= severity_order.get(sev_imp, 0) else sev_imp
            if not combined:
                combined = "Atenção"

            sf = _status_font(combined.lower() if combined else "")
            sfill = _status_fill(combined.lower() if combined else "")

            ws.cell(row=r, column=1, value=_safe_str(item.get("risco"))).font = FONT_LABEL_BOLD
            ws.cell(row=r, column=1).border = THIN_BORDER
            ws.cell(row=r, column=2, value=prob).font = FONT_VALUE
            ws.cell(row=r, column=2).border = THIN_BORDER
            ws.cell(row=r, column=2).alignment = ALIGN_CENTER
            ws.cell(row=r, column=3, value=impacto).font = FONT_VALUE
            ws.cell(row=r, column=3).border = THIN_BORDER
            ws.cell(row=r, column=3).alignment = ALIGN_CENTER
            sev_cell = ws.cell(row=r, column=4, value=combined)
            sev_cell.font = sf
            sev_cell.border = THIN_BORDER
            sev_cell.alignment = ALIGN_CENTER
            if sfill:
                sev_cell.fill = sfill

            ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
            mit = ws.cell(row=r, column=5, value=_safe_str(item.get("mitigante")))
            mit.font = FONT_SMALL
            mit.alignment = ALIGN_WRAP
            mit.border = THIN_BORDER

            is_alt = r % 2 == 0
            default_fill = FILL_LIGHT if is_alt else FILL_WHITE
            for c in [1, 2, 3, 5, 6, 7]:
                ws.cell(row=r, column=c).fill = default_fill
            for c in range(6, 8):
                ws.cell(row=r, column=c).border = THIN_BORDER
            r += 1

    # ── Stress Simulation ──
    r += 1
    kpis = analise.get("kpis", {})
    dscr = _safe(kpis.get("dscr"))
    ltv = _safe_pct(kpis.get("ltv"))

    r = _section_header(ws, r, "SIMULAÇÃO DE STRESS — VIABILIDADE")
    _write_header_row(ws, r, ["Cenário", "DSCR", "LTV", "Status", "", "", ""])
    r += 1

    scenarios = [
        ("Base (Atual)", dscr, ltv),
        ("Stress Juros +300bps", dscr * 0.85, ltv * 1.05),
        ("Stress Receita -20%", dscr * 0.80, ltv * 1.10),
        ("Stress Combinado", dscr * 0.70, ltv * 1.15),
    ]
    for nome, dscr_s, ltv_s in scenarios:
        status = "Viável" if dscr_s > 1.0 and ltv_s < 1.0 else "Inviável"
        sf = _status_font(status.lower())
        sfill = _status_fill(status.lower())
        _write_data_row(
            ws, r,
            [nome, dscr_s, ltv_s, status, "", "", ""],
            fonts=[FONT_LABEL_BOLD, FONT_VALUE_BOLD, FONT_VALUE_BOLD, sf, None, None, None],
            formats=[None, MULT_FORMAT, PCT_FORMAT, None, None, None, None],
            fills=[None, None, None, sfill, None, None, None],
        )
        r += 1

    # ── Covenants ──
    r += 1
    covenants = analise.get("covenants", {})
    clausulas = covenants.get("clausulas", [])
    if clausulas:
        r = _section_header(ws, r, "COVENANTS")
        _write_header_row(ws, r, ["Covenant", "Limite", "Atual", "Status", "", "", ""])
        r += 1

        for cl in clausulas:
            status = _safe_str(cl.get("status"))
            is_ok = "cumpr" in status.lower() or "ok" in status.lower()
            sf = FONT_OK if is_ok else FONT_ALERT
            sfill = FILL_GREEN_LIGHT if is_ok else FILL_RED_LIGHT
            _write_data_row(
                ws, r,
                [_safe_str(cl.get("covenant")), _safe_str(cl.get("limite")),
                 _safe_str(cl.get("atual")), status, "", "", ""],
                fonts=[FONT_LABEL_BOLD, FONT_VALUE, FONT_VALUE_BOLD, sf, None, None, None],
                fills=[None, None, None, sfill, None, None, None],
            )
            r += 1

    # Footer
    r += 2
    _write_footer(ws, r)


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 5 — Checklist & Lacunas
# ═══════════════════════════════════════════════════════════════════════════════

def _build_checklist(wb: Workbook, analise: dict):
    ws = wb.create_sheet("Checklist & Lacunas")
    _set_col_widths(ws, [42, 16, 42, 16, 16, 16, 16])
    _setup_print(ws)

    _write_title_bar(ws, 1, "Checklist de Documentação & Lacunas")
    _write_subtitle_bar(ws, 2, f"Gerado em {datetime.now().strftime('%d/%m/%Y')}")
    _freeze_pane(ws, "A4")

    checklist = analise.get("checklist_lacunas", {})
    r = 4

    # Summary counters
    total_p = checklist.get("total_pendencias", 0)
    total_c = checklist.get("total_criticas", 0)

    r = _section_header(ws, r, "RESUMO DE PENDÊNCIAS")

    ws.cell(row=r, column=1, value="Total de Pendências").font = FONT_LABEL_BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    pend_cell = ws.cell(row=r, column=2, value=total_p)
    pend_cell.font = FONT_VALUE_BOLD
    pend_cell.border = THIN_BORDER
    pend_cell.number_format = INT_FORMAT
    ws.cell(row=r, column=1).fill = FILL_WHITE
    pend_cell.fill = FILL_WHITE
    r += 1

    ws.cell(row=r, column=1, value="Pendências Críticas").font = FONT_LABEL_BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    crit_cell = ws.cell(row=r, column=2, value=total_c)
    crit_cell.font = FONT_ALERT if total_c > 0 else FONT_OK
    crit_cell.fill = FILL_RED_LIGHT if total_c > 0 else FILL_GREEN_LIGHT
    crit_cell.border = THIN_BORDER
    crit_cell.number_format = INT_FORMAT
    ws.cell(row=r, column=1).fill = FILL_LIGHT
    r += 2

    # ── Documentos Faltantes ──
    docs_f = checklist.get("documentos_faltantes", [])
    if docs_f:
        r = _section_header(ws, r, "DOCUMENTOS FALTANTES")
        _write_header_row(ws, r, ["Documento", "Criticidade", "Motivo / Observação", "", "", "", ""])
        r += 1
        for d in docs_f:
            crit = _safe_str(d.get("criticidade"))
            cfont = _criticidade_font(crit)
            cfill = _criticidade_fill(crit)
            motivo = _safe_str(d.get("motivo"))

            ws.cell(row=r, column=1, value=_safe_str(d.get("item"))).font = FONT_VALUE
            ws.cell(row=r, column=1).border = THIN_BORDER
            crit_c = ws.cell(row=r, column=2, value=crit)
            crit_c.font = cfont
            crit_c.fill = cfill
            crit_c.border = THIN_BORDER
            crit_c.alignment = ALIGN_CENTER
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
            mot = ws.cell(row=r, column=3, value=motivo)
            mot.font = FONT_SMALL
            mot.alignment = ALIGN_WRAP
            mot.border = THIN_BORDER

            is_alt = r % 2 == 0
            fill = FILL_LIGHT if is_alt else FILL_WHITE
            ws.cell(row=r, column=1).fill = fill
            for c in range(3, 8):
                ws.cell(row=r, column=c).border = THIN_BORDER
                ws.cell(row=r, column=c).fill = fill
            r += 1

    # ── Informações Pendentes ──
    r += 1
    info_p = checklist.get("informacoes_pendentes", [])
    if info_p:
        r = _section_header(ws, r, "INFORMAÇÕES PENDENTES")
        _write_header_row(ws, r, ["Item", "Criticidade", "Motivo / Observação", "", "", "", ""])
        r += 1
        for p in info_p:
            crit = _safe_str(p.get("criticidade"))
            cfont = _criticidade_font(crit)
            cfill = _criticidade_fill(crit)
            motivo = _safe_str(p.get("motivo"))

            ws.cell(row=r, column=1, value=_safe_str(p.get("item"))).font = FONT_VALUE
            ws.cell(row=r, column=1).border = THIN_BORDER
            crit_c = ws.cell(row=r, column=2, value=crit)
            crit_c.font = cfont
            crit_c.fill = cfill
            crit_c.border = THIN_BORDER
            crit_c.alignment = ALIGN_CENTER
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
            mot = ws.cell(row=r, column=3, value=motivo)
            mot.font = FONT_SMALL
            mot.alignment = ALIGN_WRAP
            mot.border = THIN_BORDER

            is_alt = r % 2 == 0
            fill = FILL_LIGHT if is_alt else FILL_WHITE
            ws.cell(row=r, column=1).fill = fill
            for c in range(3, 8):
                ws.cell(row=r, column=c).border = THIN_BORDER
                ws.cell(row=r, column=c).fill = fill
            r += 1

    # Footer
    r += 2
    _write_footer(ws, r)


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 6 — Cruzamento Histórico
# ═══════════════════════════════════════════════════════════════════════════════

def _build_cross_reference(wb: Workbook, analise: dict, op: dict, historico: list[dict]):
    """Aba de cruzamento com análises anteriores — memória do motor."""
    ws = wb.create_sheet("Cruzamento Histórico")
    _set_col_widths(ws, [28, 16, 16, 16, 18, 18, 20])
    _setup_print(ws)

    _write_title_bar(ws, 1, "Cruzamento com Análises Anteriores")
    _write_subtitle_bar(ws, 2, f"Histórico de {1 + len(historico)} análise(s)  |  {datetime.now().strftime('%d/%m/%Y')}")
    _freeze_pane(ws, "A5")

    r = 4
    r = _section_header(ws, r, "COMPARATIVO DE OPERAÇÕES ANALISADAS")
    _write_header_row(ws, r, ["Tomador", "Rating", "DSCR", "LTV", "Dív/EBITDA", "Parecer", "Data"])
    r += 1

    # Auto-filter
    ws.auto_filter.ref = f"A{r - 1}:G{r - 1 + 1 + len(historico)}"

    # Current analysis (highlighted)
    kpis = analise.get("kpis", {})
    rating = analise.get("rating_final", {})
    nota = rating.get("nota", "—")
    ltv_val = _safe_pct(kpis.get("ltv"))

    _write_data_row(
        ws, r,
        [
            _safe_str(op.get("tomador")),
            nota,
            _safe(kpis.get("dscr")),
            ltv_val,
            _safe(kpis.get("divida_liquida_ebitda")),
            _safe_str(rating.get("parecer")),
            datetime.now().strftime("%d/%m/%Y"),
        ],
        fonts=[FONT_VALUE_BOLD, Font(name="Calibri", size=11, bold=True,
               color=_rating_font(nota).color), FONT_VALUE_BOLD, FONT_VALUE_BOLD,
               FONT_VALUE_BOLD, FONT_VALUE_BOLD, FONT_VALUE],
        formats=[None, None, MULT_FORMAT, PCT_FORMAT, MULT_FORMAT, None, None],
        fills=[FILL_GREEN_LIGHT] * 7,  # Highlight current
    )
    r += 1

    # Historical analyses
    for item in historico:
        h_op = item.get("operacao", {})
        h_analise = item.get("analise", {})
        h_kpis = h_analise.get("kpis", {})
        h_rating = h_analise.get("rating_final", {})
        h_nota = h_rating.get("nota", "—")
        h_ltv = _safe_pct(h_kpis.get("ltv"))

        _write_data_row(
            ws, r,
            [
                _safe_str(h_op.get("tomador")),
                h_nota,
                _safe(h_kpis.get("dscr")),
                h_ltv,
                _safe(h_kpis.get("divida_liquida_ebitda")),
                _safe_str(h_rating.get("parecer")),
                _safe_str(item.get("data_analise")),
            ],
            fonts=[FONT_VALUE, Font(name="Calibri", size=10, bold=True,
                   color=_rating_font(h_nota).color), FONT_VALUE, FONT_VALUE,
                   FONT_VALUE, FONT_VALUE, FONT_VALUE],
            formats=[None, None, MULT_FORMAT, PCT_FORMAT, MULT_FORMAT, None, None],
        )
        r += 1

    # Rating evolution note
    r += 1
    r = _section_header(ws, r, "OBSERVAÇÕES")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    cell = ws.cell(row=r, column=1,
                   value="Ratings coloridos para comparação visual. "
                         "A análise atual está destacada em verde.")
    cell.font = FONT_SMALL
    cell.alignment = ALIGN_LEFT
    r += 1

    # Footer
    r += 2
    _write_footer(ws, r)


# ═══════════════════════════════════════════════════════════════════════════════
# Public API
# ═══════════════════════════════════════════════════════════════════════════════

def generate_excel(
    analise: dict,
    parametros: dict,
    output_path: str,
    historico: list[dict] | None = None,
) -> str:
    """
    Gera planilha de análise técnica de crédito boutique-quality.

    Args:
        analise: resultado do analyze_credit()
        parametros: parametros_operacao
        output_path: caminho de saída .xlsx
        historico: lista de análises anteriores para cruzamento (opcional)

    Returns:
        caminho do arquivo gerado
    """
    wb = Workbook()

    _build_resumo(wb, analise, parametros)
    _build_indicadores(wb, analise, parametros)
    _build_credito(wb, analise, parametros)
    _build_riscos(wb, analise, parametros)
    _build_checklist(wb, analise)

    if historico:
        _build_cross_reference(wb, analise, parametros, historico)

    # Cell protection on all sheets (protect structure, not individual cells)
    for ws in wb.worksheets:
        ws.protection.sheet = False  # Not locked by default; enable if needed

    wb.save(output_path)
    return output_path


def generate_excel_grupo(
    analyses: list[tuple[dict, dict]],
    grupo_nome: str,
    output_path: str,
) -> str:
    """
    Gera planilha consolidada para grupo econômico com múltiplas empresas.

    Args:
        analyses: lista de (analise, parametros) por empresa
        grupo_nome: nome do grupo econômico
        output_path: caminho de saída .xlsx

    Returns:
        caminho do arquivo gerado
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado Grupo"
    _set_col_widths(ws, [32, 22, 16, 16, 18, 18, 22])
    _setup_print(ws)

    _write_title_bar(ws, 1, f"ZYN Capital — Análise Consolidada: {grupo_nome}")
    _write_subtitle_bar(ws, 2,
        f"Data: {datetime.now().strftime('%d/%m/%Y')}  |  "
        f"{len(analyses)} empresa(s)  |  CONFIDENCIAL")
    _freeze_pane(ws, "A5")

    r = 4
    r = _section_header(ws, r, "RESUMO POR EMPRESA")
    _write_header_row(ws, r, ["Empresa", "Rating", "DSCR", "LTV", "Dív/EBITDA", "Parecer", "Volume"])
    r += 1

    # Auto-filter
    ws.auto_filter.ref = f"A{r - 1}:G{r - 1 + len(analyses)}"

    total_receita = 0.0
    total_ebitda = 0.0
    total_volume = 0.0

    for analise_item, params in analyses:
        kpis = analise_item.get("kpis", {})
        rating = analise_item.get("rating_final", {})
        vol = _safe(params.get("volume"))
        total_volume += vol
        total_receita += _safe(kpis.get("receita_liquida"))
        total_ebitda += _safe(kpis.get("ebitda"))
        nota = rating.get("nota", "—")
        ltv_val = _safe_pct(kpis.get("ltv"))

        _write_data_row(
            ws, r,
            [
                _safe_str(params.get("tomador")),
                nota,
                _safe(kpis.get("dscr")),
                ltv_val,
                _safe(kpis.get("divida_liquida_ebitda")),
                _safe_str(rating.get("parecer")),
                vol,
            ],
            fonts=[FONT_VALUE_BOLD,
                   Font(name="Calibri", size=10, bold=True, color=_rating_font(nota).color),
                   FONT_VALUE_BOLD, FONT_VALUE_BOLD, FONT_VALUE_BOLD, FONT_VALUE, FONT_VALUE_BOLD],
            formats=[None, None, MULT_FORMAT, PCT_FORMAT, MULT_FORMAT, None, BRL_FULL_FORMAT],
        )
        r += 1

    # Totals row
    r += 1
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = FILL_NAVY
        ws.cell(row=r, column=c).border = THIN_BORDER
    ws.cell(row=r, column=1, value="TOTAL DO GRUPO").font = FONT_HEADER
    ws.cell(row=r, column=1).fill = FILL_NAVY
    total_cell = ws.cell(row=r, column=7, value=total_volume)
    total_cell.font = FONT_HEADER
    total_cell.fill = FILL_NAVY
    total_cell.number_format = BRL_FULL_FORMAT
    total_cell.alignment = ALIGN_RIGHT
    r += 2

    # Consolidated KPIs
    r = _section_header(ws, r, "KPIs CONSOLIDADOS")
    margem_grupo = total_ebitda / total_receita if total_receita else 0

    consolidated = [
        ("Receita Líquida Total", total_receita, BRL_FULL_FORMAT),
        ("EBITDA Total", total_ebitda, BRL_FULL_FORMAT),
        ("Margem EBITDA Grupo", margem_grupo, PCT_FORMAT),
        ("Volume Total Operações", total_volume, BRL_FULL_FORMAT),
    ]
    for label, val, fmt in consolidated:
        _write_kv_row(ws, r, label, val, number_format=fmt, merge_value=False)
        r += 1

    # Footer
    r += 2
    _write_footer(ws, r)

    # Individual company sheets
    for analise_item, params in analyses:
        empresa = _safe_str(params.get("tomador", "Empresa"))[:31]
        _build_resumo_empresa(wb, analise_item, params, empresa)

    wb.save(output_path)
    return output_path


def _build_resumo_empresa(wb: Workbook, analise: dict, op: dict, sheet_name: str):
    """Cria aba de resumo executivo para uma empresa do grupo."""
    ws = wb.create_sheet(sheet_name)
    _set_col_widths(ws, [32, 22, 18, 18, 18, 18, 18])
    _setup_print(ws)

    rating = analise.get("rating_final", {})
    kpis = analise.get("kpis", {})

    _write_title_bar(ws, 1, f"Resumo Executivo — {_safe_str(op.get('tomador'))}")
    _write_subtitle_bar(ws, 2, f"Data: {datetime.now().strftime('%d/%m/%Y')}  |  CONFIDENCIAL")
    _freeze_pane(ws, "A4")

    r = 4
    r = _section_header(ws, r, "RATING & PARECER")

    nota = rating.get("nota", "—")
    lbl = ws.cell(row=r, column=1, value="Rating")
    lbl.font = FONT_LABEL_BOLD
    lbl.border = THIN_BORDER
    rc = ws.cell(row=r, column=2, value=nota)
    rc.font = _rating_font(nota)
    rc.border = THIN_BORDER
    rc.alignment = ALIGN_CENTER
    sf = _status_fill("ok" if nota.upper().startswith("A") else ("atenção" if nota.upper().startswith("B") else "crítico"))
    if sf:
        rc.fill = sf
    r += 1

    _write_kv_row(ws, r, "Parecer", _safe_str(rating.get("parecer")))
    r += 1
    _write_kv_row(ws, r, "Justificativa", _safe_str(rating.get("justificativa")),
                  value_font=FONT_VALUE)
    ws.cell(row=r, column=2).alignment = ALIGN_WRAP
    ws.row_dimensions[r].height = 40
    r += 2

    r = _section_header(ws, r, "INDICADORES")
    _write_header_row(ws, r, ["Indicador", "Valor", "Status", "", "", "", ""])
    r += 1

    margem = _safe_pct(kpis.get("margem_ebitda"))
    ltv = _safe_pct(kpis.get("ltv"))
    dscr = _safe(kpis.get("dscr"))
    div_ebitda = _safe(kpis.get("divida_liquida_ebitda"))

    kpi_data = [
        ("Receita Líquida", _safe(kpis.get("receita_liquida")), BRL_FULL_FORMAT, ""),
        ("EBITDA", _safe(kpis.get("ebitda")), BRL_FULL_FORMAT, ""),
        ("Margem EBITDA", margem, PCT_FORMAT,
         "Forte" if margem > 0.20 else ("Adequado" if margem > 0.15 else "Atenção")),
        ("Dív.Líq. / EBITDA", div_ebitda, MULT_FORMAT,
         "Forte" if div_ebitda < 3.5 else ("Atenção" if div_ebitda < 5 else "Crítico")),
        ("LTV", ltv, PCT_FORMAT,
         "Forte" if ltv < 0.80 else ("Atenção" if ltv < 1.0 else "Crítico")),
        ("DSCR", dscr, MULT_FORMAT,
         "Forte" if dscr > 1.2 else ("Atenção" if dscr > 1.0 else "Crítico")),
    ]
    for label, val, fmt, status in kpi_data:
        sf_font = _status_font(status) if status else FONT_VALUE
        sfill = _status_fill(status) if status else None
        _write_data_row(
            ws, r,
            [label, val, status, "", "", "", ""],
            fonts=[FONT_LABEL_BOLD, FONT_VALUE_BOLD, sf_font, None, None, None, None],
            formats=[None, fmt, None, None, None, None, None],
            fills=[None, None, sfill, None, None, None, None],
        )
        r += 1

    # Footer
    r += 2
    _write_footer(ws, r)
