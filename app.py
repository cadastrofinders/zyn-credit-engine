"""
ZYN Credit Engine — Plataforma de Análise de Crédito Estruturado
Streamlit App principal.
"""

import hashlib
import hmac
import json
import os
import time
from datetime import datetime
from pathlib import Path

import streamlit as st
from dotenv import load_dotenv

load_dotenv()

# ---------------------------------------------------------------------------
# Streamlit Cloud: inject st.secrets into os.environ so modules can read them
# ---------------------------------------------------------------------------
try:
    for key in ("ANTHROPIC_API_KEY", "NOTION_TOKEN"):
        if key in st.secrets and not os.environ.get(key):
            os.environ[key] = st.secrets[key]
except Exception:
    pass  # secrets not available (local dev)

# ---------------------------------------------------------------------------
# Authentication — Login gate + operation password
# ---------------------------------------------------------------------------
ALLOWED_DOMAIN = "zyncapital.com.br"
ALLOWED_EMAILS = {
    "danilo@zyncapital.com.br",
    "luiz@zyncapital.com.br",
    "renato@zyncapital.com.br",
}


def _get_user_password(email: str) -> str:
    """Get password for a specific user from secrets."""
    username = email.strip().lower().split("@")[0]
    try:
        users = st.secrets.get("auth_users", {})
        return users.get(username, "")
    except Exception:
        return os.environ.get(f"AUTH_{username.upper()}", "")


def _check_password(password: str, stored: str) -> bool:
    """Constant-time password comparison."""
    if not stored:
        return False
    return hmac.compare_digest(password, stored)


def _validate_email(email: str) -> bool:
    """Validate email is from allowed domain or whitelist."""
    email = email.strip().lower()
    if email in ALLOWED_EMAILS:
        return True
    if "@" in email and email.split("@")[1] == ALLOWED_DOMAIN:
        return True
    return False


def _has_any_users_configured() -> bool:
    """Check if at least one user password is configured."""
    try:
        users = st.secrets.get("auth_users", {})
        return len(users) > 0
    except Exception:
        return False


def _login_gate() -> bool:
    """Display login page and return True if authenticated."""
    if st.session_state.get("authenticated"):
        return True

    if not _has_any_users_configured():
        # No users configured — skip auth (local dev)
        return True

    st.markdown(
        """
        <div style="display:flex; justify-content:center; align-items:center; min-height:80vh;">
            <div style="max-width:420px; width:100%;">
                <div style="background:linear-gradient(135deg, #223040 0%, #2a3d52 60%, #1E6B42 100%);
                    border-radius:16px; padding:40px 36px 20px 36px; text-align:center; margin-bottom:24px;">
                    <h1 style="color:#FFFFFF; font-size:1.8rem; font-weight:800; margin:0;">
                        ZYN Credit Engine
                    </h1>
                    <p style="color:rgba(255,255,255,0.6); font-size:0.9rem; margin:8px 0 0 0;">
                        Acesso Restrito · Equipe ZYN Capital
                    </p>
                </div>
        """,
        unsafe_allow_html=True,
    )

    with st.form("login_form"):
        email = st.text_input("Email corporativo", placeholder="nome@zyncapital.com.br")
        password = st.text_input("Senha", type="password")
        submitted = st.form_submit_button("Entrar", use_container_width=True, type="primary")

        if submitted:
            if not email or not password:
                st.error("Preencha email e senha")
                return False
            if not _validate_email(email):
                st.error("Acesso restrito a emails @zyncapital.com.br")
                return False
            stored_pw = _get_user_password(email)
            if not _check_password(password, stored_pw):
                st.error("Senha incorreta")
                return False
            st.session_state.authenticated = True
            st.session_state.user_email = email.strip().lower()
            st.session_state.user_name = email.split("@")[0].capitalize()
            st.rerun()

    st.markdown(
        """
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    return False

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
# Use /tmp on Streamlit Cloud (read-only filesystem), local dir otherwise
_tmp = Path("/tmp/zyn-credit-engine")
UPLOADS_DIR = _tmp / "uploads"
OUTPUT_DIR = _tmp / "output"
CHECKLISTS_DIR = OUTPUT_DIR / "checklists"
try:
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    CHECKLISTS_DIR.mkdir(parents=True, exist_ok=True)
except OSError:
    # Fallback to local dir (dev environment)
    UPLOADS_DIR = BASE_DIR / "uploads"
    OUTPUT_DIR = BASE_DIR / "output"
    CHECKLISTS_DIR = OUTPUT_DIR / "checklists"
    UPLOADS_DIR.mkdir(exist_ok=True)
    OUTPUT_DIR.mkdir(exist_ok=True)
    CHECKLISTS_DIR.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# Module imports (lazy to allow app to load even if deps missing)
# ---------------------------------------------------------------------------
try:
    from modules.extractor import process_file, process_files_parallel, validate_cnpj
    from modules.analyzer import analyze_credit, analyze_incremental, get_analysis_summary
    from modules.docx_generator import generate_mac
    from modules.teaser_generator import generate_teaser
    from modules.excel_generator import generate_excel
    from modules.kyc_enrichment import enrich_analysis_data
    from modules.investor_matching import (
        match_investors, get_sector_benchmarks, _detect_sector,
        PRODUCT_DD_EXTRAS, GUARANTEE_TYPES,
    )
    from modules.analyzer import extract_car_codes, extract_grupo_economico
    from modules.dados_fazenda import DadosFazendaClient, get_client as get_df_client
    from modules.agro_excel_generator import generate_agro_excel

    MODULES_AVAILABLE = True
except ImportError as _imp_err:
    MODULES_AVAILABLE = False
    _IMPORT_ERROR = str(_imp_err)

# ---------------------------------------------------------------------------
# API Key check
# ---------------------------------------------------------------------------
API_KEY_SET = bool(os.environ.get("ANTHROPIC_API_KEY"))

# ---------------------------------------------------------------------------
# DD Checklist Template — 12 módulos com itens realistas
# ---------------------------------------------------------------------------
DD_CHECKLIST_TEMPLATE: dict[str, list[str]] = {
    "1. Cadastral": [
        "Cartão CNPJ atualizado",
        "Inscrição Estadual / Municipal",
        "Comprovante de endereço da sede (< 90 dias)",
        "Procurações vigentes dos representantes legais",
        "Ficha cadastral completa (KYC)",
        "Quadro societário atualizado",
        "Declaração de grupo econômico",
    ],
    "2. Societário": [
        "Contrato Social consolidado ou Estatuto Social",
        "Última alteração contratual",
        "Ata de eleição da diretoria vigente",
        "Acordo de acionistas (se houver)",
        "Organograma societário do grupo",
        "Certidão simplificada da Junta Comercial",
    ],
    "3. Financeiro": [
        "Balanço Patrimonial (último exercício auditado)",
        "Balanço Patrimonial (penúltimo exercício)",
        "DRE (último exercício)",
        "DRE (penúltimo exercício)",
        "Balancete acumulado do exercício corrente",
        "Fluxo de Caixa projetado (12 meses)",
        "Relatório de auditoria independente",
        "Notas explicativas",
    ],
    "4. Faturamento": [
        "Relatório de faturamento mensal (últimos 12 meses)",
        "Principais clientes (top 10 por receita)",
        "Contratos com clientes relevantes",
        "Concentração de receita por cliente",
        "Histórico de inadimplência de clientes",
    ],
    "5. Endividamento": [
        "Relação completa de endividamento bancário",
        "SCR / Registrato atualizado",
        "Contratos de empréstimo vigentes",
        "Cronograma de amortização por credor",
        "Cartas de anuência de credores (se necessário)",
        "Posição de derivativos / hedge",
    ],
    "6. FIDC/Fundo": [
        "Regulamento do Fundo (se aplicável)",
        "Política de crédito e elegibilidade de lastro",
        "Relatório do administrador/custodiante",
        "Composição da carteira de recebíveis",
        "Histórico de PDD e inadimplência da carteira",
        "Rating do fundo (se houver)",
    ],
    "7. Crédito": [
        "Histórico de crédito do tomador (Serasa/SPC)",
        "Score de crédito (bureau)",
        "Análise setorial",
        "Relatório de visita técnica",
        "Parecer jurídico sobre a operação",
        "Matriz de risco e mitigantes",
    ],
    "8. Sócios PF": [
        "Documentos pessoais (RG/CPF) dos sócios",
        "Declaração de IR dos sócios (último exercício)",
        "Certidão negativa de débitos fiscais (PF)",
        "Pesquisa de protestos e ações judiciais (PF)",
        "Certidão de distribuição cível e criminal",
        "Declaração de bens e patrimônio",
    ],
    "9. Certidões": [
        "CND Federal (Receita Federal / PGFN)",
        "CND Estadual (SEFAZ)",
        "CND Municipal (ISS / IPTU)",
        "CND Trabalhista (TST)",
        "CND FGTS (CEF)",
        "Certidão de protestos",
        "Certidão de distribuição cível",
        "Certidão de distribuição criminal",
    ],
    "10. Garantias": [
        "Matrícula atualizada do imóvel (< 30 dias)",
        "Laudo de avaliação do imóvel",
        "CCIR quitado (imóvel rural)",
        "CAR — Cadastro Ambiental Rural",
        "Certidão de ônus reais",
        "Apólice de seguro do bem",
        "Contrato de penhor / alienação fiduciária",
        "ITR quitado (imóvel rural)",
    ],
    "11. Jurídico": [
        "Pesquisa de processos judiciais (PJ)",
        "Contingências relevantes provisionadas",
        "Pareceres jurídicos sobre contingências",
        "Contratos materiais vigentes",
        "Licenças e alvarás operacionais",
        "Compliance e LGPD",
    ],
    "12. Institucional": [
        "Apresentação institucional da empresa",
        "Relatório de sustentabilidade / ESG",
        "Política de governança corporativa",
        "Estrutura de compliance / PLD",
        "Referências bancárias e comerciais",
    ],
}

# ---------------------------------------------------------------------------
# Custom CSS
# ---------------------------------------------------------------------------
CUSTOM_CSS = """
<style>
    /* ── Sidebar ─────────────────────────────────────────── */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1a2633 0%, #223040 40%, #2a3d52 100%);
    }
    section[data-testid="stSidebar"] * {
        color: #FFFFFF !important;
    }
    section[data-testid="stSidebar"] hr {
        border-color: rgba(255,255,255,0.12) !important;
    }
    /* ── Nav buttons premium ──────────────────────────── */
    section[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] {
        display: flex !important;
        flex-direction: column !important;
        gap: 6px !important;
    }
    section[data-testid="stSidebar"] label[data-baseweb="radio"] {
        background: rgba(255,255,255,0.05) !important;
        border: 1px solid rgba(255,255,255,0.10) !important;
        border-left: 3px solid transparent !important;
        border-radius: 6px !important;
        padding: 12px 16px !important;
        margin: 0 !important;
        transition: all 0.25s ease !important;
        cursor: pointer !important;
        display: flex !important;
        align-items: center !important;
    }
    section[data-testid="stSidebar"] label[data-baseweb="radio"] p,
    section[data-testid="stSidebar"] label[data-baseweb="radio"] span {
        font-size: 0.88rem !important;
        font-weight: 500 !important;
        letter-spacing: 0.3px !important;
    }
    section[data-testid="stSidebar"] label[data-baseweb="radio"]:hover {
        background: rgba(255,255,255,0.12) !important;
        border-color: rgba(255,255,255,0.20) !important;
        border-left-color: rgba(46,125,79,0.5) !important;
        transform: translateX(2px);
    }
    /* Selected nav item */
    section[data-testid="stSidebar"] label[data-baseweb="radio"]:has(input:checked) {
        background: rgba(46,125,79,0.20) !important;
        border-color: rgba(46,125,79,0.30) !important;
        border-left: 3px solid #2E7D4F !important;
        box-shadow: 0 2px 8px rgba(46,125,79,0.15) !important;
    }
    section[data-testid="stSidebar"] label[data-baseweb="radio"]:has(input:checked) p,
    section[data-testid="stSidebar"] label[data-baseweb="radio"]:has(input:checked) span {
        font-weight: 700 !important;
    }
    /* Hide radio circle */
    section[data-testid="stSidebar"] [data-testid="stRadio"] > div[role="radiogroup"] > label > div:first-child {
        display: none !important;
    }

    /* ── Page title ──────────────────────────────────────── */
    .main-title {
        font-size: 2.2rem;
        font-weight: 800;
        color: #223040;
        letter-spacing: -0.5px;
        margin-bottom: 0;
    }
    .main-subtitle {
        font-size: 1.1rem;
        color: #8B9197;
        margin-top: 0;
        margin-bottom: 8px;
    }

    /* ── Metric cards ────────────────────────────────────── */
    [data-testid="stMetric"] {
        background: #FFFFFF;
        border: 1px solid #E8ECF0;
        border-left: 4px solid #223040;
        padding: 16px 20px;
        border-radius: 10px;
        box-shadow: 0 1px 4px rgba(34,48,64,0.06);
        transition: box-shadow 0.2s ease;
    }
    [data-testid="stMetric"]:hover {
        box-shadow: 0 4px 12px rgba(34,48,64,0.10);
    }
    [data-testid="stMetric"] label {
        color: #8B9197 !important;
        font-size: 0.85rem !important;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    [data-testid="stMetric"] [data-testid="stMetricValue"] {
        color: #223040 !important;
        font-weight: 700 !important;
    }

    /* ── KPI strip (Nova Análise > Análise tab) ──────────── */
    .kpi-card {
        background: #FFFFFF;
        border: 1px solid #E8ECF0;
        border-radius: 10px;
        padding: 16px 12px;
        text-align: center;
        box-shadow: 0 1px 4px rgba(34,48,64,0.06);
    }
    .kpi-label {
        font-size: 0.72rem;
        color: #8B9197;
        text-transform: uppercase;
        letter-spacing: 0.6px;
        margin-bottom: 4px;
    }
    .kpi-value {
        font-size: 1.5rem;
        font-weight: 800;
        color: #223040;
    }

    /* ── Rating badges (AAA-D scale) ───────────────────── */
    .rating-badge {
        display: inline-block;
        padding: 12px 32px;
        border-radius: 12px;
        font-size: 2.4rem;
        font-weight: 800;
        color: #FFFFFF;
        text-align: center;
        letter-spacing: 2px;
        box-shadow: 0 4px 14px rgba(0,0,0,0.15);
    }
    .rating-AAA { background: linear-gradient(135deg, #0D5C2F, #1E8B4F); }
    .rating-AA { background: linear-gradient(135deg, #1E6B42, #2E9B62); }
    .rating-A { background: linear-gradient(135deg, #2E7D4F, #45B06A); }
    .rating-BBB { background: linear-gradient(135deg, #223040, #3A5570); }
    .rating-BB { background: linear-gradient(135deg, #5C6B08, #8B9B0E); }
    .rating-B { background: linear-gradient(135deg, #7D6608, #B8960E); }
    .rating-C { background: linear-gradient(135deg, #E65100, #FF7A22); }
    .rating-D { background: linear-gradient(135deg, #922B21, #C0392B); }

    /* ── Parecer colors ──────────────────────────────────── */
    .parecer-favoravel { color: #1E6B42; font-weight: 700; font-size: 1.1rem; }
    .parecer-ressalvas { color: #7D6608; font-weight: 700; font-size: 1.1rem; }
    .parecer-desfavoravel { color: #922B21; font-weight: 700; font-size: 1.1rem; }

    /* ── Section headers ─────────────────────────────────── */
    .section-header {
        background: linear-gradient(90deg, #223040, #2a3d52);
        color: #FFFFFF;
        padding: 10px 18px;
        border-radius: 8px;
        margin-top: 18px;
        font-weight: 600;
        letter-spacing: 0.3px;
    }

    /* ── Status badges (DD checklist) ────────────────────── */
    .status-ok {
        background-color: #EAF4EE;
        color: #1E6B42;
        padding: 3px 10px;
        border-radius: 6px;
        font-weight: 600;
        font-size: 0.8rem;
    }
    .status-pendente {
        background-color: #FDECEA;
        color: #922B21;
        padding: 3px 10px;
        border-radius: 6px;
        font-weight: 600;
        font-size: 0.8rem;
    }
    .status-desatualizado {
        background-color: #FEF9E7;
        color: #7D6608;
        padding: 3px 10px;
        border-radius: 6px;
        font-weight: 600;
        font-size: 0.8rem;
    }

    /* ── Tabs ─────────────────────────────────────────────── */
    .stTabs [data-baseweb="tab-list"] {
        gap: 4px;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px 8px 0 0;
        padding: 10px 20px;
        font-weight: 600;
    }
    .stTabs [aria-selected="true"] {
        background-color: #223040 !important;
        color: #FFFFFF !important;
    }

    /* ── File uploader ───────────────────────────────────── */
    [data-testid="stFileUploader"] {
        border: 2px dashed #D0D5DA;
        border-radius: 12px;
        padding: 8px;
        transition: border-color 0.2s ease;
    }
    [data-testid="stFileUploader"]:hover {
        border-color: #223040;
    }

    /* ── Expander (analysis sections) ────────────────────── */
    .streamlit-expanderHeader {
        background-color: #F2F4F6 !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        color: #223040 !important;
    }

    /* ── Progress bars ───────────────────────────────────── */
    .stProgress > div > div > div {
        background-color: #1E6B42 !important;
    }

    /* ── Info / Warning / Error boxes ────────────────────── */
    .stAlert {
        border-radius: 10px !important;
    }

    /* ── Footer ──────────────────────────────────────────── */
    .footer-text {
        color: rgba(255,255,255,0.5);
        font-size: 0.75rem;
        text-align: center;
        padding-top: 24px;
        letter-spacing: 0.5px;
    }

    /* ── General polish ──────────────────────────────────── */
    .block-container {
        padding-top: 2rem !important;
    }
    .stButton > button {
        border-radius: 8px;
        font-weight: 600;
        transition: all 0.2s ease;
    }
    .stButton > button:hover {
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(34,48,64,0.15);
    }
    .stDownloadButton > button {
        background-color: #1E6B42 !important;
        color: #FFFFFF !important;
        border: none !important;
        border-radius: 8px;
        font-weight: 600;
        padding: 8px 24px;
    }
    .stDownloadButton > button:hover {
        background-color: #246E45 !important;
        box-shadow: 0 4px 12px rgba(30,125,79,0.25);
    }
</style>
"""

# ---------------------------------------------------------------------------
# Session State Initialization
# ---------------------------------------------------------------------------
DEFAULTS: dict = {
    "operacoes": [],
    "current_op": None,
    "uploaded_files": [],
    "extracted_data": {},
    "analysis": None,
    "step": 0,
    "dd_status": {},
}

for key, default in DEFAULTS.items():
    if key not in st.session_state:
        st.session_state[key] = default

# ---------------------------------------------------------------------------
# Persistence — save/load extraction + analysis + op to disk so they survive
# reloads and redeploys on Streamlit Cloud
# ---------------------------------------------------------------------------
CACHE_FILE = OUTPUT_DIR / "session_cache.json"


def _save_cache():
    """Persist extracted_data, analysis, and current_op to disk."""
    try:
        payload = {
            "extracted_data": st.session_state.extracted_data or {},
            "analysis": st.session_state.analysis,
            "current_op": st.session_state.current_op,
        }
        CACHE_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2))
    except Exception:
        pass


def _load_cache():
    """Restore cached data if session_state is empty and cache file exists."""
    if st.session_state.extracted_data or not CACHE_FILE.exists():
        return
    try:
        payload = json.loads(CACHE_FILE.read_text())
        if payload.get("extracted_data"):
            st.session_state.extracted_data = payload["extracted_data"]
        if payload.get("analysis"):
            st.session_state.analysis = payload["analysis"]
        if payload.get("current_op"):
            st.session_state.current_op = payload["current_op"]
    except Exception:
        pass


_load_cache()

# ---------------------------------------------------------------------------
# History — save completed analyses (GitHub persistent + local fallback)
# ---------------------------------------------------------------------------
HISTORY_DIR = OUTPUT_DIR / "historico"
HISTORY_DIR.mkdir(exist_ok=True)

from modules.github_storage import (
    save_analysis as _gh_save,
    list_analyses as _gh_list,
    load_analysis as _gh_load,
    delete_analysis as _gh_delete,
)


def _save_to_history(op: dict, analise: dict, extracted: dict):
    """Save analysis to GitHub (persistent) + local (cache)."""
    tomador = op.get("tomador", "desconhecido").replace(" ", "_").replace("/", "-")
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"{tomador}_{ts}.json"
    payload = {
        "operacao": op,
        "analise": analise,
        "extracted_data": extracted,
        "data_analise": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    # Save locally (fast access)
    (HISTORY_DIR / filename).write_text(json.dumps(payload, ensure_ascii=False, indent=2))
    # Save to GitHub (persistent across deploys)
    _gh_save(filename, payload)
    # Auto-save checklist for this client
    tomador_orig = op.get("tomador", "desconhecido")
    if st.session_state.dd_status:
        _save_checklist(tomador_orig, st.session_state.dd_status, list(_detected_doc_types()))


def _list_history() -> list[dict]:
    """List analyses: GitHub first (persistent), local fallback."""
    # Try GitHub first
    gh_files = _gh_list()
    if gh_files:
        items = []
        for f in gh_files:
            try:
                data = _gh_load(f["name"])
                if data:
                    items.append(data)
            except Exception:
                continue
        if items:
            return items

    # Fallback to local files
    items = []
    for f in sorted(HISTORY_DIR.glob("*.json"), reverse=True):
        try:
            data = json.loads(f.read_text())
            data["_filename"] = f.name
            items.append(data)
        except Exception:
            continue
    return items


@st.cache_data(ttl=300)
def _list_history_cached() -> list[dict]:
    """Cached version of _list_history (5 min TTL)."""
    return _list_history()


def _load_operacoes_from_history():
    """Populate operacoes from history files if list is empty."""
    if st.session_state.operacoes:
        return
    for item in _list_history():
        op = item.get("operacao", {})
        if op:
            rating_final = item.get("analise", {}).get("rating_final", {})
            if rating_final:
                op["rating"] = rating_final.get("nota", op.get("rating", "—"))
                op["parecer"] = rating_final.get("parecer", op.get("parecer", "—"))
            if "status" not in op or op["status"] == "Em Andamento":
                op["status"] = "Concluída"
            st.session_state.operacoes.append(op)


_load_operacoes_from_history()


def _load_from_history(filename: str) -> dict | None:
    """Load analysis: local first, then GitHub."""
    path = HISTORY_DIR / filename
    if path.exists():
        return json.loads(path.read_text())
    # Try GitHub
    return _gh_load(filename)


def _delete_from_history(filename: str):
    """Delete from both local and GitHub."""
    path = HISTORY_DIR / filename
    if path.exists():
        path.unlink()
    _gh_delete(filename)


# ---------------------------------------------------------------------------
# Agro consultation history — persistent via GitHub
# ---------------------------------------------------------------------------
AGRO_HISTORY_FILE = "data/agro_consultas/historico.json"


def _save_agro_consulta(consulta: dict):
    """Append agro consultation to persistent history on GitHub."""
    history = _load_agro_history()
    history.append(consulta)
    # Keep last 100 consultations
    if len(history) > 100:
        history = history[-100:]
    _gh_save(f"../agro_consultas/historico.json", {"consultas": history})
    # Also save locally
    local_dir = OUTPUT_DIR / "agro_consultas"
    local_dir.mkdir(exist_ok=True)
    (local_dir / "historico.json").write_text(
        json.dumps({"consultas": history}, ensure_ascii=False, indent=2)
    )


def _load_agro_history() -> list[dict]:
    """Load agro consultation history from GitHub or local."""
    # Try local first (fast)
    local_path = OUTPUT_DIR / "agro_consultas" / "historico.json"
    if local_path.exists():
        try:
            data = json.loads(local_path.read_text())
            return data.get("consultas", [])
        except Exception:
            pass
    # Try GitHub
    try:
        data = _gh_load("../agro_consultas/historico.json")
        if data and "consultas" in data:
            return data["consultas"]
    except Exception:
        pass
    return []


# ---------------------------------------------------------------------------
# Checklist persistence — one checklist per client
# ---------------------------------------------------------------------------
# Expanded mapping: extracted doc tipo → checklist module + item(s)
TIPO_TO_CHECKLIST: dict[str, list[tuple[str, str]]] = {
    # Cadastral
    "cnpj": [("1. Cadastral", "Cartão CNPJ atualizado")],
    "procuracao": [("1. Cadastral", "Procurações vigentes dos representantes legais")],
    "kyc": [("1. Cadastral", "Ficha cadastral completa (KYC)")],
    # Societário
    "contrato": [
        ("2. Societário", "Contrato Social consolidado ou Estatuto Social"),
        ("2. Societário", "Última alteração contratual"),
    ],
    "ata": [("2. Societário", "Ata de eleição da diretoria vigente")],
    "organograma": [("2. Societário", "Organograma societário do grupo")],
    "certidao_junta": [("2. Societário", "Certidão simplificada da Junta Comercial")],
    # Financeiro
    "balanco": [
        ("3. Financeiro", "Balanço Patrimonial (último exercício auditado)"),
        ("3. Financeiro", "Balanço Patrimonial (penúltimo exercício)"),
    ],
    "dre": [
        ("3. Financeiro", "DRE (último exercício)"),
        ("3. Financeiro", "DRE (penúltimo exercício)"),
    ],
    "balancete": [("3. Financeiro", "Balancete acumulado do exercício corrente")],
    "fluxo_caixa": [("3. Financeiro", "Fluxo de Caixa projetado (12 meses)")],
    "auditoria": [
        ("3. Financeiro", "Relatório de auditoria independente"),
        ("3. Financeiro", "Notas explicativas"),
    ],
    # Faturamento
    "faturamento": [
        ("4. Faturamento", "Relatório de faturamento mensal (últimos 12 meses)"),
        ("4. Faturamento", "Principais clientes (top 10 por receita)"),
    ],
    "contrato_cliente": [("4. Faturamento", "Contratos com clientes relevantes")],
    # Endividamento
    "endividamento": [("5. Endividamento", "Relação completa de endividamento bancário")],
    "scr": [("5. Endividamento", "SCR / Registrato atualizado")],
    "contrato_emprestimo": [("5. Endividamento", "Contratos de empréstimo vigentes")],
    # FIDC
    "regulamento_fundo": [("6. FIDC/Fundo", "Regulamento do Fundo (se aplicável)")],
    "politica_credito": [("6. FIDC/Fundo", "Política de crédito e elegibilidade de lastro")],
    "carteira_recebiveis": [("6. FIDC/Fundo", "Composição da carteira de recebíveis")],
    "rating_fundo": [("6. FIDC/Fundo", "Rating do fundo (se houver)")],
    # Crédito
    "serasa": [("7. Crédito", "Histórico de crédito do tomador (Serasa/SPC)")],
    "score": [("7. Crédito", "Score de crédito (bureau)")],
    "parecer_juridico": [("7. Crédito", "Parecer jurídico sobre a operação")],
    # Sócios PF
    "doc_socio": [("8. Sócios PF", "Documentos pessoais (RG/CPF) dos sócios")],
    "irpf": [("8. Sócios PF", "Declaração de IR dos sócios (último exercício)")],
    # Certidões
    "certidao": [
        ("9. Certidões", "CND Federal (Receita Federal / PGFN)"),
        ("9. Certidões", "CND Estadual (SEFAZ)"),
        ("9. Certidões", "CND Municipal (ISS / IPTU)"),
        ("9. Certidões", "CND Trabalhista (TST)"),
        ("9. Certidões", "CND FGTS (CEF)"),
    ],
    "certidao_protesto": [("9. Certidões", "Certidão de protestos")],
    "certidao_civel": [("9. Certidões", "Certidão de distribuição cível")],
    # Garantias
    "matricula": [
        ("10. Garantias", "Matrícula atualizada do imóvel (< 30 dias)"),
        ("10. Garantias", "Certidão de ônus reais"),
    ],
    "laudo_avaliacao": [("10. Garantias", "Laudo de avaliação do imóvel")],
    "ccir_car": [
        ("10. Garantias", "CCIR quitado (imóvel rural)"),
        ("10. Garantias", "CAR — Cadastro Ambiental Rural"),
    ],
    "seguro": [("10. Garantias", "Apólice de seguro do bem")],
    "itr": [("10. Garantias", "ITR quitado (imóvel rural)")],
    "alienacao": [("10. Garantias", "Contrato de penhor / alienação fiduciária")],
    # Jurídico
    "processos": [("11. Jurídico", "Pesquisa de processos judiciais (PJ)")],
    "contingencias": [("11. Jurídico", "Contingências relevantes provisionadas")],
    "licencas": [("11. Jurídico", "Licenças e alvarás operacionais")],
    # Institucional
    "apresentacao": [("12. Institucional", "Apresentação institucional da empresa")],
    "esg": [("12. Institucional", "Relatório de sustentabilidade / ESG")],
    # Genéricos — nome do arquivo contém pistas
    "planejamento": [("4. Faturamento", "Relatório de faturamento mensal (últimos 12 meses)")],
    "producao": [("4. Faturamento", "Relatório de faturamento mensal (últimos 12 meses)")],
}


def _checklist_filename(tomador: str) -> str:
    """Generate a filename for a client's checklist."""
    safe = tomador.strip().replace(" ", "_").replace("/", "-")
    return f"checklist_{safe}.json"


def _save_checklist(tomador: str, dd_status: dict, extracted_types: list[str] | None = None):
    """Save checklist status for a specific client."""
    payload = {
        "tomador": tomador,
        "dd_status": dd_status,
        "extracted_types": extracted_types or [],
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    fname = _checklist_filename(tomador)
    (CHECKLISTS_DIR / fname).write_text(json.dumps(payload, ensure_ascii=False, indent=2))


def _load_checklist(tomador: str) -> dict | None:
    """Load a saved checklist for a specific client."""
    fname = _checklist_filename(tomador)
    path = CHECKLISTS_DIR / fname
    if path.exists():
        try:
            return json.loads(path.read_text())
        except Exception:
            return None
    return None


def _list_checklists() -> list[dict]:
    """List all saved checklists."""
    items = []
    for f in sorted(CHECKLISTS_DIR.glob("*.json"), reverse=True):
        try:
            data = json.loads(f.read_text())
            data["_filename"] = f.name
            items.append(data)
        except Exception:
            pass
    return items


def _auto_populate_checklist(dd_status: dict, extracted_data: dict) -> tuple[dict, list[str]]:
    """Auto-populate checklist from extracted documents. Returns (updated_status, matched_types)."""
    matched_types: list[str] = []

    # Collect all detected types
    tipos_detectados: set[str] = set()
    filenames_lower: list[str] = []
    for fname, item in extracted_data.items():
        classificacao = item.get("classificacao", {})
        tipo = classificacao.get("tipo")
        if tipo:
            tipos_detectados.add(tipo)
        filenames_lower.append(fname.lower())

    # Map tipos to checklist items
    for tipo in tipos_detectados:
        mappings = TIPO_TO_CHECKLIST.get(tipo, [])
        for modulo, item_name in mappings:
            if modulo in dd_status and item_name in dd_status[modulo]:
                if dd_status[modulo][item_name] != "OK":
                    dd_status[modulo][item_name] = "OK"
                    matched_types.append(f"{modulo} → {item_name}")

    # Heuristic: try to match filenames to checklist items
    filename_hints = {
        "scr": [("5. Endividamento", "SCR / Registrato atualizado")],
        "registrato": [("5. Endividamento", "SCR / Registrato atualizado")],
        "serasa": [("7. Crédito", "Histórico de crédito do tomador (Serasa/SPC)")],
        "spc": [("7. Crédito", "Histórico de crédito do tomador (Serasa/SPC)")],
        "fgts": [("9. Certidões", "CND FGTS (CEF)")],
        "trabalhist": [("9. Certidões", "CND Trabalhista (TST)")],
        "irpf": [("8. Sócios PF", "Declaração de IR dos sócios (último exercício)")],
        "procuracao": [("1. Cadastral", "Procurações vigentes dos representantes legais")],
        "itr": [("10. Garantias", "ITR quitado (imóvel rural)")],
        "ccir": [("10. Garantias", "CCIR quitado (imóvel rural)")],
        "car": [("10. Garantias", "CAR — Cadastro Ambiental Rural")],
        "seguro": [("10. Garantias", "Apólice de seguro do bem")],
        "avaliacao": [("10. Garantias", "Laudo de avaliação do imóvel")],
        "laudo": [("10. Garantias", "Laudo de avaliação do imóvel")],
        "auditoria": [("3. Financeiro", "Relatório de auditoria independente")],
        "faturamento": [("4. Faturamento", "Relatório de faturamento mensal (últimos 12 meses)")],
        "producao": [("4. Faturamento", "Relatório de faturamento mensal (últimos 12 meses)")],
        "planejamento": [("4. Faturamento", "Relatório de faturamento mensal (últimos 12 meses)")],
        "endividamento": [("5. Endividamento", "Relação completa de endividamento bancário")],
        "organograma": [("2. Societário", "Organograma societário do grupo")],
    }

    for fn in filenames_lower:
        for hint, mappings in filename_hints.items():
            if hint in fn:
                for modulo, item_name in mappings:
                    if modulo in dd_status and item_name in dd_status[modulo]:
                        if dd_status[modulo][item_name] != "OK":
                            dd_status[modulo][item_name] = "OK"
                            matched_types.append(f"{modulo} → {item_name} (arquivo: {fn})")

    return dd_status, matched_types


def _generate_checklist_excel(tomador: str, dd_status: dict) -> bytes:
    """Generate a branded ZYN checklist Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist DD"

    # ZYN Colors
    navy_fill = PatternFill(start_color="223040", end_color="223040", fill_type="solid")
    green_fill = PatternFill(start_color="2E7D4F", end_color="2E7D4F", fill_type="solid")
    ok_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    pend_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    desat_fill = PatternFill(start_color="FBE9E7", end_color="FBE9E7", fill_type="solid")
    header_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    white_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    bold_font = Font(name="Calibri", bold=True, size=11)
    normal_font = Font(name="Calibri", size=10)
    module_font = Font(name="Calibri", bold=True, size=11, color="223040")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20

    # Title row
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = f"CHECKLIST DUE DILIGENCE — {tomador}"
    cell.font = white_font
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Subtitle row
    ws.merge_cells("A2:D2")
    cell = ws["A2"]
    cell.value = f"ZYN Capital · {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    cell.font = Font(name="Calibri", size=10, color="FFFFFF")
    cell.fill = green_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 25

    # Summary row
    total_items = sum(len(itens) for itens in dd_status.values())
    total_ok = sum(1 for itens in dd_status.values() for s in itens.values() if s == "OK")
    total_pend = sum(1 for itens in dd_status.values() for s in itens.values() if s == "PENDENTE")
    total_desat = sum(1 for itens in dd_status.values() for s in itens.values() if s == "DESATUALIZADO")
    pct = f"{total_ok / total_items * 100:.0f}%" if total_items > 0 else "0%"

    ws.merge_cells("A3:D3")
    cell = ws["A3"]
    cell.value = f"Progresso: {pct} ({total_ok}/{total_items})  |  ✅ OK: {total_ok}  |  ⏳ Pendente: {total_pend}  |  ⚠️ Desatualizado: {total_desat}"
    cell.font = Font(name="Calibri", size=10, bold=True)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 28

    # Headers
    row = 5
    for col_idx, (header, width) in enumerate([("#", 6), ("Documento", 55), ("Status", 18), ("Observação", 20)], 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = bold_font
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[row].height = 25

    row = 6
    item_num = 0
    for modulo, itens in DD_CHECKLIST_TEMPLATE.items():
        # Module header row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value=modulo)
        cell.font = module_font
        cell.fill = PatternFill(start_color="E3EBF1", end_color="E3EBF1", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        ws.row_dimensions[row].height = 22
        row += 1

        for item_name in itens:
            item_num += 1
            status = dd_status.get(modulo, {}).get(item_name, "PENDENTE")

            # Number
            cell = ws.cell(row=row, column=1, value=item_num)
            cell.font = normal_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            # Document name
            cell = ws.cell(row=row, column=2, value=item_name)
            cell.font = normal_font
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border

            # Status with color
            cell = ws.cell(row=row, column=3, value=status)
            cell.font = Font(name="Calibri", size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            if status == "OK":
                cell.fill = ok_fill
                cell.font = Font(name="Calibri", size=10, bold=True, color="2E7D4F")
            elif status == "PENDENTE":
                cell.fill = pend_fill
                cell.font = Font(name="Calibri", size=10, bold=True, color="F57C00")
            elif status == "DESATUALIZADO":
                cell.fill = desat_fill
                cell.font = Font(name="Calibri", size=10, bold=True, color="D32F2F")

            # Observation (empty for user to fill)
            cell = ws.cell(row=row, column=4, value="")
            cell.font = normal_font
            cell.border = thin_border

            row += 1

    # Freeze panes
    ws.freeze_panes = "A6"

    # Print settings
    ws.print_title_rows = "1:5"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Session management
# ---------------------------------------------------------------------------
def _clear_session():
    """Reset session state for a new analysis."""
    st.session_state.extracted_data = {}
    st.session_state.analysis = None
    st.session_state.current_op = None
    st.session_state.uploaded_files = []
    st.session_state.step = 0
    st.session_state.dd_status = {}
    st.session_state.pop("_checklist_loaded_for", None)
    # Increment form counter to force Streamlit to recreate widgets with fresh values
    st.session_state.form_counter = st.session_state.get("form_counter", 0) + 1
    # Clear disk cache too
    if CACHE_FILE.exists():
        CACHE_FILE.unlink()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _to_num(val, default=0):
    """Safely convert any value to float."""
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        cleaned = val.replace("R$", "").replace("%", "").replace("x", "").replace("X", "").strip()
        cleaned = cleaned.replace(".", "").replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return default
    return default


def _fmt_brl(value, compact: bool = False) -> str:
    """Formata valor em R$ brasileiro. compact=True para caber em st.metric."""
    v = _to_num(value)
    if v == 0 and value:
        return str(value)
    if compact:
        if abs(v) >= 1_000_000_000:
            return f"R${v / 1_000_000_000:,.1f}B"
        if abs(v) >= 1_000_000:
            return f"R${v / 1_000_000:,.1f}MM"
        if abs(v) >= 1_000:
            return f"R${v / 1_000:,.0f}mil"
        return f"R${v:,.0f}"
    if abs(v) >= 1_000_000_000:
        return f"R$ {v / 1_000_000_000:,.2f} B"
    if abs(v) >= 1_000_000:
        return f"R$ {v / 1_000_000:,.2f} MM"
    if abs(v) >= 1_000:
        return f"R$ {v / 1_000:,.1f} mil"
    return f"R$ {v:,.2f}"


def _confidence_color(conf: float) -> str:
    if conf >= 0.85:
        return "green"
    if conf >= 0.6:
        return "orange"
    return "red"


def _rating_class(nota: str) -> str:
    valid_ratings = {"AAA", "AA", "A", "BBB", "BB", "B", "C", "D"}
    return f"rating-{nota}" if nota in valid_ratings else "rating-BBB"


def _parecer_class(parecer: str) -> str:
    if "Desfavorável" in parecer:
        return "parecer-desfavoravel"
    if "Ressalvas" in parecer:
        return "parecer-ressalvas"
    return "parecer-favoravel"


def _count_ops_by_status(status: str) -> int:
    return sum(1 for op in st.session_state.operacoes if op.get("status") == status)


def _upsert_operacao(op: dict):
    """Add or update operation in the list (dedup by tomador+cnpj)."""
    tomador = op.get("tomador", "").strip().lower()
    cnpj = op.get("cnpj", "").strip()
    for i, existing in enumerate(st.session_state.operacoes):
        if existing.get("tomador", "").strip().lower() == tomador and existing.get("cnpj", "").strip() == cnpj:
            st.session_state.operacoes[i] = op
            return
    st.session_state.operacoes.append(op)


def _detected_doc_types() -> set[str]:
    """Retorna conjunto de tipos de documento detectados na extração."""
    tipos = set()
    for item in st.session_state.extracted_data.values():
        classificacao = item.get("classificacao", {})
        tipo = classificacao.get("tipo")
        if tipo:
            tipos.add(tipo)
    return tipos


# ---------------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------------
def page_dashboard():
    # ── Hero Section ──────────────────────────────────────────────────────
    st.markdown(
        """
        <div style="
            background: linear-gradient(135deg, #223040 0%, #2a3d52 60%, #1E6B42 100%);
            border-radius: 16px;
            padding: 40px 36px 32px 36px;
            margin-bottom: 24px;
        ">
            <h1 style="color:#FFFFFF; font-size:2.4rem; font-weight:800; margin:0; letter-spacing:-0.5px;">
                ZYN Credit Engine
            </h1>
            <p style="color:rgba(255,255,255,0.7); font-size:1.05rem; margin:6px 0 0 0;">
                Motor de Análise de Crédito Estruturado &nbsp;·&nbsp; Análise Setorial + Investor Matching &nbsp;·&nbsp; Powered by Claude AI
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ── KPI Cards ─────────────────────────────────────────────────────────
    total = len(st.session_state.operacoes)
    em_andamento = _count_ops_by_status("Em Andamento")
    concluidas = _count_ops_by_status("Concluída")

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(
            f"""<div style="background:#FFFFFF; border:1px solid #E8ECF0; border-left:4px solid #223040;
            border-radius:10px; padding:20px; box-shadow:0 1px 4px rgba(34,48,64,0.06);">
            <p style="color:#8B9197; font-size:0.75rem; text-transform:uppercase; letter-spacing:0.8px; margin:0 0 6px 0;">Total Operações</p>
            <p style="color:#223040; font-size:2rem; font-weight:800; margin:0;">{total}</p>
            </div>""", unsafe_allow_html=True)
    with col2:
        st.markdown(
            f"""<div style="background:#FFFFFF; border:1px solid #E8ECF0; border-left:4px solid #7D6608;
            border-radius:10px; padding:20px; box-shadow:0 1px 4px rgba(34,48,64,0.06);">
            <p style="color:#8B9197; font-size:0.75rem; text-transform:uppercase; letter-spacing:0.8px; margin:0 0 6px 0;">Em Andamento</p>
            <p style="color:#7D6608; font-size:2rem; font-weight:800; margin:0;">{em_andamento}</p>
            </div>""", unsafe_allow_html=True)
    with col3:
        st.markdown(
            f"""<div style="background:#FFFFFF; border:1px solid #E8ECF0; border-left:4px solid #1E6B42;
            border-radius:10px; padding:20px; box-shadow:0 1px 4px rgba(34,48,64,0.06);">
            <p style="color:#8B9197; font-size:0.75rem; text-transform:uppercase; letter-spacing:0.8px; margin:0 0 6px 0;">Concluídas</p>
            <p style="color:#1E6B42; font-size:2rem; font-weight:800; margin:0;">{concluidas}</p>
            </div>""", unsafe_allow_html=True)
    with col4:
        taxa_sucesso = f"{(concluidas / total * 100):.0f}%" if total > 0 else "—"
        st.markdown(
            f"""<div style="background:#FFFFFF; border:1px solid #E8ECF0; border-left:4px solid #922B21;
            border-radius:10px; padding:20px; box-shadow:0 1px 4px rgba(34,48,64,0.06);">
            <p style="color:#8B9197; font-size:0.75rem; text-transform:uppercase; letter-spacing:0.8px; margin:0 0 6px 0;">Taxa de Conclusão</p>
            <p style="color:#223040; font-size:2rem; font-weight:800; margin:0;">{taxa_sucesso}</p>
            </div>""", unsafe_allow_html=True)

    st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)

    # ── Operações Recentes ────────────────────────────────────────────────
    if st.session_state.operacoes:
        st.markdown(
            '<p style="font-size:1.1rem; font-weight:700; color:#223040; margin-bottom:8px;">Operações Recentes</p>',
            unsafe_allow_html=True,
        )
        rows = []
        for op in reversed(st.session_state.operacoes[-20:]):
            rows.append({
                "Tomador": op.get("tomador", ""),
                "CNPJ": op.get("cnpj", ""),
                "Tipo": op.get("tipo_operacao", ""),
                "Volume": _fmt_brl(op.get("volume", 0)),
                "Status": op.get("status", ""),
                "Rating": op.get("rating", "—"),
                "Data": op.get("data_criacao", ""),
            })
        st.dataframe(rows, use_container_width=True, hide_index=True)
    else:
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        st.markdown(
            """
            <div style="background:#F2F4F6; border-radius:12px; padding:40px; text-align:center;">
                <p style="font-size:1.6rem; margin:0 0 8px 0;">Nenhuma operação registrada</p>
                <p style="color:#8B9197; margin:0;">Clique em <b>Nova Análise</b> no menu lateral para iniciar sua primeira análise de crédito.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # ── Pipeline rápido ───────────────────────────────────────────────────
    st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)
    st.markdown(
        """
        <div style="background:linear-gradient(90deg, #223040, #2a3d52); border-radius:12px; padding:24px 28px;">
            <p style="color:#FFFFFF; font-weight:700; font-size:1rem; margin:0 0 12px 0;">Como funciona</p>
            <div style="display:flex; gap:16px; flex-wrap:wrap;">
                <div style="flex:1; min-width:140px; background:rgba(255,255,255,0.08); border-radius:8px; padding:14px; text-align:center;">
                    <p style="color:#A8B4C0; font-size:0.7rem; text-transform:uppercase; letter-spacing:1px; margin:0 0 4px 0;">Passo 1</p>
                    <p style="color:#FFFFFF; font-weight:600; font-size:0.9rem; margin:0;">Upload de Docs</p>
                </div>
                <div style="flex:1; min-width:140px; background:rgba(255,255,255,0.08); border-radius:8px; padding:14px; text-align:center;">
                    <p style="color:#A8B4C0; font-size:0.7rem; text-transform:uppercase; letter-spacing:1px; margin:0 0 4px 0;">Passo 2</p>
                    <p style="color:#FFFFFF; font-weight:600; font-size:0.9rem; margin:0;">Extração via Sonnet</p>
                </div>
                <div style="flex:1; min-width:140px; background:rgba(255,255,255,0.08); border-radius:8px; padding:14px; text-align:center;">
                    <p style="color:#A8B4C0; font-size:0.7rem; text-transform:uppercase; letter-spacing:1px; margin:0 0 4px 0;">Passo 3</p>
                    <p style="color:#FFFFFF; font-weight:600; font-size:0.9rem; margin:0;">Análise via Opus</p>
                </div>
                <div style="flex:1; min-width:140px; background:rgba(255,255,255,0.08); border-radius:8px; padding:14px; text-align:center;">
                    <p style="color:#A8B4C0; font-size:0.7rem; text-transform:uppercase; letter-spacing:1px; margin:0 0 4px 0;">Passo 4</p>
                    <p style="color:#FFFFFF; font-weight:600; font-size:0.9rem; margin:0;">MAC .docx Pronto</p>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def page_nova_analise():
    st.markdown(
        """
        <div style="margin-bottom:20px;">
            <h1 style="color:#223040; font-size:1.8rem; font-weight:800; margin:0;">Nova Análise de Crédito</h1>
            <p style="color:#8B9197; font-size:0.95rem; margin:4px 0 0 0;">
                Preencha os dados da operação, faça upload dos documentos e gere o MAC automaticamente.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if not API_KEY_SET:
        st.warning(
            "**ANTHROPIC_API_KEY** não configurada. "
            "A extração e análise via Claude não funcionarão. "
            "Configure os Secrets no Streamlit Cloud."
        )

    if not MODULES_AVAILABLE:
        st.error(f"Erro ao importar módulos: {_IMPORT_ERROR}")
        return

    tab_upload, tab_extracao, tab_analise, tab_mac = st.tabs([
        "1 - Upload & Dados",
        "2 - Extração",
        "3 - Análise",
        "4 - Documentos",
    ])

    # ------------------------------------------------------------------
    # TAB 1 — Upload & Dados
    # ------------------------------------------------------------------
    with tab_upload:
        st.markdown(
            '<p style="font-size:1.1rem; font-weight:700; color:#223040; margin-bottom:4px;">Parâmetros da Operação</p>',
            unsafe_allow_html=True,
        )
        st.caption("Preencha os campos abaixo com os dados do tomador e da operação.")

        _fc = st.session_state.get("form_counter", 0)
        with st.form(f"form_parametros_{_fc}", clear_on_submit=False):
            col_a, col_b = st.columns(2)

            with col_a:
                tomador = st.text_input("Tomador *", value=st.session_state.current_op.get("tomador", "") if st.session_state.current_op else "")
                cnpj = st.text_input("CNPJ", placeholder="00.000.000/0001-00", value=st.session_state.current_op.get("cnpj", "") if st.session_state.current_op else "")
                tipo_operacao = st.selectbox(
                    "Tipo de Operação",
                    ["CRI", "CRA", "CPR-F", "SLB", "NC/CCB", "FIDC", "Fiagro", "Debenture",
                     "Compra de Estoque", "Precatorios", "NPL", "SCP"],
                    index=0,
                )
                volume = st.number_input("Volume (R$)", min_value=0.0, format="%.2f", value=st.session_state.current_op.get("volume", 0.0) if st.session_state.current_op else 0.0)
                prazo_meses = st.number_input("Prazo (meses)", min_value=1, value=st.session_state.current_op.get("prazo_meses", 12) if st.session_state.current_op else 12)

            with col_b:
                taxa = st.text_input("Taxa", placeholder="CDI+4%", value=st.session_state.current_op.get("taxa", "") if st.session_state.current_op else "")
                amortizacao = st.selectbox("Amortização", ["SAC", "Price", "Bullet"])
                garantias_text = st.text_area(
                    "Garantias (uma por linha)",
                    value=st.session_state.current_op.get("garantias_text", "") if st.session_state.current_op else "",
                    help="Tipos: Real, Fiduciária, Aval, Coobrigação, Alienação, Penhor",
                )
                tipo_captacao = st.selectbox("Tipo de Captação", ["Captação PJ", "Captação PF"])
                instrumento = st.text_input("Instrumento", placeholder="NC / CCB", value=st.session_state.current_op.get("instrumento", "") if st.session_state.current_op else "")

            socio_responsavel = st.selectbox(
                "Sócio Responsável",
                ["Danilo Salasar", "Renato Barison", "Luiz Roberto Evangelista"],
            )

            submitted = st.form_submit_button("Salvar e Prosseguir", use_container_width=True)

        if submitted:
            if not tomador.strip():
                st.error("O campo **Tomador** é obrigatório.")
            elif cnpj.strip() and MODULES_AVAILABLE and not validate_cnpj(cnpj.strip()):
                st.warning("CNPJ inválido — verifique os dígitos. Parâmetros salvos mesmo assim.")
                # Still allow saving with invalid CNPJ (might be placeholder)
                garantias_list = [g.strip() for g in garantias_text.strip().split("\n") if g.strip()]
                op = {
                    "tomador": tomador.strip(),
                    "cnpj": cnpj.strip(),
                    "tipo_operacao": tipo_operacao,
                    "volume": volume,
                    "prazo_meses": prazo_meses,
                    "taxa": taxa.strip(),
                    "amortizacao": amortizacao,
                    "garantias": garantias_list,
                    "garantias_text": garantias_text,
                    "tipo_captacao": tipo_captacao,
                    "instrumento": instrumento.strip(),
                    "socio_responsavel": socio_responsavel,
                    "status": "Em Andamento",
                    "rating": "—",
                    "data_criacao": datetime.now().strftime("%Y-%m-%d %H:%M"),
                }
                st.session_state.current_op = op
                _upsert_operacao(op)
                _save_cache()
            else:
                garantias_list = [g.strip() for g in garantias_text.strip().split("\n") if g.strip()]

                op = {
                    "tomador": tomador.strip(),
                    "cnpj": cnpj.strip(),
                    "tipo_operacao": tipo_operacao,
                    "volume": volume,
                    "prazo_meses": prazo_meses,
                    "taxa": taxa.strip(),
                    "amortizacao": amortizacao,
                    "garantias": garantias_list,
                    "garantias_text": garantias_text,
                    "tipo_captacao": tipo_captacao,
                    "instrumento": instrumento.strip(),
                    "socio_responsavel": socio_responsavel,
                    "status": "Em Andamento",
                    "rating": "—",
                    "data_criacao": datetime.now().strftime("%Y-%m-%d %H:%M"),
                }
                st.session_state.current_op = op
                _upsert_operacao(op)
                _save_cache()
                st.success("Parâmetros salvos com sucesso.")

        st.markdown("---")
        st.markdown("### Upload de Documentos")

        uploaded = st.file_uploader(
            "Arraste ou selecione os documentos",
            accept_multiple_files=True,
            type=["pdf", "xlsx", "xls", "png", "jpg", "jpeg", "docx", "pptx"],
        )

        if uploaded:
            st.session_state.uploaded_files = []
            for f in uploaded:
                file_bytes = f.read()
                st.session_state.uploaded_files.append({
                    "name": f.name,
                    "size": len(file_bytes),
                    "bytes": file_bytes,
                })

            st.markdown("**Arquivos carregados:**")
            for uf in st.session_state.uploaded_files:
                size_kb = uf["size"] / 1024
                st.markdown(f"- **{uf['name']}** — {size_kb:,.1f} KB")

    # ------------------------------------------------------------------
    # TAB 2 — Extração
    # ------------------------------------------------------------------
    with tab_extracao:
        st.markdown("### Extração de Dados")

        if not st.session_state.uploaded_files:
            st.info("Nenhum arquivo carregado. Faça o upload na aba anterior.")
        else:
            st.markdown(f"**{len(st.session_state.uploaded_files)}** arquivo(s) disponível(is) para extração.")
            for uf in st.session_state.uploaded_files:
                st.markdown(f"- {uf['name']}")

            if st.button("Iniciar Extração", use_container_width=True, type="primary"):
                if not API_KEY_SET:
                    st.error("ANTHROPIC_API_KEY não configurada. Não é possível realizar a extração.")
                else:
                    total_files = len(st.session_state.uploaded_files)
                    # Progress container with colored status
                    progress_bar = st.progress(0, text=f"🔄 Preparando extração de {total_files} arquivos...")
                    status_container = st.container()
                    status_text = status_container.empty()
                    stats_text = status_container.empty()
                    errors: list[str] = []
                    _extraction_stats = {"ok": 0, "cache": 0, "error": 0, "current": ""}

                    # Preparar lista de arquivos para processamento paralelo
                    files_list = [(uf["bytes"], uf["name"]) for uf in st.session_state.uploaded_files]

                    def _extraction_progress(filename, idx, total, result):
                        has_err = "error" in result.get("dados", {})
                        cached = result.get("_from_cache", False)
                        pct = idx / total

                        if cached:
                            _extraction_stats["cache"] += 1
                        elif has_err:
                            _extraction_stats["error"] += 1
                        else:
                            _extraction_stats["ok"] += 1

                        if has_err:
                            errors.append(f"{filename}: {result['dados']['error']}")

                        # Color-coded progress text
                        if pct < 0.33:
                            phase = "🟡 Classificando e extraindo"
                        elif pct < 0.66:
                            phase = "🟠 Processando documentos"
                        else:
                            phase = "🟢 Finalizando"

                        # Truncar nome do arquivo se muito longo
                        fname_short = filename if len(filename) <= 35 else filename[:32] + "..."
                        progress_bar.progress(pct, text=f"{phase} — **{fname_short}** ({idx}/{total})")

                        # Stats line
                        ok = _extraction_stats["ok"]
                        cache = _extraction_stats["cache"]
                        err = _extraction_stats["error"]
                        parts = []
                        if ok:
                            parts.append(f"✅ {ok} extraído(s)")
                        if cache:
                            parts.append(f"📦 {cache} do cache")
                        if err:
                            parts.append(f"❌ {err} erro(s)")
                        remaining = total - idx
                        est_seconds = remaining * 4  # ~4s per file estimate
                        est_min = est_seconds // 60
                        est_sec = est_seconds % 60
                        time_str = f"{est_min}m{est_sec:02d}s" if est_min > 0 else f"{est_sec}s"
                        stats_text.caption(f"{' · '.join(parts)} · ⏱️ ~{time_str} restante(s)")

                    results = process_files_parallel(
                        files_list,
                        max_workers=3,
                        progress_callback=_extraction_progress,
                    )

                    # Detect missing files (silently failed / timed out)
                    missing_files = [fn for _, fn in files_list if fn not in results]
                    for mf in missing_files:
                        results[mf] = {
                            "classificacao": {"tipo": "outro", "confianca": 0.0, "descricao": "", "error": "Arquivo não processado (timeout ou erro silencioso)"},
                            "dados": {"error": f"Arquivo não processado: {mf}"},
                        }
                        errors.append(f"{mf}: não processado (timeout ou erro)")

                    progress_bar.progress(1.0, text=f"✅ Extração concluída — {len(results)}/{total_files} arquivo(s)")
                    status_text.empty()
                    stats_text.empty()
                    st.session_state.extracted_data = results
                    _save_cache()

                    # Resumo final — contagem clara
                    cached_count = sum(1 for r in results.values() if r.get("_from_cache"))
                    error_count = sum(1 for r in results.values() if "error" in r.get("dados", {}))
                    new_count = len(results) - cached_count - error_count
                    st.success(f"**{new_count}** novo(s) · **{cached_count}** do cache · **{error_count}** erro(s) · **{len(results)}** total")

                    if errors:
                        with st.expander(f"⚠️ {len(errors)} erro(s) — clique para ver"):
                            for err in errors:
                                st.error(f"❌ {err}")

            # Show extraction results
            if st.session_state.extracted_data:
                st.markdown("---")
                st.markdown("### Resultados da Extração")

                for fname, result in st.session_state.extracted_data.items():
                    classificacao = result.get("classificacao", {})
                    dados = result.get("dados", {})
                    tipo = classificacao.get("tipo", "desconhecido")
                    confianca = classificacao.get("confianca", 0.0)
                    descricao = classificacao.get("descricao", "")
                    has_error = "error" in dados

                    status_icon = "❌ Erro" if has_error else "✅ Extraído"
                    conf_color = _confidence_color(confianca)

                    with st.expander(f"{status_icon} — **{fname}**", expanded=not has_error):
                        col1, col2 = st.columns([1, 2])
                        with col1:
                            st.markdown(f"**Tipo detectado:** `{tipo}`")
                            st.markdown(
                                f"**Confiança:** :{conf_color}[{confianca:.0%}]"
                            )
                            if descricao:
                                st.caption(descricao)
                        with col2:
                            if has_error:
                                st.error(f"Erro: {dados['error']}")
                            else:
                                st.json(dados, expanded=False)

                # Checklist automático de tipos documentais
                st.markdown("---")
                st.markdown("### Checklist Automático de Documentos")

                tipos_esperados = {
                    "balanco": "Balanço Patrimonial",
                    "dre": "DRE",
                    "balancete": "Balancete",
                    "matricula": "Matrícula de Imóvel",
                    "contrato": "Contrato",
                    "certidao": "Certidão",
                    "ccir_car": "CCIR / CAR",
                }
                tipos_detectados = _detected_doc_types()

                col_ok, col_pend = st.columns(2)
                with col_ok:
                    st.markdown("**Documentos encontrados:**")
                    encontrados = [
                        label for tipo, label in tipos_esperados.items() if tipo in tipos_detectados
                    ]
                    if encontrados:
                        for item in encontrados:
                            st.markdown(f"- ✅ {item}")
                    else:
                        st.caption("Nenhum tipo padrão detectado.")

                with col_pend:
                    st.markdown("**Documentos não encontrados:**")
                    faltantes = [
                        label for tipo, label in tipos_esperados.items() if tipo not in tipos_detectados
                    ]
                    if faltantes:
                        for item in faltantes:
                            st.markdown(f"- ⚠️ {item}")
                    else:
                        st.markdown("Todos os tipos padrão foram detectados.")

    # ------------------------------------------------------------------
    # TAB 3 — Análise
    # ------------------------------------------------------------------
    with tab_analise:
        st.markdown("### Análise de Crédito")

        if not st.session_state.current_op:
            st.info("Preencha os parâmetros da operação na aba **Upload & Dados** antes de gerar a análise.")
        elif not st.session_state.extracted_data:
            st.info("Execute a extração de documentos na aba **Extração** antes de gerar a análise.")
        else:
            op = st.session_state.current_op
            st.markdown(f"**Tomador:** {op['tomador']}  |  **Tipo:** {op['tipo_operacao']}  |  **Volume:** {_fmt_brl(op['volume'])}")

            # Build dados_extraidos — accumulate multiple docs of same type
            dados_para_analise: dict = {}
            total_docs = 0
            for fname, result in st.session_state.extracted_data.items():
                classificacao = result.get("classificacao", {})
                dados = result.get("dados", {})
                tipo = classificacao.get("tipo", "outro")
                if "error" not in dados:
                    total_docs += 1
                    if tipo in dados_para_analise:
                        # Append with filename prefix to distinguish
                        key = f"{tipo}_{total_docs}"
                        dados_para_analise[key] = dados
                    else:
                        dados_para_analise[tipo] = dados

            st.markdown(f"Dados disponíveis: **{total_docs}** documento(s) em **{len(dados_para_analise)}** entrada(s).")

            if st.button("Gerar Análise de Crédito", use_container_width=True, type="primary"):
                if not API_KEY_SET:
                    st.error("ANTHROPIC_API_KEY não configurada. Não é possível realizar a análise.")
                else:
                    # Analysis progress with phases
                    analysis_progress = st.progress(0, text="🔄 Preparando análise...")
                    status_container = st.empty()
                    _analysis_phases = {
                        "Setor detectado": 0.10,
                        "Enviando": 0.15,
                        "Conectando": 0.20,
                        "Recebendo resposta": 0.40,
                        "Resposta completa": 0.75,
                        "Gerando matching": 0.85,
                        "Enriquecendo": 0.90,
                    }
                    def _update_status(msg):
                        # Detect phase from message and update progress
                        pct = 0.10
                        phase_icon = "🟡"
                        for keyword, phase_pct in _analysis_phases.items():
                            if keyword.lower() in msg.lower():
                                pct = phase_pct
                                break
                        if pct < 0.25:
                            phase_icon = "🟡"
                        elif pct < 0.60:
                            phase_icon = "🟠"
                        else:
                            phase_icon = "🟢"
                        # Truncar msg se muito longa
                        msg_short = msg if len(msg) <= 80 else msg[:77] + "..."
                        analysis_progress.progress(pct, text=f"{phase_icon} {msg_short}")
                        status_container.caption(f"⏳ Claude Opus está analisando {total_docs} documento(s)...")

                    with st.spinner("Analisando com Claude Opus..."):
                        try:
                            # KYC enrichment — busca dados públicos do CNPJ
                            cnpj = op.get("cnpj", "")
                            if cnpj and cnpj != "N/I":
                                dados_para_analise = dict(dados_para_analise)  # copy
                                enrich_analysis_data(cnpj, dados_para_analise, status_callback=_update_status)

                            # Grupo Econômico — detecta CPFs e CNPJs em todos os docs
                            _update_status("🔍 Mapeando grupo econômico nos documentos...")
                            cpf_cnpj_input = op.get("cnpj", "") or ""
                            grupo = extract_grupo_economico(
                                st.session_state.extracted_data,
                                cpf_cnpj_principal=cpf_cnpj_input,
                            )
                            if grupo["total_membros"] > 0:
                                dados_para_analise["grupo_economico"] = grupo
                                _update_status(
                                    f"🔍 Grupo: {len(grupo['cpfs'])} CPF(s) + {len(grupo['cnpjs'])} CNPJ(s)"
                                )

                            # Dados Fazenda — consulta ambiental agro (CAR, NDVI, embargos)
                            tipo_op = op.get("tipo_operacao", "")
                            is_agro = tipo_op in ("CRA", "CPR-F", "SLB", "Fiagro") or any(
                                kw in str(dados_para_analise).lower()
                                for kw in ("rural", "safra", "hectare", "soja", "milho", "agro", "car ")
                            )
                            if is_agro:
                                try:
                                    _update_status("🌱 Detectando CARs nos documentos...")
                                    car_codes = extract_car_codes(st.session_state.extracted_data)
                                    df_client = get_df_client()
                                    if car_codes and df_client:
                                        _update_status(f"🌱 Consultando Dados Fazenda — {len(car_codes)} CAR(s)...")
                                        df_resultado = df_client.consulta_grupo(car_codes)
                                        dados_para_analise["dados_fazenda"] = df_resultado
                                        # Cruzamento SIGEF
                                        cruzamento = df_client.cruzar_grupo_sigef(car_codes)
                                        dados_para_analise["cruzamento_sigef"] = cruzamento
                                        _update_status(f"🌱 Dados Fazenda: {df_resultado.get('total_propriedades', 0)} prop. · Score: {df_resultado.get('score_ambiental_grupo', 'N/D')}")
                                    elif car_codes:
                                        dados_para_analise["car_codes_detectados"] = car_codes
                                        _update_status(f"⚠ {len(car_codes)} CAR(s) detectado(s) mas Dados Fazenda não configurado")
                                except Exception as e:
                                    _update_status(f"⚠ Dados Fazenda: {str(e)[:50]}")

                            analise = analyze_credit(dados_para_analise, op, status_callback=_update_status)
                            st.session_state.analysis = analise

                            # Update operation record
                            rating_final = analise.get("rating_final", {})
                            op["rating"] = rating_final.get("nota", "—")
                            op["parecer"] = rating_final.get("parecer", "—")
                            op["status"] = "Concluída"
                            st.session_state.current_op = op
                            _save_cache()

                            analysis_progress.progress(1.0, text="✅ Análise concluída!")
                            status_container.empty()
                            _save_to_history(op, analise, st.session_state.extracted_data)
                            st.success("Análise concluída e salva no histórico.")
                        except Exception as e:
                            analysis_progress.empty()
                            status_container.empty()
                            st.error(f"Erro durante a análise: {e}")
                            import traceback
                            with st.expander("Detalhes do erro"):
                                st.code(traceback.format_exc(), language=None)

            # Display analysis results
            if st.session_state.analysis:
                analise = st.session_state.analysis
                rating_final = analise.get("rating_final", {})
                kpis = analise.get("kpis", {})
                nota = rating_final.get("nota", "N/A")
                parecer = rating_final.get("parecer", "N/A")

                st.markdown("---")

                # Rating + Parecer header
                col_rating, col_parecer = st.columns([1, 2])
                with col_rating:
                    st.markdown(
                        f'<div class="rating-badge {_rating_class(nota)}">{nota}</div>',
                        unsafe_allow_html=True,
                    )
                    st.caption("Rating Final")
                with col_parecer:
                    st.markdown(
                        f'<span class="{_parecer_class(parecer)}" style="font-size:1.5rem;">{parecer}</span>',
                        unsafe_allow_html=True,
                    )

                # KPIs row
                st.markdown("---")
                st.markdown("### Indicadores-Chave")
                k1, k2, k3, k4, k5, k6 = st.columns(6)
                k1.metric("Receita Líq.", _fmt_brl(kpis.get("receita_liquida", 0), compact=True))
                k2.metric("EBITDA", _fmt_brl(kpis.get("ebitda", 0), compact=True))

                margem = _to_num(kpis.get("margem_ebitda", 0))
                margem_str = f"{margem:.1%}" if 0 < margem <= 1 else f"{margem:.1f}%" if margem else str(kpis.get("margem_ebitda", "N/D"))
                k3.metric("Margem EBITDA", margem_str)

                div_ebitda = _to_num(kpis.get("divida_liquida_ebitda", 0))
                k4.metric("Dív.Líq./EBITDA", f"{div_ebitda:.2f}x" if div_ebitda else str(kpis.get("divida_liquida_ebitda", "N/D")))

                ltv = _to_num(kpis.get("ltv", 0))
                ltv_str = f"{ltv:.1%}" if 0 < ltv <= 1 else f"{ltv:.1f}%" if ltv else str(kpis.get("ltv", "N/D"))
                k5.metric("LTV", ltv_str)

                dscr = _to_num(kpis.get("dscr", 0))
                k6.metric("DSCR", f"{dscr:.2f}x" if dscr else str(kpis.get("dscr", "N/D")))

                # 10 analysis sections
                st.markdown("---")
                st.markdown("### Seções da Análise")

                secoes = [
                    ("tomador", "1. Tomador"),
                    ("patrimonio", "2. Patrimônio"),
                    ("producao", "3. Produção"),
                    ("capital", "4. Estrutura de Capital"),
                    ("operacao", "5. Operação"),
                    ("pagamento", "6. Capacidade de Pagamento"),
                    ("onus", "7. Ônus e Restrições"),
                    ("riscos", "8. Riscos"),
                    ("covenants", "9. Covenants"),
                    ("cronograma", "10. Cronograma"),
                ]

                for key, label in secoes:
                    secao_data = analise.get(key, {})
                    rating_secao = secao_data.get("rating_secao", "N/A")
                    flags = secao_data.get("flags", [])

                    with st.expander(f"**{label}** — _{rating_secao}_"):
                        # Show all string fields
                        for campo, valor in secao_data.items():
                            if campo in ("rating_secao", "flags"):
                                continue
                            if isinstance(valor, str) and valor:
                                st.markdown(f"**{campo.replace('_', ' ').title()}:**")
                                st.markdown(valor)
                            elif isinstance(valor, dict):
                                st.markdown(f"**{campo.replace('_', ' ').title()}:**")
                                st.json(valor, expanded=False)
                            elif isinstance(valor, list) and valor:
                                st.markdown(f"**{campo.replace('_', ' ').title()}:**")
                                if isinstance(valor[0], dict):
                                    st.dataframe(valor, use_container_width=True, hide_index=True)
                                else:
                                    for item in valor:
                                        st.markdown(f"- {item}")

                        if flags:
                            st.markdown("**Flags de Atenção:**")
                            for flag in flags:
                                st.warning(flag)

                # Checklist de Lacunas
                checklist = analise.get("checklist_lacunas", {})
                docs_faltantes = checklist.get("documentos_faltantes", [])
                info_pendentes = checklist.get("informacoes_pendentes", [])
                total_pend = checklist.get("total_pendencias", 0)
                total_crit = checklist.get("total_criticas", 0)

                if docs_faltantes or info_pendentes:
                    st.markdown("---")
                    st.markdown(f"### Checklist de Lacunas — {total_pend} pendência(s), {total_crit} crítica(s)")

                    if docs_faltantes:
                        st.markdown("#### Documentos Faltantes")
                        df_docs = []
                        for doc in docs_faltantes:
                            df_docs.append({
                                "Item": doc.get("item", ""),
                                "Criticidade": doc.get("criticidade", ""),
                                "Motivo": doc.get("motivo", ""),
                            })
                        st.dataframe(df_docs, use_container_width=True, hide_index=True)

                    if info_pendentes:
                        st.markdown("#### Informações Pendentes")
                        df_info = []
                        for info in info_pendentes:
                            df_info.append({
                                "Item": info.get("item", ""),
                                "Criticidade": info.get("criticidade", ""),
                                "Motivo": info.get("motivo", ""),
                            })
                        st.dataframe(df_info, use_container_width=True, hide_index=True)

                # Investor Matching
                investor_data = analise.get("investor_matching", {})
                investidores = investor_data.get("investidores_sugeridos", [])
                if investidores:
                    st.markdown("---")
                    setor_det = investor_data.get("setor_detectado", "N/A")
                    st.markdown(f"### Investidores Sugeridos — Setor: {setor_det.title()}")

                    # Count by source
                    fontes = {}
                    for inv in investidores:
                        f = inv.get("fonte", "Outro")
                        fontes[f] = fontes.get(f, 0) + 1
                    fonte_str = " | ".join(f"{k}: {v}" for k, v in fontes.items())
                    st.caption(f"Ranking unificado: {fonte_str} ({len(investidores)} total)")

                    inv_rows = []
                    for inv in investidores:
                        inv_rows.append({
                            "Score": f"{inv.get('score', 0):.0f}",
                            "Investidor": inv["nome"],
                            "Fonte": inv.get("fonte", ""),
                            "Perfil": inv.get("perfil", ""),
                            "Motivos": " | ".join(inv.get("motivos", [])),
                        })
                    st.dataframe(inv_rows, use_container_width=True, hide_index=True)

                # Botão Nova Análise — fora do bloco de geração
                st.markdown("---")
                if st.button("Iniciar Nova Análise", use_container_width=True, key="clear_after_analysis"):
                    _clear_session()
                    st.rerun()

    # ------------------------------------------------------------------
    # TAB 4 — Documentos
    # ------------------------------------------------------------------
    with tab_mac:
        st.markdown("### Geração de Documentos")

        if not st.session_state.analysis:
            st.info("Execute a análise de crédito na aba **Análise** antes de gerar documentos.")
        else:
            analise = st.session_state.analysis
            op = st.session_state.current_op

            st.markdown("#### Resumo da Análise")
            summary = get_analysis_summary(analise)
            st.code(summary, language=None)

            # --- Gerar TODOS em paralelo ---
            if st.button("⚡ Gerar Todos os Documentos", use_container_width=True, type="primary"):
                import concurrent.futures
                tomador_clean = (op.get("tomador", "operacao") or "operacao").replace(" ", "_").replace("/", "-")
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

                xl_name = f"AnaliseTecnica_{tomador_clean}_{timestamp}.xlsx"
                mac_name = f"MAC_{tomador_clean}_{timestamp}.docx"
                teaser_name = f"Teaser_{tomador_clean}_{timestamp}.pptx"
                xl_path = str(OUTPUT_DIR / xl_name)
                mac_path = str(OUTPUT_DIR / mac_name)
                teaser_path = str(OUTPUT_DIR / teaser_name)

                hist = _list_history()
                results = {}
                progress_bar = st.progress(0, text="Gerando 3 documentos em paralelo...")

                def _gen_excel():
                    generate_excel(analise, op, xl_path, hist if hist else None)
                    return xl_path

                def _gen_mac():
                    return generate_mac(analise, op, mac_path)

                def _gen_teaser():
                    generate_teaser(analise, op, teaser_path)
                    return teaser_path

                with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
                    futures = {
                        executor.submit(_gen_excel): "excel",
                        executor.submit(_gen_mac): "mac",
                        executor.submit(_gen_teaser): "teaser",
                    }
                    done_count = 0
                    for future in concurrent.futures.as_completed(futures):
                        name = futures[future]
                        done_count += 1
                        try:
                            results[name] = future.result(timeout=120)
                            progress_bar.progress(done_count / 3, text=f"✅ {name.upper()} concluído ({done_count}/3)")
                        except Exception as e:
                            results[name] = None
                            st.error(f"Erro ao gerar {name.upper()}: {e}")
                            progress_bar.progress(done_count / 3, text=f"❌ {name.upper()} erro ({done_count}/3)")

                progress_bar.progress(1.0, text="✅ Documentos prontos!")

                # Download buttons
                col_dl1, col_dl2, col_dl3 = st.columns(3)

                if results.get("excel"):
                    with col_dl1:
                        with open(xl_path, "rb") as f:
                            st.download_button("📊 Baixar Excel", data=f.read(), file_name=xl_name,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True)

                if results.get("mac"):
                    with col_dl2:
                        with open(mac_path, "rb") as f:
                            st.download_button("📄 Baixar MAC", data=f.read(), file_name=mac_name,
                                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                                use_container_width=True)

                if results.get("teaser"):
                    with col_dl3:
                        with open(teaser_path, "rb") as f:
                            st.download_button("📑 Baixar Teaser", data=f.read(), file_name=teaser_name,
                                mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                                use_container_width=True)

                op["status"] = "Concluída"
                st.session_state.current_op = op

            # --- Botões individuais ---
            st.markdown("---")
            col_xl, col_mac_btn, col_teaser_btn = st.columns(3)

            with col_xl:
                if st.button("Análise Técnica (.xlsx)", use_container_width=True):
                    try:
                        tomador_clean = (op.get("tomador", "operacao") or "operacao").replace(" ", "_").replace("/", "-")
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        xl_name = f"AnaliseTecnica_{tomador_clean}_{timestamp}.xlsx"
                        xl_path = str(OUTPUT_DIR / xl_name)

                        with st.spinner("Gerando Planilha..."):
                            hist = _list_history()
                            generate_excel(analise, op, xl_path, hist if hist else None)

                        with open(xl_path, "rb") as f:
                            st.download_button("📊 Baixar Excel", data=f.read(), file_name=xl_name,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True)
                    except Exception as e:
                        st.error(f"Erro ao gerar planilha: {e}")

            with col_mac_btn:
                if st.button("MAC (.docx)", use_container_width=True):
                    try:
                        tomador_clean = (op.get("tomador", "operacao") or "operacao").replace(" ", "_").replace("/", "-")
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        filename = f"MAC_{tomador_clean}_{timestamp}.docx"
                        output_path = str(OUTPUT_DIR / filename)

                        with st.spinner("Gerando MAC..."):
                            generated_path = generate_mac(analise, op, output_path)

                        with open(generated_path, "rb") as f:
                            st.download_button("📄 Baixar MAC", data=f.read(), file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                                use_container_width=True)
                        op["status"] = "Concluída"
                        st.session_state.current_op = op
                    except Exception as e:
                        st.error(f"Erro ao gerar MAC: {e}")

            with col_teaser_btn:
                if st.button("Teaser (.pptx)", use_container_width=True):
                    try:
                        tomador_clean = (op.get("tomador", "operacao") or "operacao").replace(" ", "_").replace("/", "-")
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        teaser_name = f"Teaser_{tomador_clean}_{timestamp}.pptx"
                        teaser_path = str(OUTPUT_DIR / teaser_name)

                        with st.spinner("Gerando Teaser..."):
                            import logging
                            logging.warning(f"[TEASER] analise keys: {list(analise.keys()) if isinstance(analise, dict) else type(analise)}")
                            logging.warning(f"[TEASER] op keys: {list(op.keys()) if isinstance(op, dict) else type(op)}")
                            if isinstance(analise, dict) and 'tomador' in analise and isinstance(analise['tomador'], dict):
                                logging.warning(f"[TEASER] tomador keys: {list(analise['tomador'].keys())}")
                            generate_teaser(analise, op, teaser_path)

                        with open(teaser_path, "rb") as f:
                            st.download_button("📑 Baixar Teaser", data=f.read(), file_name=teaser_name,
                                mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                                use_container_width=True)
                    except Exception as e:
                        st.error(f"Erro ao gerar Teaser: {e}")


def page_historico():
    st.markdown(
        """
        <div style="margin-bottom:20px;">
            <h2 style="color:#223040; margin:0;">Historico de Analises</h2>
            <p style="color:#8B9197; margin:4px 0 0 0;">Analises concluidas salvas automaticamente</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    historico = _list_history()

    if not historico:
        st.info("Nenhuma analise salva ainda. Execute uma analise na aba **Nova Analise** para comecar.")
        return

    st.markdown(f"**{len(historico)}** analise(s) salva(s).")
    st.markdown("---")

    for i, item in enumerate(historico):
        op = item.get("operacao", {})
        analise = item.get("analise", {})
        data = item.get("data_analise", "—")
        rating = analise.get("rating_final", {})
        nota = rating.get("nota", "—")
        parecer = rating.get("parecer", "—")
        tomador = op.get("tomador", "N/A")
        tipo = op.get("tipo_operacao", "N/A")
        volume = op.get("volume", 0)
        filename = item.get("_filename", "")

        # Color for rating badge (AAA-D scale)
        cor_nota = {"AAA": "#0D5C2F", "AA": "#1E6B42", "A": "#2E7D4F", "BBB": "#223040",
                    "BB": "#5C6B08", "B": "#7D6608", "C": "#E65100", "D": "#922B21"}.get(nota, "#8B9197")

        with st.expander(
            f"**{tomador}** — {tipo} — {_fmt_brl(volume)} | Rating: {nota} | {data}",
            expanded=False,
        ):
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Rating", nota)
            col2.metric("Parecer", parecer)
            col3.metric("DSCR", f"{analise.get('kpis', {}).get('dscr', 0):.2f}x")
            col4.metric("LTV", f"{analise.get('kpis', {}).get('ltv', 0):.1%}" if isinstance(analise.get('kpis', {}).get('ltv'), (int, float)) and analise.get('kpis', {}).get('ltv', 0) <= 1 else f"{analise.get('kpis', {}).get('ltv', 0):.1f}%")

            st.markdown("---")

            # Justificativa
            justificativa = rating.get("justificativa", "—")
            st.markdown(f"**Justificativa:** {justificativa}")

            # Recomendacoes
            recs = rating.get("recomendacoes", [])
            if recs:
                st.markdown("**Recomendacoes:**")
                for r in recs:
                    st.markdown(f"- {r}")

            # Flags
            todas_flags = []
            for secao in ["tomador", "patrimonio", "producao", "capital", "operacao", "pagamento", "onus", "riscos", "covenants", "cronograma"]:
                for flag in analise.get(secao, {}).get("flags", []):
                    todas_flags.append(f"[{secao.title()}] {flag}")
            if todas_flags:
                st.markdown(f"**Flags de Atencao ({len(todas_flags)}):**")
                for fl in todas_flags:
                    st.markdown(f"- {fl}")

            # Lacunas
            lacunas = analise.get("checklist_lacunas", {})
            docs_falt = lacunas.get("documentos_faltantes", [])
            info_pend = lacunas.get("informacoes_pendentes", [])
            if docs_falt or info_pend:
                st.markdown(f"**Pendencias:** {lacunas.get('total_pendencias', 0)} | Criticas: {lacunas.get('total_criticas', 0)}")
                for d in docs_falt:
                    st.markdown(f"- [{d.get('criticidade', '—')}] {d.get('item', '—')}")
                for p in info_pend:
                    st.markdown(f"- [{p.get('criticidade', '—')}] {p.get('item', '—')}")

            st.markdown("---")

            # Complementar Análise — upload additional docs, merge, re-analyze
            with st.container():
                st.markdown("**Complementar Análise**")
                complement_files = st.file_uploader(
                    "Upload de documentos complementares",
                    accept_multiple_files=True,
                    key=f"complement_{i}",
                    type=["pdf", "png", "jpg", "jpeg", "xlsx", "xls", "csv", "docx"],
                )
                if complement_files and st.button(
                    "Extrair e Re-analisar",
                    key=f"reanalyze_{i}",
                    use_container_width=True,
                    type="primary",
                ):
                    existing_extracted = item.get("extracted_data", {})
                    new_extracted = dict(existing_extracted)
                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    # Extract complement files
                    comp_files = [(uploaded.read(), uploaded.name) for uploaded in complement_files]

                    if len(comp_files) == 1:
                        # Single file — direct extraction (no thread overhead)
                        file_bytes, file_name = comp_files[0]
                        status_text.info(f"Extraindo {file_name}...")
                        progress_bar.progress(0.3, text=f"Extraindo {file_name}...")
                        result = process_file(file_bytes, file_name)
                        comp_results = {file_name: result}
                    else:
                        def _comp_progress(filename, idx, total, result):
                            progress_bar.progress(idx / total, text=f"Extraindo {filename}... ({idx}/{total})")
                            status_text.info(f"Processando {filename} ({idx}/{total})")
                        comp_results = process_files_parallel(comp_files, max_workers=3, progress_callback=_comp_progress)

                    new_extracted.update(comp_results)

                    progress_bar.progress(1.0)
                    status_text.info("Documentos extraídos. Atualizando análise (incremental)...")

                    # Build data from NEW docs only (for incremental analysis)
                    novos_dados = {}
                    for fname, result in comp_results.items():
                        classificacao = result.get("classificacao", {})
                        dados = result.get("dados", {})
                        tipo = classificacao.get("tipo", "outro")
                        if "error" not in dados:
                            novos_dados[tipo] = dados

                    analise_anterior = item.get("analise", {})

                    try:
                        def _update_status(msg):
                            status_text.info(f"⏳ {msg}")

                        nova_analise = analyze_incremental(
                            analise_anterior=analise_anterior,
                            novos_dados_extraidos=novos_dados,
                            parametros_operacao=op,
                            status_callback=_update_status,
                        )

                        # Update history file with new analysis
                        updated_payload = {
                            "operacao": op,
                            "analise": nova_analise,
                            "extracted_data": new_extracted,
                            "data_analise": datetime.now().strftime("%Y-%m-%d %H:%M"),
                        }
                        (HISTORY_DIR / filename).write_text(
                            json.dumps(updated_payload, ensure_ascii=False, indent=2)
                        )
                        _gh_save(filename, updated_payload)
                        status_text.empty()
                        progress_bar.empty()
                        st.success(f"Análise atualizada com {len(complement_files)} documento(s) (incremental).")
                        st.rerun()
                    except Exception as ex:
                        status_text.empty()
                        progress_bar.empty()
                        st.error(f"Erro na re-análise: {ex}")

            st.markdown("---")

            # Action buttons
            col_excel, col_mac, col_teaser, col_load, col_del = st.columns(5)

            with col_excel:
                if st.button("Análise Técnica (.xlsx)", key=f"excel_{i}", use_container_width=True, type="primary"):
                    try:
                        excel_path = OUTPUT_DIR / f"AnaliseTecnica_{tomador.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
                        # Pass historico for cross-reference
                        other_analyses = [h for h in historico if h.get("_filename") != filename]
                        generate_excel(analise, op, str(excel_path), other_analyses if other_analyses else None)
                        with open(excel_path, "rb") as f:
                            st.download_button(
                                label="Baixar Análise (.xlsx)",
                                data=f.read(),
                                file_name=excel_path.name,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"dl_excel_{i}",
                                use_container_width=True,
                            )
                        st.success("Planilha de Análise Técnica gerada.")
                    except Exception as e:
                        st.error(f"Erro ao gerar Excel: {e}")

            with col_mac:
                if st.button("MAC (.docx)", key=f"mac_{i}", use_container_width=True):
                    try:
                        mac_path = OUTPUT_DIR / f"MAC_{tomador.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.docx"
                        generate_mac(analise, op, str(mac_path))
                        with open(mac_path, "rb") as f:
                            st.download_button(
                                label="Baixar MAC (.docx)",
                                data=f.read(),
                                file_name=mac_path.name,
                                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                                key=f"dl_mac_{i}",
                                use_container_width=True,
                            )
                        st.success("MAC gerado com sucesso.")
                    except Exception as e:
                        st.error(f"Erro ao gerar MAC: {e}")

            with col_teaser:
                if st.button("Teaser (.pptx)", key=f"teaser_{i}", use_container_width=True):
                    try:
                        teaser_path = OUTPUT_DIR / f"Teaser_{tomador.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pptx"
                        generate_teaser(analise, op, str(teaser_path))
                        with open(teaser_path, "rb") as f:
                            st.download_button(
                                label="Baixar Teaser (.pptx)",
                                data=f.read(),
                                file_name=teaser_path.name,
                                mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                                key=f"dl_teaser_{i}",
                                use_container_width=True,
                            )
                        st.success("Teaser gerado com sucesso.")
                    except Exception as e:
                        st.error(f"Erro ao gerar Teaser: {e}")

            with col_load:
                if st.button("Carregar", key=f"load_{i}", use_container_width=True):
                    st.session_state.analysis = analise
                    st.session_state.current_op = op
                    if item.get("extracted_data"):
                        st.session_state.extracted_data = item["extracted_data"]
                    _save_cache()
                    st.success("Análise carregada na sessão.")
                    st.rerun()

            with col_del:
                if st.button("Excluir", key=f"del_{i}", use_container_width=True):
                    _delete_from_history(filename)
                    st.success("Analise excluida.")
                    st.rerun()


def page_checklist_dd():
    st.markdown(
        """
        <div style="margin-bottom:20px;">
            <h1 style="color:#223040; font-size:1.8rem; font-weight:800; margin:0;">Checklist Due Diligence</h1>
            <p style="color:#8B9197; font-size:0.95rem; margin:4px 0 0 0;">
                12 módulos &nbsp;·&nbsp; Cruzamento automático com documentos extraídos
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ── Tab layout: current checklist vs saved checklists ──
    tab_atual, tab_salvos = st.tabs(["Checklist Atual", "Checklists Salvos"])

    with tab_atual:
        # Initialize DD status from session state (base + product-specific)
        if not st.session_state.dd_status:
            for modulo, itens in DD_CHECKLIST_TEMPLATE.items():
                st.session_state.dd_status[modulo] = {}
                for item in itens:
                    st.session_state.dd_status[modulo][item] = "PENDENTE"

            # Add product-specific DD items based on operation type
            if st.session_state.current_op and MODULES_AVAILABLE:
                tipo_op = st.session_state.current_op.get("tipo_operacao", "")
                product_extras = PRODUCT_DD_EXTRAS.get(tipo_op, {})
                for modulo, itens in product_extras.items():
                    if modulo not in st.session_state.dd_status:
                        st.session_state.dd_status[modulo] = {}
                    for item in itens:
                        if item not in st.session_state.dd_status[modulo]:
                            st.session_state.dd_status[modulo][item] = "PENDENTE"

        # Try to load saved checklist for current client (only once per client)
        tomador = st.session_state.current_op.get("tomador", "") if st.session_state.current_op else ""
        if tomador and st.session_state.get("_checklist_loaded_for") != tomador:
            saved = _load_checklist(tomador)
            if saved and saved.get("dd_status"):
                st.session_state.dd_status = saved["dd_status"]
            st.session_state["_checklist_loaded_for"] = tomador

        # Auto-populate from extracted data
        auto_matches: list[str] = []
        if st.session_state.extracted_data:
            st.session_state.dd_status, auto_matches = _auto_populate_checklist(
                st.session_state.dd_status, st.session_state.extracted_data
            )

        if auto_matches:
            with st.expander(f"🔄 **{len(auto_matches)} documentos cruzados automaticamente**", expanded=False):
                for m in auto_matches:
                    st.markdown(f"- ✅ {m}")

        # Dashboard summary
        total_items = 0
        total_ok = 0
        total_pendente = 0
        total_desatualizado = 0

        for modulo, itens in st.session_state.dd_status.items():
            for item, status in itens.items():
                total_items += 1
                if status == "OK":
                    total_ok += 1
                elif status == "PENDENTE":
                    total_pendente += 1
                elif status == "DESATUALIZADO":
                    total_desatualizado += 1

        # KPI cards
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Total Itens", total_items)
        col2.metric("OK", total_ok)
        col3.metric("Pendente", total_pendente)
        col4.metric("Desatualizado", total_desatualizado)

        if total_items > 0:
            overall_pct = total_ok / total_items
            st.markdown("**Progresso Geral**")
            st.progress(overall_pct, text=f"{overall_pct:.0%} concluído ({total_ok}/{total_items})")

        st.markdown("---")

        # Current operation context
        if st.session_state.current_op:
            st.markdown(f"**Operação:** {tomador} — {st.session_state.current_op.get('tipo_operacao', 'N/A')}")
            st.markdown("---")

        # Module-by-module display (base + product-specific)
        status_options = ["OK", "PENDENTE", "DESATUALIZADO"]
        status_icons = {"OK": "✅", "PENDENTE": "⏳", "DESATUALIZADO": "⚠️"}

        # Build full module list: base template + any extras in dd_status
        all_modules: dict[str, list[str]] = {}
        for modulo, itens in DD_CHECKLIST_TEMPLATE.items():
            all_modules[modulo] = list(itens)
        for modulo, itens_dict in st.session_state.dd_status.items():
            if modulo not in all_modules:
                all_modules[modulo] = list(itens_dict.keys())

        for modulo, itens in all_modules.items():
            modulo_status = st.session_state.dd_status.get(modulo, {})
            ok_count = sum(1 for s in modulo_status.values() if s == "OK")
            total_mod = len(itens)
            pct = ok_count / total_mod if total_mod > 0 else 0

            with st.expander(f"**{modulo}** — {ok_count}/{total_mod} ({pct:.0%})", expanded=False):
                st.progress(pct)

                for item in itens:
                    current_status = st.session_state.dd_status.get(modulo, {}).get(item, "PENDENTE")
                    col_item, col_status = st.columns([3, 1])

                    with col_item:
                        icon = status_icons.get(current_status, "⏳")
                        st.markdown(f"{icon} {item}")

                    with col_status:
                        new_status = st.selectbox(
                            "Status",
                            status_options,
                            index=status_options.index(current_status),
                            key=f"dd_{modulo}_{item}",
                            label_visibility="collapsed",
                        )
                        if new_status != current_status:
                            st.session_state.dd_status[modulo][item] = new_status
                            # Auto-save on change
                            if tomador:
                                _save_checklist(tomador, st.session_state.dd_status, list(_detected_doc_types()))
                            st.rerun()

        # ── Action buttons ──
        st.markdown("---")
        col_save, col_excel = st.columns(2)

        with col_save:
            if st.button("💾 Salvar Checklist", use_container_width=True, key="save_checklist"):
                if tomador:
                    _save_checklist(tomador, st.session_state.dd_status, list(_detected_doc_types()))
                    st.success(f"Checklist salvo para **{tomador}**.")
                else:
                    st.warning("Preencha os dados da operação primeiro (aba Nova Análise).")

        with col_excel:
            if st.button("📊 Gerar Checklist Excel", use_container_width=True, key="gen_checklist_xlsx"):
                try:
                    label = tomador or "geral"
                    excel_bytes = _generate_checklist_excel(label, st.session_state.dd_status)
                    safe_name = label.replace(" ", "_").replace("/", "-")
                    st.download_button(
                        label="📥 Baixar Checklist (.xlsx)",
                        data=excel_bytes,
                        file_name=f"Checklist_DD_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                    # Also save to disk
                    if tomador:
                        _save_checklist(tomador, st.session_state.dd_status, list(_detected_doc_types()))
                except Exception as e:
                    st.error(f"Erro ao gerar Excel: {e}")

    # ── Tab: Saved checklists ──
    with tab_salvos:
        st.markdown("### Checklists Salvos por Cliente")
        checklists = _list_checklists()

        if not checklists:
            st.info("Nenhum checklist salvo ainda. Realize uma análise e salve o checklist.")
        else:
            for i, ck in enumerate(checklists):
                ck_tomador = ck.get("tomador", "Desconhecido")
                ck_updated = ck.get("updated_at", "—")
                ck_status = ck.get("dd_status", {})

                # Count stats
                ck_total = sum(len(v) for v in ck_status.values())
                ck_ok = sum(1 for itens in ck_status.values() for s in itens.values() if s == "OK")
                ck_pct = f"{ck_ok / ck_total * 100:.0f}%" if ck_total > 0 else "0%"

                with st.expander(f"**{ck_tomador}** — {ck_pct} concluído ({ck_ok}/{ck_total}) · Atualizado: {ck_updated}"):
                    # Summary per module
                    for modulo in DD_CHECKLIST_TEMPLATE:
                        mod_items = ck_status.get(modulo, {})
                        mod_ok = sum(1 for s in mod_items.values() if s == "OK")
                        mod_total = len(mod_items)
                        mod_pct = mod_ok / mod_total if mod_total > 0 else 0
                        bar_color = "🟢" if mod_pct >= 0.8 else ("🟡" if mod_pct >= 0.4 else "🔴")
                        st.markdown(f"{bar_color} **{modulo}** — {mod_ok}/{mod_total}")

                    col_load, col_xlsx, col_del = st.columns(3)
                    with col_load:
                        if st.button("Carregar", key=f"load_ck_{i}", use_container_width=True):
                            st.session_state.dd_status = ck_status
                            st.session_state["_checklist_loaded_for"] = ck_tomador
                            st.success(f"Checklist de **{ck_tomador}** carregado.")
                            st.rerun()

                    with col_xlsx:
                        if st.button("Excel", key=f"xlsx_ck_{i}", use_container_width=True):
                            try:
                                excel_bytes = _generate_checklist_excel(ck_tomador, ck_status)
                                safe_name = ck_tomador.replace(" ", "_").replace("/", "-")
                                st.download_button(
                                    label="📥 Baixar",
                                    data=excel_bytes,
                                    file_name=f"Checklist_DD_{safe_name}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True,
                                    key=f"dl_ck_{i}",
                                )
                            except Exception as e:
                                st.error(f"Erro: {e}")

                    with col_del:
                        if st.button("Excluir", key=f"del_ck_{i}", use_container_width=True):
                            path = CHECKLISTS_DIR / ck.get("_filename", "")
                            if path.exists():
                                path.unlink()
                            st.success(f"Checklist de **{ck_tomador}** excluído.")
                            st.rerun()


def page_investidores():
    st.markdown(
        """
        <div style="margin-bottom:20px;">
            <h1 style="color:#223040; font-size:1.8rem; font-weight:800; margin:0;">Investidores & Matching</h1>
            <p style="color:#8B9197; font-size:0.95rem; margin:4px 0 0 0;">
                Base de investidores calibrada pelo Pipeline ZYN &nbsp;·&nbsp; Matching por perfil de operacao
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if not MODULES_AVAILABLE:
        st.error(f"Erro ao importar modulos: {_IMPORT_ERROR}")
        return

    tab_matching, tab_base, tab_benchmarks = st.tabs(["Matching", "Base de Investidores", "Benchmarks Setoriais"])

    with tab_matching:
        st.markdown("### Simulador de Matching")
        st.caption("Simule quais investidores tem maior aderencia ao perfil da sua operacao.")

        col_a, col_b = st.columns(2)
        with col_a:
            sim_tipo = st.selectbox(
                "Tipo de Operacao",
                ["CRI", "CRA", "CPR-F", "SLB", "NC/CCB", "FIDC", "Fiagro", "Debenture",
                 "Compra de Estoque", "Precatorios", "NPL", "SCP"],
                key="sim_tipo",
            )
            sim_volume = st.number_input("Volume (R$)", min_value=0.0, value=30_000_000.0, format="%.0f", key="sim_vol")
        with col_b:
            sim_setor = st.selectbox(
                "Setor (auto-detectado se vazio)",
                ["(auto)", "agro", "imobiliario", "industria", "fidc", "special_sits"],
                key="sim_setor",
            )

        if st.button("Buscar Investidores", use_container_width=True, type="primary"):
            setor = None if sim_setor == "(auto)" else sim_setor
            results = match_investors(
                tipo_operacao=sim_tipo,
                volume=sim_volume,
                setor=setor,
                top_n=10,
            )

            if not results:
                st.warning("Nenhum investidor encontrado para este perfil.")
            else:
                st.success(f"**{len(results)}** investidores encontrados.")
                for inv in results:
                    score = inv.get("score", 0)
                    # Color based on score
                    if score >= 70:
                        bar_color = "#1E6B42"
                    elif score >= 40:
                        bar_color = "#223040"
                    else:
                        bar_color = "#7D6608"

                    with st.expander(f"**{inv['nome']}** — {score:.0f} pts | {inv.get('deals_pipeline', 0)} deals ZYN"):
                        st.markdown(f"**Perfil:** {inv.get('perfil', '')}")
                        st.markdown(f"**Instrumentos:** {', '.join(inv.get('instrumentos', []))}")
                        st.markdown(f"**Setores:** {', '.join(inv.get('setores', []))}")
                        st.markdown("**Motivos do match:**")
                        for m in inv.get("motivos", []):
                            st.markdown(f"- {m}")

        # If there's a current analysis, show auto-match
        if st.session_state.analysis:
            investor_data = st.session_state.analysis.get("investor_matching", {})
            investidores = investor_data.get("investidores_sugeridos", [])
            if investidores:
                st.markdown("---")
                st.markdown("### Matching da Analise Atual")
                op = st.session_state.current_op or {}
                st.markdown(f"**{op.get('tomador', 'N/A')}** — {op.get('tipo_operacao', '')} — {_fmt_brl(op.get('volume', 0))}")
                inv_rows = []
                for inv in investidores:
                    inv_rows.append({
                        "Score": f"{inv.get('score', 0):.0f}",
                        "Investidor": inv["nome"],
                        "Deals ZYN": inv.get("deals_pipeline", 0),
                        "Perfil": inv.get("perfil", ""),
                        "Motivos": " | ".join(inv.get("motivos", [])),
                    })
                st.dataframe(inv_rows, use_container_width=True, hide_index=True)

    with tab_base:
        st.markdown("### Base de Investidores ZYN")
        st.caption("43+ investidores institucionais mapeados a partir do Pipeline ZYN.")

        from modules.investor_matching import INVESTOR_DATABASE

        for inv in sorted(INVESTOR_DATABASE, key=lambda x: -x["deals_pipeline"]):
            with st.expander(f"**{inv['nome']}** — {inv['deals_pipeline']} deals | R${inv['volume_min']/1e6:.0f}-{inv['volume_max']/1e6:.0f}MM"):
                col1, col2, col3 = st.columns(3)
                col1.metric("Deals no Pipeline", inv["deals_pipeline"])
                col2.metric("Volume Min", _fmt_brl(inv["volume_min"]))
                col3.metric("Volume Max", _fmt_brl(inv["volume_max"]))
                st.markdown(f"**Perfil:** {inv['perfil']}")
                st.markdown(f"**Instrumentos:** {', '.join(inv['instrumentos'])}")
                st.markdown(f"**Setores:** {', '.join(inv['setores'])}")

    with tab_benchmarks:
        st.markdown("### Benchmarks Setoriais")
        st.caption("Referencias calibradas pelo mercado brasileiro de credito estruturado e Pipeline ZYN.")

        for setor_key in ["agro", "imobiliario", "industria", "fidc", "special_sits"]:
            bench = get_sector_benchmarks(setor_key)
            with st.expander(f"**{bench['setor']}**", expanded=False):
                col1, col2, col3, col4 = st.columns(4)
                if bench.get("margem_ebitda_media"):
                    col1.metric("Margem EBITDA", f"{bench['margem_ebitda_media']:.0%}")
                else:
                    col1.metric("Margem EBITDA", "N/A")
                if bench.get("divida_ebitda_media"):
                    col2.metric("Div/EBITDA", f"{bench['divida_ebitda_media']:.1f}x")
                else:
                    col2.metric("Div/EBITDA", "N/A")
                if bench.get("dscr_medio"):
                    col3.metric("DSCR", f"{bench['dscr_medio']:.1f}x")
                else:
                    col3.metric("DSCR", "N/A")
                if bench.get("ltv_maximo_recomendado"):
                    col4.metric("LTV Max", f"{bench['ltv_maximo_recomendado']:.0%}")
                else:
                    col4.metric("LTV Max", "N/A")

                st.markdown(f"**Volume mediano:** {_fmt_brl(bench['volume_mediano'])}")
                st.markdown(f"**Prazo tipico:** {bench['prazo_tipico_meses']} meses")
                st.markdown(f"**Instrumentos comuns:** {', '.join(bench.get('instrumentos_comuns', []))}")
                st.markdown(f"**Garantias tipicas:** {', '.join(bench.get('garantias_tipicas', []))}")

                if bench.get("riscos_especificos"):
                    st.markdown("**Riscos setoriais:**")
                    for r in bench["riscos_especificos"]:
                        st.markdown(f"- {r}")

                if bench.get("notas"):
                    st.info(bench["notas"])

        # Guarantee types reference
        st.markdown("---")
        st.markdown("### Taxonomia de Garantias ZYN")
        for tipo, info in GUARANTEE_TYPES.items():
            with st.expander(f"**{tipo}** — {info['descricao']}"):
                st.markdown("**Exemplos:**")
                for ex in info["exemplos"]:
                    st.markdown(f"- {ex}")
                st.markdown("**Documentos necessarios:**")
                for doc in info["docs_necessarios"]:
                    st.markdown(f"- {doc}")


# ---------------------------------------------------------------------------
# Page: Consulta Agro — Dados Fazenda
# ---------------------------------------------------------------------------
def page_consulta_agro():
    st.markdown(
        """
        <div style="margin-bottom:20px;">
            <h1 style="color:#223040; font-size:1.8rem; font-weight:800; margin:0;">
                🌱 Consulta Agro — Dados Fazenda
            </h1>
            <p style="color:#8B9197; font-size:0.95rem; margin:4px 0 0 0;">
                Consulta ambiental automatizada: CAR, NDVI, embargos IBAMA, sobreposições e SIGEF.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if not MODULES_AVAILABLE:
        st.error(f"Módulos não disponíveis: {_IMPORT_ERROR}")
        return

    # Input
    st.markdown("### Busca por CPF, CNPJ ou Código CAR")
    col_input, col_tipo = st.columns([3, 1])
    with col_input:
        busca_input = st.text_input(
            "CPF, CNPJ ou Código CAR",
            placeholder="Ex: 123.456.789-09 ou MT-5106240-3A41E6F0E6FF4444BB2C1C97EF8D08BA",
            key="agro_busca_input",
        )
    with col_tipo:
        tipo_busca = st.selectbox("Tipo", ["Auto-detectar", "CPF", "CNPJ", "CAR"], key="agro_tipo_busca")

    if st.button("🔍 Consultar", use_container_width=True, type="primary", key="agro_btn_consultar"):
        if not busca_input.strip():
            st.error("Informe um CPF, CNPJ ou código CAR.")
            return

        df_client = get_df_client()
        if not df_client:
            st.error(
                "**Dados Fazenda não configurado.** "
                "Adicione DADOS_FAZENDA_EMAIL e DADOS_FAZENDA_PASSWORD nos Secrets."
            )
            return

        valor = busca_input.strip()
        digits = __import__("re").sub(r"[^\dA-Za-z\-/]", "", valor)

        # Auto-detect type
        if tipo_busca == "Auto-detectar":
            digits_only = __import__("re").sub(r"[^\d]", "", valor)
            if len(digits_only) == 11:
                tipo_busca = "CPF"
            elif len(digits_only) == 14:
                tipo_busca = "CNPJ"
            elif "-" in valor and len(valor) > 20:
                tipo_busca = "CAR"
            else:
                st.error("Não foi possível detectar o tipo. Selecione manualmente.")
                return

        progress = st.progress(0, text="🔄 Autenticando no Dados Fazenda...")
        status = st.empty()
        results_container = st.container()

        try:
            # Step 1: Authenticate
            df_client._authenticate()
            progress.progress(0.15, text="🔄 Buscando propriedades...")

            if tipo_busca == "CAR":
                # Direct CAR lookup — base nacional completa
                car_code = valor.upper().strip()
                progress.progress(0.25, text=f"🌱 Consultando CAR na base nacional: {car_code[:40]}...")
                prop_result = df_client.consulta_car_aberta(car_code)
                resultado = {
                    "total_propriedades": 1,
                    "area_total_ha": 0,
                    "propriedades": [prop_result],
                    "alertas_consolidados": prop_result.get("alertas", []),
                    "score_ambiental_grupo": prop_result.get("score_ambiental", "N/D"),
                    "resumo": f"Consulta CAR {car_code}: Score {prop_result.get('score_ambiental', 'N/D')}",
                }
                cruzamento = {}
                progress.progress(1.0, text="✅ Consulta concluída!")
            else:
                # CPF/CNPJ — busca todas propriedades monitoradas na conta
                progress.progress(0.20, text=f"🔍 Buscando propriedades vinculadas ao {tipo_busca}...")
                all_props = df_client.get_properties()
                car_codes = [p.get("car_code", "") for p in all_props if p.get("car_code")]
                status.info(f"📋 {len(car_codes)} propriedade(s) na base Dados Fazenda. Consultando todas...")

                if not car_codes:
                    progress.empty()
                    st.warning("Nenhuma propriedade encontrada na conta.")
                    return

                # Consulta aberta para cada CAR
                propriedades = []
                alertas_todos = []
                for i, car in enumerate(car_codes):
                    pct = 0.25 + (0.65 * (i + 1) / len(car_codes))
                    progress.progress(pct, text=f"🌱 {i+1}/{len(car_codes)} — {car[:35]}...")
                    prop_result = df_client.consulta_car_aberta(car)
                    propriedades.append(prop_result)
                    alertas_todos.extend(prop_result.get("alertas", []))

                # Score consolidado
                scores = [p.get("score_ambiental", "Verde") for p in propriedades]
                if "Vermelho" in scores:
                    score_grupo = "Vermelho"
                elif "Amarelo" in scores:
                    score_grupo = "Amarelo"
                else:
                    score_grupo = "Verde"

                area_total = sum(
                    p.get("ndvi", {}).get("area_ha", 0) for p in propriedades
                )

                resultado = {
                    "total_propriedades": len(propriedades),
                    "area_total_ha": area_total,
                    "propriedades": propriedades,
                    "alertas_consolidados": alertas_todos,
                    "score_ambiental_grupo": score_grupo,
                    "resumo": f"Grupo com {len(propriedades)} propriedade(s) — Score: {score_grupo}",
                }
                cruzamento = df_client.cruzar_grupo_sigef(car_codes)
                progress.progress(1.0, text="✅ Consulta concluída!")
            status.empty()

            # Save to session for potential use in analysis
            st.session_state["agro_consulta"] = {
                "busca": valor,
                "tipo": tipo_busca,
                "resultado": resultado,
                "cruzamento": cruzamento,
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
            }

            # Persist to GitHub history
            try:
                _save_agro_consulta(st.session_state["agro_consulta"])
            except Exception:
                pass  # non-blocking — history save failure should not affect UX

        except Exception as e:
            progress.empty()
            status.empty()
            st.error(f"Erro na consulta: {e}")
            import traceback
            with st.expander("Detalhes do erro"):
                st.code(traceback.format_exc(), language=None)
            return

        # ── Display Results ──
        with results_container:
            st.markdown("---")

            # Header with score
            score = resultado.get("score_ambiental_grupo", "N/D")
            score_colors = {"Verde": "#2E7D4F", "Amarelo": "#EAB308", "Vermelho": "#DC2626"}
            score_color = score_colors.get(score, "#8B9197")

            col_s1, col_s2, col_s3, col_s4 = st.columns(4)
            col_s1.markdown(
                f'<div style="text-align:center; padding:12px; background:{score_color}; '
                f'border-radius:8px; color:white; font-size:1.3rem; font-weight:800;">'
                f'{score}</div>',
                unsafe_allow_html=True,
            )
            col_s1.caption("Score Ambiental")
            col_s2.metric("Propriedades", resultado.get("total_propriedades", 0))
            col_s3.metric("Área Total", f"{resultado.get('area_total_ha', 0):,.0f} ha")
            col_s4.metric("Alertas", len(resultado.get("alertas_consolidados", [])))

            # Alerts
            alertas = resultado.get("alertas_consolidados", [])
            if alertas:
                st.markdown("### ⚠️ Alertas")
                for alerta in alertas:
                    if "🔴" in alerta or "Vermelho" in alerta:
                        st.error(alerta)
                    elif "⚠" in alerta:
                        st.warning(alerta)
                    else:
                        st.success(alerta)

            # Properties detail
            propriedades = resultado.get("propriedades", [])
            if propriedades:
                st.markdown("### 🗺️ Propriedades Consultadas")
                for i, prop in enumerate(propriedades):
                    car = prop.get("car_code", "N/D")
                    score_prop = prop.get("score_ambiental", "N/D")
                    prop_color = score_colors.get(score_prop, "#8B9197")

                    with st.expander(
                        f"**{car[:50]}** — {score_prop}",
                        expanded=(i == 0),
                    ):
                        # NDVI
                        ndvi = prop.get("ndvi", {})
                        if ndvi:
                            st.markdown("**📈 NDVI (Vegetação)**")
                            cn1, cn2, cn3, cn4 = st.columns(4)
                            cn1.metric("NDVI Médio", f"{ndvi.get('ndvi_mean', 0):.3f}")
                            cn2.metric("NDVI Mediana", f"{ndvi.get('ndvi_median', 0):.3f}")
                            cn3.metric("Tendência", ndvi.get("tendencia", "N/D"))
                            cn4.metric("Cobertura Vegetal", f"{ndvi.get('cobertura_vegetal_pct', 0):.0f}%")

                        # Embargos
                        embargos = prop.get("embargos", {})
                        if embargos:
                            st.markdown("**🚫 Embargos**")
                            tem_embargo = embargos.get("tem_embargo", False)
                            if tem_embargo:
                                st.error(f"⛔ EMBARGO ATIVO — {embargos.get('detalhes', 'Ver detalhes')}")
                            else:
                                st.success("✅ Sem embargos ativos")

                        # Overlaps
                        sobrep = prop.get("sobreposicoes", {})
                        if sobrep:
                            st.markdown("**🗺️ Sobreposições Ambientais**")
                            items = [
                                ("Terra Indígena", sobrep.get("terras_indigenas")),
                                ("Quilombola", sobrep.get("quilombolas")),
                                ("Unidade Conservação", sobrep.get("unidades_conservacao")),
                                ("Assentamento", sobrep.get("assentamentos")),
                            ]
                            for label, data in items:
                                if data and data.get("tem_sobreposicao"):
                                    st.error(f"⛔ Sobreposição com {label} detectada")
                                elif data:
                                    st.success(f"✅ {label}: Sem sobreposição")

                        # INCRA
                        incra = prop.get("incra", {})
                        if incra and incra.get("certificado"):
                            st.success(f"✅ INCRA: Certificado — {incra.get('codigo', '')}")
                        elif incra:
                            st.warning("⚠ INCRA: Não certificado")

            # SIGEF Cross-reference
            if cruzamento:
                st.markdown("### 🔗 Cruzamento SIGEF")
                cob = cruzamento.get("cobertura_pct", 0)
                cc1, cc2, cc3 = st.columns(3)
                cc1.metric("Cobertura", f"{cob:.0f}%")
                cc2.metric("No Monitoramento", cruzamento.get("propriedades_cadastradas", 0))
                cc3.metric("Nos Documentos", cruzamento.get("propriedades_documentos", 0))

                alertas_sigef = cruzamento.get("alertas", [])
                for a in alertas_sigef:
                    if "🔴" in a:
                        st.error(a)
                    elif "⚠" in a:
                        st.warning(a)
                    else:
                        st.success(a)

                nao_cad = cruzamento.get("nao_cadastradas", [])
                if nao_cad:
                    st.markdown("**CARs não monitorados:**")
                    for nc in nao_cad:
                        st.markdown(f"- ⚠ `{nc['car_code']}`")

            # Summary for MAC
            resumo = resultado.get("resumo", "")
            if resumo:
                st.markdown("### 📝 Resumo para MAC")
                st.info(resumo)

            # Excel download button
            st.markdown("---")
            try:
                excel_bytes = generate_agro_excel(st.session_state["agro_consulta"])
                ts_file = datetime.now().strftime("%Y%m%d_%H%M")
                busca_clean = __import__("re").sub(r"[^\w\-]", "_", valor)
                st.download_button(
                    "📥 Baixar Relatório Excel",
                    data=excel_bytes,
                    file_name=f"ConsultaAgro_{busca_clean}_{ts_file}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as excel_err:
                st.warning(f"Erro ao gerar Excel: {excel_err}")

    # Show previous consultation if exists
    if "agro_consulta" in st.session_state and not st.session_state.get("_agro_btn_clicked"):
        prev = st.session_state["agro_consulta"]
        resultado = prev.get("resultado", {})
        cruzamento = prev.get("cruzamento", {})

        if resultado:
            st.markdown("---")
            st.caption(f"Última consulta: **{prev['busca']}** ({prev['tipo']}) — {prev['timestamp']}")

            score = resultado.get("score_ambiental_grupo", "N/D")
            score_colors = {"Verde": "#2E7D4F", "Amarelo": "#EAB308", "Vermelho": "#DC2626"}
            score_color = score_colors.get(score, "#8B9197")

            col_s1, col_s2, col_s3 = st.columns(3)
            col_s1.markdown(
                f'<div style="text-align:center; padding:8px; background:{score_color}; '
                f'border-radius:8px; color:white; font-weight:800;">{score}</div>',
                unsafe_allow_html=True,
            )
            col_s2.metric("Propriedades", resultado.get("total_propriedades", 0))
            col_s3.metric("Área Total", f"{resultado.get('area_total_ha', 0):,.0f} ha")

            alertas = resultado.get("alertas_consolidados", [])
            for a in alertas[:5]:
                if "🔴" in a:
                    st.error(a)
                elif "⚠" in a:
                    st.warning(a)
                else:
                    st.success(a)

    # ── History Section ──
    st.markdown("---")
    st.markdown("### 📋 Histórico de Consultas")

    agro_history = _load_agro_history()
    if agro_history:
        # Show most recent first, limit to 20
        recent = list(reversed(agro_history[-20:]))
        history_rows = []
        for h in recent:
            res = h.get("resultado", {})
            score = res.get("score_ambiental_grupo", "N/D")
            score_emoji = {"Verde": "🟢", "Amarelo": "🟡", "Vermelho": "🔴"}.get(score, "⚪")
            history_rows.append({
                "Data": h.get("timestamp", "N/D"),
                "Busca": h.get("busca", "N/D"),
                "Tipo": h.get("tipo", "N/D"),
                "Score": f"{score_emoji} {score}",
                "Propriedades": res.get("total_propriedades", 0),
                "Alertas": len(res.get("alertas_consolidados", [])),
            })

        st.dataframe(
            history_rows,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Data": st.column_config.TextColumn("Data", width="medium"),
                "Busca": st.column_config.TextColumn("Busca", width="large"),
                "Tipo": st.column_config.TextColumn("Tipo", width="small"),
                "Score": st.column_config.TextColumn("Score", width="medium"),
                "Propriedades": st.column_config.NumberColumn("Props", width="small"),
                "Alertas": st.column_config.NumberColumn("Alertas", width="small"),
            },
        )

        # Allow re-viewing a past consultation
        selected_idx = st.selectbox(
            "Recarregar consulta anterior",
            range(len(recent)),
            format_func=lambda i: f"{recent[i].get('timestamp', 'N/D')} — {recent[i].get('busca', 'N/D')} ({recent[i].get('tipo', '')})",
            key="agro_history_select",
        )
        if st.button("🔄 Carregar consulta selecionada", key="agro_load_history"):
            st.session_state["agro_consulta"] = recent[selected_idx]
            st.rerun()
    else:
        st.info("Nenhuma consulta anterior registrada.")


# ---------------------------------------------------------------------------
# Main Layout
# ---------------------------------------------------------------------------
def main():
    st.set_page_config(
        page_title="ZYN Credit Engine",
        page_icon="🏦",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

    # ── Authentication Gate ──
    if not _login_gate():
        return

    # Sidebar
    with st.sidebar:
        st.markdown(
            """
            <div style="padding:8px 0 4px 0;">
                <p style="color:#FFFFFF; font-size:1.5rem; font-weight:800; letter-spacing:-0.3px; margin:0;">
                    ZYN Capital
                </p>
                <p style="color:rgba(255,255,255,0.45); font-size:0.78rem; margin:2px 0 0 0; letter-spacing:0.5px;">
                    CRÉDITO ESTRUTURADO & M&A
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.markdown("---")

        st.markdown(
            '<p style="color:rgba(255,255,255,0.35); font-size:0.65rem; text-transform:uppercase; '
            'letter-spacing:1.5px; margin:0 0 4px 4px;">Navegação</p>',
            unsafe_allow_html=True,
        )

        page = st.radio(
            "Navegação",
            ["Dashboard", "Nova Análise", "Historico", "Checklist DD", "Investidores", "Consulta Agro"],
            label_visibility="collapsed",
        )

        st.markdown("---")

        # Quick stats in sidebar
        total = len(st.session_state.operacoes)
        em_andamento = _count_ops_by_status("Em Andamento")
        concluidas_sb = _count_ops_by_status("Concluída")
        st.markdown(
            f"""
            <div style="padding:4px 8px;">
                <p style="color:rgba(255,255,255,0.35); font-size:0.65rem; text-transform:uppercase;
                letter-spacing:1.5px; margin:0 0 8px 0;">Resumo</p>
                <div style="display:flex; justify-content:space-between; margin-bottom:6px;">
                    <span style="color:rgba(255,255,255,0.6); font-size:0.8rem;">Operações</span>
                    <span style="color:#FFFFFF; font-weight:700; font-size:0.8rem;">{total}</span>
                </div>
                <div style="display:flex; justify-content:space-between; margin-bottom:6px;">
                    <span style="color:rgba(255,255,255,0.6); font-size:0.8rem;">Concluídas</span>
                    <span style="color:#2E7D4F; font-weight:700; font-size:0.8rem;">{concluidas_sb}</span>
                </div>
                <div style="display:flex; justify-content:space-between;">
                    <span style="color:rgba(255,255,255,0.6); font-size:0.8rem;">Em andamento</span>
                    <span style="color:#EAB308; font-weight:700; font-size:0.8rem;">{em_andamento}</span>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.markdown("---")

        if st.session_state.current_op:
            if st.button("🔄 Limpar e Nova Análise", use_container_width=True, key="sidebar_clear"):
                _clear_session()
                st.rerun()

        st.markdown("---")

        # User info + logout
        user_name = st.session_state.get("user_name", "")
        user_email = st.session_state.get("user_email", "")
        if user_name:
            st.markdown(
                f"""<div style="padding:4px 8px;">
                    <p style="color:rgba(255,255,255,0.5); font-size:0.7rem; margin:0 0 2px 0;">Conectado como</p>
                    <p style="color:#FFFFFF; font-size:0.85rem; font-weight:600; margin:0;">{user_name}</p>
                    <p style="color:rgba(255,255,255,0.4); font-size:0.7rem; margin:2px 0 0 0;">{user_email}</p>
                </div>""",
                unsafe_allow_html=True,
            )
            if st.button("Sair", use_container_width=True, key="logout_btn"):
                for k in ["authenticated", "ops_authenticated", "user_email", "user_name"]:
                    st.session_state.pop(k, None)
                st.rerun()

        st.markdown(
            '<p class="footer-text">ZYN Credit Engine v2.0</p>',
            unsafe_allow_html=True,
        )

    # Page routing
    if page == "Dashboard":
        page_dashboard()
    elif page == "Nova Análise":
        page_nova_analise()
    elif page == "Historico":
        page_historico()
    elif page == "Checklist DD":
        page_checklist_dd()
    elif page == "Investidores":
        page_investidores()
    elif page == "Consulta Agro":
        page_consulta_agro()


if __name__ == "__main__":
    main()
